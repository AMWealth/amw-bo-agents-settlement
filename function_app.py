import logging
import os
import io

import re
import json
import base64
import hashlib
import zipfile
from datetime import datetime, timedelta, timezone, date
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

import azure.functions as func
import requests
import psycopg2
import psycopg2.extras
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pdfplumber

app = func.FunctionApp()

# =============================================================================
# ENV
# =============================================================================
TENANT_ID = os.environ.get("TENANT_ID", "").strip()
CLIENT_ID = os.environ.get("CLIENT_ID", "").strip()
CLIENT_SECRET = os.environ.get("CLIENT_SECRET", "").strip()

GRAPH_MAILBOX = os.environ.get("GRAPH_MAILBOX", "back.office@amwealth.ae").strip()
_mailbox2 = os.environ.get("GRAPH_MAILBOX_2", "").strip()
GRAPH_MAILBOXES = [m for m in [GRAPH_MAILBOX, _mailbox2] if m]
LOOKBACK_HOURS = int(os.environ.get("SETTLEMENT_LOOKBACK_HOURS", "72"))
PG_CONN_STRING = os.environ.get("PG_CONN_STRING", "").strip()

# Template codes that parse the email body (no attachment needed).
# All other templates require a PDF/Excel attachment.
EMAIL_BODY_TEMPLATES = frozenset({
    "STONEX_REPO_EMAIL",
    "GRANT_WESTOVER_REPO_EMAIL",
    "FAB_REPO_EMAIL",
})

GRAPH_SCOPE = "https://graph.microsoft.com/.default"
GRAPH_BASE = "https://graph.microsoft.com/v1.0"

REPORT_TO = os.environ.get("SETTLEMENT_REPORT_TO", "k.malkova@amwealth.ae").strip()
TEST_MODE = os.environ.get("TEST_MODE", "false").strip().lower() in ("1", "true", "yes")
TEST_EMAIL = "k.malkova@amwealth.ae"

# Optional fallback only
ALLOWED_SENDERS_FALLBACK = {
    "intl.email.confirms@instinet.com",
    "operations@gtnme.com",
    "donovan.landry@capitalunionbank.com",
    "charles.carroll@capitalunionbank.com",
    "emiliano.cabrera@capitalunionbank.com",
    "s.voll@camcapmarkets.com",
    "bo.tdsm@zarattinibank.ch",
    "backoffice@ashendenfinance.ch",
    "donotreplysecuritiesconfirmations@stonex.com",
    "bplmailer@bpl-bondpartners.ch",
    "emccarthy@seaportglobal.com",
    "eohara@seaportglobal.com",
    "settlement@bridport.ch",
    "opsseclendingrepo@stonex.com",
    "statements@stonex.com",
    # New senders added 2026-03-18
    "grant.westover@stonex.com",
    "amna.anwar@bankfab.com",
    "umar.malik@bankfab.com",
    "vijuraj.thandalath@bankfab.com",
    # Emirates NBD added 2026-03-24
    "validationstryops@emiratesnbd.com",
    "confirmationstryops@tanfeeth.ae",
    # FAB SWIFT MT545/MT547 settlement confirmations
    "noreply@bankfab.com",
    # ENBD Securities Order Confirmation Reports (GCM / Sincy Joji)
    "sincyjo@emiratesnbd.com",
}

TEST_SENDERS_DEFAULT = [
    "operations@gtnme.com",
    "donovan.landry@capitalunionbank.com",
    "charles.carroll@capitalunionbank.com",
    "emiliano.cabrera@capitalunionbank.com",
    "s.voll@camcapmarkets.com",
    "intl.email.confirms@instinet.com",
    "bo.tdsm@zarattinibank.ch",
    "backoffice@ashendenfinance.ch",
    "donotreplysecuritiesconfirmations@stonex.com",
    "bplmailer@bpl-bondpartners.ch",
    "statements@stonex.com",
    "emccarthy@seaportglobal.com",
    "eohara@seaportglobal.com",
    "settlement@bridport.ch",
    "opsseclendingrepo@stonex.com",
    "grant.westover@stonex.com",
    "amna.anwar@bankfab.com",
    "umar.malik@bankfab.com",
    "vijuraj.thandalath@bankfab.com",
    "validationstryops@emiratesnbd.com",
    "confirmationstryops@tanfeeth.ae",
]

# =============================================================================
# APP / DB
# =============================================================================
def get_conn():
    logging.warning("DEBUG DB MODE: using PG_CONN_STRING")
    if not PG_CONN_STRING:
        raise RuntimeError("Missing required env var: PG_CONN_STRING")

    conn_str = PG_CONN_STRING
    if "sslmode=" not in conn_str.lower():
        conn_str += " sslmode=require"

    return psycopg2.connect(conn_str, connect_timeout=30)


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def start_agent_run(conn, agent_name: str) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            insert into back_office_auto.agent_runs (agent_name, started_at, status, note)
            values (%s, now(), %s, %s)
            returning id
            """,
            (agent_name, "RUNNING", f"{agent_name} started"),
        )
        run_id = cur.fetchone()[0]
    conn.commit()
    return run_id


def finish_agent_run(conn, run_id: int, status: str, note: str):
    with conn.cursor() as cur:
        cur.execute(
            """
            update back_office_auto.agent_runs
               set finished_at = now(),
                   status = %s,
                   note = %s
             where id = %s
            """,
            (status, note, run_id),
        )
    conn.commit()


def email_already_processed(conn, internet_message_id: str) -> bool:
    with conn.cursor() as cur:
        cur.execute(
            """
            select 1
            from back_office_auto.settlement_emails
            where internet_message_id = %s
            limit 1
            """,
            (internet_message_id,),
        )
        return cur.fetchone() is not None


def clear_reconciliation_run_rows(conn, run_id: int):
    with conn.cursor() as cur:
        cur.execute(
            "delete from back_office_auto.settlement_reconciliation where run_id = %s",
            (run_id,),
        )
    conn.commit()


def load_mapping(conn) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            """
            select *
            from back_office_auto.counterparty_email_mapping
            where is_active = true
            """
        )
        for row in cur.fetchall():
            sender = (row["email_address_of_counterparty"] or "").strip().lower()
            if sender:
                out[sender] = dict(row)
    return out


def get_allowed_senders(mapping_by_sender: Dict[str, Dict[str, Any]]) -> set:
    allowed = set()

    for sender, row in mapping_by_sender.items():
        sender_key = (sender or "").strip().lower()
        if not sender_key:
            continue

        if row.get("is_active") is True:
            allowed.add(sender_key)

    if not allowed:
        allowed = set(ALLOWED_SENDERS_FALLBACK)

    return allowed


def is_sender_allowed(sender: str, allowed_senders: set) -> bool:
    """Return True if sender is in the exact allowlist OR matches a domain fallback."""
    key = (sender or "").strip().lower()
    return key in allowed_senders or _fallback_by_domain(key) is not None


# =============================================================================
# GRAPH
# =============================================================================
def get_graph_token() -> str:
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": GRAPH_SCOPE,
        "grant_type": "client_credentials",
    }
    r = requests.post(url, data=data, timeout=60)
    r.raise_for_status()
    return r.json()["access_token"]


def graph_headers(token: str) -> Dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def normalize_email_address(sender_obj: Dict[str, Any]) -> str:
    try:
        return (sender_obj.get("emailAddress", {}).get("address") or "").strip().lower()
    except Exception:
        return ""


def list_recent_messages(token: str, mailbox: str, since_dt: datetime) -> List[Dict[str, Any]]:
    since_iso = since_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    filter_str = f"receivedDateTime ge {since_iso}"

    url = (
        f"{GRAPH_BASE}/users/{mailbox}/mailFolders/Inbox/messages"
        f"?$top=100"
        f"&$select=id,internetMessageId,subject,receivedDateTime,from,hasAttachments,bodyPreview"
        f"&$orderby=receivedDateTime desc"
        f"&$filter={filter_str}"
    )

    results: List[Dict[str, Any]] = []
    while url:
        r = requests.get(url, headers=graph_headers(token), timeout=120)
        r.raise_for_status()
        data = r.json()
        results.extend(data.get("value", []))
        url = data.get("@odata.nextLink")

    return results


def list_recent_messages_by_sender_python_filter(
    token: str,
    mailbox: str,
    sender_email: str,
    since_dt: datetime,
    top_n: int = 3,
    require_attachments: bool = False,
) -> List[Dict[str, Any]]:
    sender_email = (sender_email or "").strip().lower()
    all_msgs = list_recent_messages(token, mailbox, since_dt)
    filtered = [
        m for m in all_msgs
        if normalize_email_address(m.get("from", {})) == sender_email
        and (not require_attachments or m.get("hasAttachments") is True)
    ]
    filtered.sort(key=lambda x: x.get("receivedDateTime") or "", reverse=True)
    return filtered[:top_n]


def get_message_attachments(token: str, mailbox: str, message_id: str) -> List[Dict[str, Any]]:
    url = f"{GRAPH_BASE}/users/{mailbox}/messages/{message_id}/attachments?$top=100"
    r = requests.get(url, headers=graph_headers(token), timeout=120)
    r.raise_for_status()
    data = r.json()
    return data.get("value", [])


def get_message_full(token: str, mailbox: str, message_id: str) -> Dict[str, Any]:
    url = (
        f"{GRAPH_BASE}/users/{mailbox}/messages/{message_id}"
        f"?$select=id,internetMessageId,subject,receivedDateTime,from,hasAttachments,body,bodyPreview"
    )
    r = requests.get(url, headers=graph_headers(token), timeout=120)
    r.raise_for_status()
    return r.json()


def get_attachment_content_bytes(
    token: str,
    mailbox: str,
    message_id: str,
    attachment_id: str,
) -> bytes:
    url = f"{GRAPH_BASE}/users/{mailbox}/messages/{message_id}/attachments/{attachment_id}"
    r = requests.get(url, headers=graph_headers(token), timeout=120)
    r.raise_for_status()
    data = r.json()

    content_b64 = data.get("contentBytes")
    if not content_b64:
        raise RuntimeError(f"Attachment {attachment_id} has no contentBytes")

    return base64.b64decode(content_b64)


# =============================================================================
# COMMON HELPERS
# =============================================================================
def clean_text(s: Optional[str]) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def clean_col(s: Any) -> str:
    s = clean_text(str(s or ""))
    s = re.sub(r"[^a-zA-Z0-9%]+", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_").lower()


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def infer_file_type(filename: str) -> str:
    f = filename.lower()
    if f.endswith(".xlsx"):
        return "xlsx"
    if f.endswith(".xlsm"):
        return "xlsm"
    if f.endswith(".xls"):
        return "xls"
    if f.endswith(".pdf"):
        return "pdf"
    if f.endswith(".zip"):
        return "zip"
    if f.endswith(".csv"):
        return "csv"
    if f.endswith(".msg"):
        return "msg"
    return "other"


def strip_html_tags(html: str) -> str:
    if not html:
        return ""
    html = re.sub(r"(?is)<(script|style).*?>.*?(</\1>)", " ", html)
    html = re.sub(r"(?i)<br\s*/?>", "\n", html)
    html = re.sub(r"(?i)</p>", "\n", html)
    html = re.sub(r"<[^>]+>", " ", html)
    html = html.replace("&nbsp;", " ")
    html = html.replace("&amp;", "&")
    html = html.replace("&lt;", "<")
    html = html.replace("&gt;", ">")
    return clean_text(html)


def parse_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except Exception:
            return None

    s = clean_text(str(value))
    if not s:
        return None

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    s = s.replace(",", "")
    s = s.replace("%", "")
    s = s.replace("'", "")          # Zarattini/Bridport: "200'000" → "200000"
    s = re.sub(r"^[A-Z]{3}\s+", "", s)
    # CUB OCR artifact: "1 15,628" (space inside number after currency strip)
    s = re.sub(r"^(\d)\s+(\d)", r"\1\2", s)
    s = re.sub(r"[^0-9.\-]", "", s).strip()

    if s in ("", "-", ".", "-."):
        return None

    try:
        num = Decimal(s)
        return -num if negative and num > 0 else num
    except InvalidOperation:
        return None


def parse_date_any(value: Any, prefer_day_first: bool = True) -> Optional[date]:
    if value is None:
        return None

    if hasattr(value, "date") and not isinstance(value, str):
        try:
            return value.date()
        except Exception:
            pass

    s = str(value).strip()
    if not s:
        return None

    s = re.sub(r"\s+", " ", s)

    # Strip trailing non-date garbage: "03-Mar-26 Last Execution Time" → "03-Mar-26"
    # but keep month names: "18 February 2025" stays intact
    s_trimmed = re.match(r"^(\d{1,4}[-/. ]\w{1,9}[-/. ]\d{2,4})", s)
    if s_trimmed:
        s = s_trimmed.group(1).strip()

    patterns_day_first = [
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d %b %Y",
        "%d %B %Y",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%d-%b-%y",     # CUB: "20-Feb-26"
        "%d/%m/%y",     # short year slash
        "%d.%m.%y",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
    ]

    patterns_month_first = [
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%m/%d/%y",
        "%b %d %Y",
        "%B %d %Y",
    ]

    patterns = patterns_day_first + patterns_month_first if prefer_day_first else patterns_month_first + patterns_day_first

    for p in patterns:
        try:
            return datetime.strptime(s, p).date()
        except Exception:
            continue

    for dayfirst in ([True, False] if prefer_day_first else [False, True]):
        try:
            dt = pd.to_datetime(s, errors="coerce", dayfirst=dayfirst)
            if pd.notna(dt):
                return dt.date()
        except Exception:
            continue

    return None


def parse_datetime_any(value: Any, prefer_day_first: bool = True) -> Optional[datetime]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    s = clean_text(str(value))
    if not s:
        return None

    patterns_day_first = [
        "%d/%m/%Y %I:%M:%S %p",
        "%d-%m-%Y %I:%M:%S %p",
        "%d.%m.%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%d %b %Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M:%S %p",
    ]

    patterns_month_first = [
        "%m/%d/%Y %I:%M:%S %p",
        "%m-%d-%Y %I:%M:%S %p",
        "%m/%d/%Y %H:%M:%S",
        "%m-%d-%Y %H:%M:%S",
        "%b %d %Y %I:%M:%S %p",
    ]

    patterns = patterns_day_first + patterns_month_first if prefer_day_first else patterns_month_first + patterns_day_first

    for p in patterns:
        try:
            return datetime.strptime(s, p)
        except Exception:
            continue

    for dayfirst in ([True, False] if prefer_day_first else [False, True]):
        try:
            dt = pd.to_datetime(s, errors="coerce", dayfirst=dayfirst)
            if pd.notna(dt):
                return dt.to_pydatetime()
        except Exception:
            continue

    return None


def normalize_side(value: Optional[str], template_code: Optional[str] = None) -> Optional[str]:
    if value is None:
        return None

    raw = str(value).strip()
    if not raw:
        return None

    v = re.sub(r"\s+", " ", raw).strip().upper()

    if v in {"B", "BUY", "BOT", "BOUGHT"}:
        return "BUY"

    if v in {"S", "SELL", "SLD", "SOLD"}:
        return "SELL"

    # Bondpartners: "Your purchase" = AM Wealth buys, "Your sale" = AM Wealth sells
    if "YOUR PURCHASE" in v:
        return "BUY"
    if "YOUR SALE" in v:
        return "SELL"

    if "WE SOLD TO YOU" in v:
        return "BUY"

    if "WE BOUGHT FROM YOU" in v:
        return "SELL"

    # Ashenden: "YOU SOLD IN USD" / "YOU BOUGHT IN USD"
    if "YOU SOLD" in v:
        return "SELL"

    if "YOU BOUGHT" in v:
        return "BUY"

    # StoneX FI: "We confirm your BUY/SELL transaction"
    if "CONFIRM YOUR BUY" in v:
        return "BUY"

    if "CONFIRM YOUR SELL" in v:
        return "SELL"

    # FAB REPO: "AM Wealth enters Reverse Repo (lends cash/ borrows securities)"
    # Reverse Repo = AM Wealth lends cash, borrows securities → BUY side (receives securities)
    if "REVERSE REPO" in v or "LENDS CASH" in v:
        return "BUY"

    if v == "BUYS":
        return "BUY"

    if v == "SELLS":
        return "SELL"

    if "BUY" in v and "SELL" not in v:
        return "BUY"

    if "SELL" in v and "BUY" not in v:
        return "SELL"

    if "LENDS COLLATERAL" in v:
        return "SELL"

    if "BORROWS COLLATERAL" in v:
        return "BUY"

    return None


def validate_side(side: Optional[str]) -> List[str]:
    if side not in {"BUY", "SELL"}:
        return [f"invalid_side:{side}"]
    return []


def validate_trade_dates(trade_date: Optional[date], value_date: Optional[date]) -> List[str]:
    issues = []
    today = now_utc().date()

    if trade_date:
        if trade_date < today - timedelta(days=30):
            issues.append(f"trade_date_too_old:{trade_date}")
        if trade_date > today + timedelta(days=2):
            issues.append(f"trade_date_in_future:{trade_date}")

    if trade_date and value_date:
        if value_date < trade_date - timedelta(days=2):
            issues.append(f"value_date_before_trade_date:{value_date}")
        if value_date > trade_date + timedelta(days=30):
            issues.append(f"value_date_too_far:{value_date}")

    return issues


def validate_trade_date_vs_email(trade_date: Optional[date], email_received_at: Optional[datetime]) -> List[str]:
    if not trade_date or not email_received_at:
        return []

    email_date = email_received_at.date()
    delta = abs((trade_date - email_date).days)

    if delta > 14:
        return [f"trade_date_far_from_email_date:{trade_date}_vs_{email_date}"]
    return []


def pick_first(row: Dict[str, Any], keys: List[str]) -> Any:
    for k in keys:
        if k in row and row[k] not in (None, "", "nan"):
            return row[k]
    return None


def rx(pattern: str, text: str, flags: int = re.IGNORECASE | re.MULTILINE) -> Optional[str]:
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else None


def finalize_trade_validation(trade: Dict[str, Any], email_received_at: Optional[datetime]):
    issues = []

    if not trade.get("isin"):
        issues.append("missing_isin")

    issues.extend(validate_side(trade.get("side")))
    issues.extend(validate_trade_dates(trade.get("trade_date"), trade.get("value_date")))
    issues.extend(validate_trade_date_vs_email(trade.get("trade_date"), email_received_at))

    if not trade.get("trade_date"):
        issues.append("missing_trade_date")

    if not trade.get("counterparty_reference"):
        issues.append("missing_counterparty_reference")

    if issues:
        trade["validation_status"] = "NEEDS_REVIEW"
        trade["validation_note"] = ", ".join(issues)
    else:
        trade["validation_status"] = "PARSED"
        trade["validation_note"] = None


# Domain-based fallback: any sender @capitalunionbank.com → CUB_PDF, etc.
SENDER_DOMAIN_FALLBACK: Dict[str, Dict[str, str]] = {
    "capitalunionbank.com": {"template_code": "CUB_PDF", "broker_name": "Capital Union Bank Ltd."},
    "emiratesnbd.com": {"template_code": "ENBD_PDF", "broker_name": "Emirates NBD"},
    "tanfeeth.ae": {"template_code": "ENBD_PDF", "broker_name": "Emirates NBD"},
    "bankfab.com": {"template_code": "FAB_SWIFT_PDF", "broker_name": "First Abu Dhabi Bank PJSC"},
}


def _fallback_by_domain(sender: str) -> Optional[Dict[str, str]]:
    """Return domain fallback dict if sender domain matches, else None."""
    domain = sender.split("@", 1)[-1].lower() if "@" in sender else ""
    return SENDER_DOMAIN_FALLBACK.get(domain)


def resolve_broker_name_from_mapping(sender: str, mapping_by_sender: Dict[str, Dict[str, Any]]) -> str:
    sender_key = (sender or "").strip().lower()
    row = mapping_by_sender.get(sender_key)

    if not row:
        fb = _fallback_by_domain(sender_key)
        if fb:
            return fb["broker_name"]
        return sender_key

    alias_name = clean_text(row.get("counterparty_alias"))
    counterparty_name = clean_text(row.get("counterparty"))

    if alias_name:
        return alias_name
    if counterparty_name:
        return counterparty_name
    return sender_key


def detect_template_from_mapping(sender: str, mapping_by_sender: Dict[str, Dict[str, Any]]) -> Optional[str]:
    sender_key = (sender or "").strip().lower()
    row = mapping_by_sender.get(sender_key)
    if not row:
        fb = _fallback_by_domain(sender_key)
        if fb:
            return fb["template_code"]
        return None
    return clean_text(row.get("template_code")) or None


def trade_dedup_key(trade: Dict[str, Any]) -> str:
    return "|".join([
        clean_text(str(trade.get("internet_message_id") or "")),
        clean_text(str(trade.get("isin") or "")),
        clean_text(str(trade.get("side") or "")),
        clean_text(str(trade.get("trade_date") or "")),
        clean_text(str(trade.get("value_date") or "")),
        clean_text(str(trade.get("quantity") or trade.get("nominal") or "")),
        clean_text(str(trade.get("price") or trade.get("price_in_percentage") or "")),
    ])


def build_generic_reference(
    isin: Optional[str],
    side: Optional[str],
    trade_date: Optional[date],
    value_date: Optional[date],
    qty: Optional[Decimal],
    price: Optional[Decimal],
    nominal: Optional[Decimal] = None,
) -> str:
    return "|".join([
        clean_text(str(isin or "")),
        clean_text(str(side or "")),
        clean_text(str(trade_date or "")),
        clean_text(str(value_date or "")),
        clean_text(str(qty if qty is not None else nominal if nominal is not None else "")),
        clean_text(str(price if price is not None else "")),
    ])


def normalize_trade_signs(trade: Dict[str, Any]) -> Dict[str, Any]:
    for field in [
        "quantity",
        "price",
        "consideration",
        "commission",
        "net_amount",
        "nominal",
        "price_in_percentage",
        "accrued_interest",
    ]:
        val = trade.get(field)
        dec = parse_decimal(val)
        if dec is not None:
            trade[field] = abs(dec)
    return trade


# =============================================================================
# INSERT HELPERS
# =============================================================================
def insert_settlement_email(
    conn,
    internet_message_id: str,
    message_id: str,
    sender: str,
    subject: str,
    received_at: Optional[datetime],
    status: str,
    note: str,
    mailbox: str,
    attachment_count: int,
    parsed_trade_count: int,
    processing_run_id: Optional[int],
) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            insert into back_office_auto.settlement_emails
            (
                internet_message_id,
                message_id,
                sender,
                subject,
                received_at,
                processed_at,
                status,
                note,
                mailbox,
                attachment_count,
                parsed_trade_count,
                processing_run_id
            )
            values (%s, %s, %s, %s, %s, now(), %s, %s, %s, %s, %s, %s)
            on conflict (internet_message_id) do update
               set processed_at = now(),
                   status = excluded.status,
                   note = excluded.note,
                   mailbox = excluded.mailbox,
                   attachment_count = excluded.attachment_count,
                   parsed_trade_count = excluded.parsed_trade_count,
                   processing_run_id = excluded.processing_run_id
            returning id
            """,
            (
                internet_message_id,
                message_id,
                sender,
                subject,
                received_at,
                status,
                note,
                mailbox,
                attachment_count,
                parsed_trade_count,
                processing_run_id,
            ),
        )
        row = cur.fetchone()
    conn.commit()
    return row[0]


def insert_settlement_file(
    conn,
    internet_message_id: str,
    file_name: str,
    file_type: str,
    file_hash: str,
    attachment_size: int,
    attachment_order: Optional[int],
    parent_zip_file_name: Optional[str],
    parse_status: str,
    parse_note: Optional[str],
) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            insert into back_office_auto.settlement_files
            (
                internet_message_id,
                file_name,
                file_type,
                file_hash,
                created_at,
                attachment_size,
                attachment_order,
                parent_zip_file_name,
                parse_status,
                parse_note
            )
            values (%s, %s, %s, %s, now(), %s, %s, %s, %s, %s)
            on conflict do nothing
            returning id
            """,
            (
                internet_message_id,
                file_name,
                file_type,
                file_hash,
                attachment_size,
                attachment_order,
                parent_zip_file_name,
                parse_status,
                parse_note,
            ),
        )
        row = cur.fetchone()

    conn.commit()

    if row:
        return row[0]

    with conn.cursor() as cur:
        cur.execute(
            """
            select id
            from back_office_auto.settlement_files
            where internet_message_id = %s
              and file_hash = %s
            order by id desc
            limit 1
            """,
            (internet_message_id, file_hash),
        )
        existing = cur.fetchone()
    conn.commit()
    return existing[0]


def upsert_settlement_trade(conn, trade: Dict[str, Any]) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            insert into back_office_auto.settlement_trades
            (
                internet_message_id,
                source_file,
                source_type,
                broker_name,
                account_number,
                security_name,
                isin,
                sedol,
                side,
                trade_date,
                value_date,
                quantity,
                price,
                price_currency,
                consideration,
                commission,
                stamp_duty,
                transaction_levy,
                trading_fee,
                afrc_fee,
                net_amount,
                settlement_terms,
                counterparty_reference,
                validation_status,
                validation_note,
                created_at,
                nominal,
                price_in_percentage,
                accrued_interest,
                settlement_currency,
                parser_template,
                raw_json,
                matched_by,
                processing_run_id,
                file_id,
                email_id,
                side_original_text,
                trade_date_original_text,
                value_date_original_text,
                our_ssi
            )
            values
            (
                %(internet_message_id)s,
                %(source_file)s,
                %(source_type)s,
                %(broker_name)s,
                %(account_number)s,
                %(security_name)s,
                %(isin)s,
                %(sedol)s,
                %(side)s,
                %(trade_date)s,
                %(value_date)s,
                %(quantity)s,
                %(price)s,
                %(price_currency)s,
                %(consideration)s,
                %(commission)s,
                %(stamp_duty)s,
                %(transaction_levy)s,
                %(trading_fee)s,
                %(afrc_fee)s,
                %(net_amount)s,
                %(settlement_terms)s,
                %(counterparty_reference)s,
                %(validation_status)s,
                %(validation_note)s,
                now(),
                %(nominal)s,
                %(price_in_percentage)s,
                %(accrued_interest)s,
                %(settlement_currency)s,
                %(parser_template)s,
                %(raw_json)s,
                %(matched_by)s,
                %(processing_run_id)s,
                %(file_id)s,
                %(email_id)s,
                %(side_original_text)s,
                %(trade_date_original_text)s,
                %(value_date_original_text)s,
                %(our_ssi)s
            )
            on conflict (internet_message_id, counterparty_reference)
            do update set
                source_file = excluded.source_file,
                source_type = excluded.source_type,
                broker_name = excluded.broker_name,
                account_number = excluded.account_number,
                security_name = excluded.security_name,
                isin = excluded.isin,
                sedol = excluded.sedol,
                side = excluded.side,
                trade_date = excluded.trade_date,
                value_date = excluded.value_date,
                quantity = excluded.quantity,
                price = excluded.price,
                price_currency = excluded.price_currency,
                consideration = excluded.consideration,
                commission = excluded.commission,
                stamp_duty = excluded.stamp_duty,
                transaction_levy = excluded.transaction_levy,
                trading_fee = excluded.trading_fee,
                afrc_fee = excluded.afrc_fee,
                net_amount = excluded.net_amount,
                settlement_terms = excluded.settlement_terms,
                validation_status = excluded.validation_status,
                validation_note = excluded.validation_note,
                nominal = excluded.nominal,
                price_in_percentage = excluded.price_in_percentage,
                accrued_interest = excluded.accrued_interest,
                settlement_currency = excluded.settlement_currency,
                parser_template = excluded.parser_template,
                raw_json = excluded.raw_json,
                matched_by = excluded.matched_by,
                processing_run_id = excluded.processing_run_id,
                file_id = excluded.file_id,
                email_id = excluded.email_id,
                side_original_text = excluded.side_original_text,
                trade_date_original_text = excluded.trade_date_original_text,
                value_date_original_text = excluded.value_date_original_text,
                our_ssi = COALESCE(excluded.our_ssi, settlement_trades.our_ssi)
            returning id
            """,
            trade,
        )
        row = cur.fetchone()
    conn.commit()
    return row[0]


# =============================================================================
# EXCEL HELPERS / PARSERS
# =============================================================================
def extract_excel_sheets(file_bytes: bytes, filename: str) -> List[pd.DataFrame]:
    dataframes: List[pd.DataFrame] = []
    lower_name = filename.lower()

    if lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm"):
        try:
            xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
            for sheet_name in xls.sheet_names:
                try:
                    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, engine="openpyxl")
                    if df is not None and not df.empty:
                        dataframes.append(df)
                except Exception as e:
                    logging.warning("Failed reading sheet %s in %s: %s", sheet_name, filename, e)
            return dataframes
        except Exception as e:
            logging.exception("Unable to read xlsx/xlsm file %s: %s", filename, e)
            return dataframes

    if lower_name.endswith(".xls"):
        try:
            xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="xlrd")
            for sheet_name in xls.sheet_names:
                try:
                    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, engine="xlrd")
                    if df is not None and not df.empty:
                        dataframes.append(df)
                except Exception as e:
                    logging.warning("Failed reading xls sheet %s in %s: %s", sheet_name, filename, e)
            return dataframes
        except Exception as e:
            logging.exception("Unable to read xls file %s: %s", filename, e)
            return dataframes

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        for ws in wb.worksheets:
            rows = list(ws.values)
            if not rows:
                continue
            header = [clean_text(x) for x in rows[0]]
            body = rows[1:]
            df = pd.DataFrame(body, columns=header)
            if not df.empty:
                dataframes.append(df)
    except Exception as e:
        logging.exception("Fallback excel reader failed for %s: %s", filename, e)

    return dataframes


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [clean_col(c) for c in df.columns]
    return df


def find_gtn_header_row(df: pd.DataFrame) -> Optional[int]:
    for idx in range(len(df)):
        row_values = [clean_text(str(x)).lower() for x in df.iloc[idx].tolist() if x is not None]
        row_text = " | ".join(row_values)

        if (
            "symbol" in row_text
            and "side" in row_text
            and "quantity" in row_text
            and ("stl.date" in row_text or "stl_date" in row_text or "stl date" in row_text)
            and ("isin code" in row_text or "isin_code" in row_text or "isin" in row_text)
        ):
            return idx
    return None


def rebuild_gtn_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    raw = df.copy().reset_index(drop=True)

    header_row_idx = find_gtn_header_row(raw)
    if header_row_idx is None:
        logging.warning("GTN header row not found")
        return pd.DataFrame()

    header = [clean_col(x) for x in raw.iloc[header_row_idx].tolist()]
    data = raw.iloc[header_row_idx + 1 :].copy()
    data.columns = header
    data = data.reset_index(drop=True)

    data = data.loc[:, [c for c in data.columns if c and not str(c).startswith("unnamed")]]
    data = data.dropna(how="all")
    return data


def extract_isin_from_text(text: str) -> Optional[str]:
    m = re.search(r"\b[A-Z]{2}[A-Z0-9]{9,12}\b", text)
    return m.group(0) if m else None


def extract_date_from_text(text: str) -> Optional[str]:
    m = re.search(r"\b\d{2}/\d{2}/\d{4}\b", text)
    return m.group(0) if m else None


def extract_datetime_from_text(text: str) -> Optional[str]:
    m = re.search(r"\b\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}(?:\s*[ap]m)?\b", text, re.IGNORECASE)
    return m.group(0) if m else None


def parse_gtn_dataframe_rows(df: pd.DataFrame) -> List[Dict[str, Any]]:
    table = rebuild_gtn_dataframe(df)
    if table.empty:
        return []

    rows: List[Dict[str, Any]] = []

    for _, row in table.iterrows():
        rowd = {str(k): (None if pd.isna(v) else v) for k, v in row.to_dict().items()}

        row_text = " ".join([clean_text(str(v)) for v in rowd.values() if v is not None]).strip()
        row_text_upper = row_text.upper()

        if not row_text:
            continue

        if (
            "SUB TOTAL" in row_text_upper
            or "GRAND TOTAL" in row_text_upper
            or "SETTLEMENT INSTRUCTIONS" in row_text_upper
            or "THIS TRADE CONFIRMATION IS PROVIDED TO YOU" in row_text_upper
            or "SYMBOL SIDE QUANTITY PRICE" in row_text_upper
        ):
            continue

        symbol = pick_first(rowd, ["symbol"])
        side = pick_first(rowd, ["side"])
        quantity = pick_first(rowd, ["quantity"])
        price = pick_first(rowd, ["price"])
        gross = pick_first(rowd, ["gross"])
        brok_com = pick_first(rowd, ["brok_com", "brok_com.", "broker_commission", "commission"])
        net_settle = pick_first(rowd, ["net_settle", "net_amount"])
        tr_date = pick_first(rowd, ["tr_date", "trade_date"])
        stl_date = pick_first(rowd, ["stl_date", "settlement_date", "value_date"])
        isin_code = pick_first(rowd, ["isin_code", "isin"])

        if not symbol:
            continue

        if not side:
            if " BUY " in f" {row_text_upper} ":
                side = "Buy"
            elif " SELL " in f" {row_text_upper} ":
                side = "Sell"

        if not isin_code:
            isin_code = extract_isin_from_text(row_text)

        if not tr_date:
            tr_date = extract_datetime_from_text(row_text)

        if not stl_date:
            stl_date = extract_date_from_text(row_text)

        if not side or not isin_code or not tr_date:
            continue

        rows.append({
            "symbol": symbol,
            "side_raw": side,
            "quantity": quantity,
            "price": price,
            "gross": gross,
            "commission": brok_com,
            "net_amount": net_settle,
            "trade_datetime_raw": tr_date,
            "settlement_date_raw": stl_date,
            "isin": isin_code,
            "raw_row": rowd,
        })

    logging.info("GTN rebuilt dataframe parsed rows=%s", len(rows))
    return rows


def parse_gtn_excel(
    df: pd.DataFrame,
    internet_message_id: str,
    source_file: str,
    sender: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    mapping_by_sender: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    parsed_rows = parse_gtn_dataframe_rows(df)

    for row in parsed_rows:
        trade_dt = parse_datetime_any(row["trade_datetime_raw"], prefer_day_first=False)
        trade_date = trade_dt.date() if trade_dt else parse_date_any(row["trade_datetime_raw"], prefer_day_first=False)
        value_date = parse_date_any(row["settlement_date_raw"], prefer_day_first=False)

        trade = {
            "internet_message_id": internet_message_id,
            "source_file": source_file,
            "source_type": "excel",
            "broker_name": resolve_broker_name_from_mapping(sender, mapping_by_sender),
            "account_number": None,
            "security_name": clean_text(row["symbol"]),
            "isin": clean_text(row["isin"]),
            "sedol": None,
            "side": normalize_side(row["side_raw"], "GTN_XLS_PDF"),
            "trade_date": trade_date,
            "value_date": value_date,
            "quantity": parse_decimal(row["quantity"]),
            "price": parse_decimal(row["price"]),
            "price_currency": "USD",
            "consideration": parse_decimal(row["gross"]),
            "commission": parse_decimal(row["commission"]),
            "stamp_duty": None,
            "transaction_levy": None,
            "trading_fee": None,
            "afrc_fee": None,
            "net_amount": parse_decimal(row["net_amount"]),
            "settlement_terms": None,
            "counterparty_reference": None,
            "nominal": None,
            "price_in_percentage": None,
            "accrued_interest": None,
            "settlement_currency": "USD",
            "parser_template": "GTN_XLS_PDF",
            "raw_json": json.dumps(row["raw_row"], default=str),
            "matched_by": None,
            "processing_run_id": processing_run_id,
            "file_id": file_id,
            "email_id": email_id,
            "side_original_text": clean_text(str(row["side_raw"])) if row["side_raw"] is not None else None,
            "trade_date_original_text": clean_text(str(row["trade_datetime_raw"])) if row["trade_datetime_raw"] is not None else None,
            "value_date_original_text": clean_text(str(row["settlement_date_raw"])) if row["settlement_date_raw"] is not None else None,
            "validation_status": None,
            "validation_note": None,
            "our_ssi": None,
        }
        trade = normalize_trade_signs(trade)
        trade["counterparty_reference"] = build_generic_reference(
            trade["isin"], trade["side"], trade["trade_date"], trade["value_date"], trade["quantity"], trade["price"]
        )
        finalize_trade_validation(trade, email_received_at)
        out.append(trade)

    return out


def parse_instinet_excel(
    df: pd.DataFrame,
    internet_message_id: str,
    source_file: str,
    sender: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    mapping_by_sender: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    df = normalize_dataframe(df)
    df = df.dropna(how="all")
    if df.empty:
        return out

    for _, row in df.iterrows():
        rowd = {str(k): (None if pd.isna(v) else v) for k, v in row.to_dict().items()}

        isin = pick_first(rowd, ["isin", "isin_code", "security_nb", "security_number"])
        security_name = pick_first(rowd, ["security_name", "sec_descr", "description", "symbol_name", "security", "symbol"])
        side_raw = pick_first(rowd, ["side", "buy_sell", "b_s"])
        trade_date_raw = pick_first(rowd, ["trade_date", "tr_date", "tradedate"])
        value_date_raw = pick_first(rowd, ["settl_date", "settle_date", "settlement_date", "value_date", "stl_date"])
        quantity = pick_first(rowd, ["quantity", "qty", "units"])
        price = pick_first(rowd, ["price", "trade_price"])
        price_currency = pick_first(rowd, ["currency", "tr_currency", "settl_currency", "price_currency"])
        consideration = pick_first(rowd, ["consideration", "gross", "principal_amount", "amount"])
        commission = pick_first(rowd, ["commission", "brok_com", "brokerage", "broker_commission"])
        net_amount = pick_first(rowd, ["net", "net_settle", "net_amount", "total_net"])
        counterparty_reference = pick_first(rowd, ["trans", "trans_no", "transaction_no", "counterparty_reference", "reference", "ref"])

        if not counterparty_reference:
            counterparty_reference = "|".join([
                clean_text(str(isin or "")),
                clean_text(str(normalize_side(side_raw, "INSTINET_XLSM") or "")),
                clean_text(str(parse_date_any(trade_date_raw, prefer_day_first=True) or "")),
                clean_text(str(parse_decimal(quantity) or parse_decimal(consideration) or "")),
            ])

        if not isin and not security_name:
            continue

        trade = {
            "internet_message_id": internet_message_id,
            "source_file": source_file,
            "source_type": "excel",
            "broker_name": resolve_broker_name_from_mapping(sender, mapping_by_sender),
            "account_number": None,
            "security_name": clean_text(security_name),
            "isin": clean_text(isin),
            "sedol": None,
            "side": normalize_side(side_raw, "INSTINET_XLSM"),
            "trade_date": parse_date_any(trade_date_raw, prefer_day_first=True),
            "value_date": parse_date_any(value_date_raw, prefer_day_first=True),
            "quantity": parse_decimal(quantity),
            "price": parse_decimal(price),
            "price_currency": clean_text(price_currency),
            "consideration": parse_decimal(consideration),
            "commission": parse_decimal(commission),
            "stamp_duty": None,
            "transaction_levy": None,
            "trading_fee": None,
            "afrc_fee": None,
            "net_amount": parse_decimal(net_amount),
            "settlement_terms": None,
            "counterparty_reference": clean_text(counterparty_reference),
            "nominal": None,
            "price_in_percentage": None,
            "accrued_interest": None,
            "settlement_currency": clean_text(price_currency),
            "parser_template": "INSTINET_XLSM",
            "raw_json": json.dumps(rowd, default=str),
            "matched_by": None,
            "processing_run_id": processing_run_id,
            "file_id": file_id,
            "email_id": email_id,
            "side_original_text": clean_text(str(side_raw)) if side_raw is not None else None,
            "trade_date_original_text": clean_text(str(trade_date_raw)) if trade_date_raw is not None else None,
            "value_date_original_text": clean_text(str(value_date_raw)) if value_date_raw is not None else None,
            "validation_status": None,
            "validation_note": None,
            "our_ssi": None,
        }

        if not trade["trade_date"] and trade.get("isin") and trade.get("quantity") and trade.get("price"):
            continue

        trade = normalize_trade_signs(trade)
        finalize_trade_validation(trade, email_received_at)
        out.append(trade)

    return out


# =============================================================================
# PDF TEXT HELPERS
# =============================================================================
def extract_pdf_text(file_bytes: bytes) -> str:
    texts: List[str] = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                try:
                    texts.append(page.extract_text() or "")
                except Exception:
                    continue
    except Exception as e:
        logging.warning("extract_pdf_text: could not open PDF (%s), returning empty text", e)
        return ""
    return "\n".join(texts)


def build_trade_dict(
    *,
    internet_message_id: str,
    source_file: str,
    source_type: str,
    broker_name: str,
    security_name: Optional[str],
    isin: Optional[str],
    side: Optional[str],
    trade_date: Optional[date],
    value_date: Optional[date],
    quantity: Optional[Decimal],
    price: Optional[Decimal],
    price_currency: Optional[str],
    consideration: Optional[Decimal],
    commission: Optional[Decimal],
    net_amount: Optional[Decimal],
    settlement_terms: Optional[str],
    counterparty_reference: Optional[str],
    nominal: Optional[Decimal],
    price_in_percentage: Optional[Decimal],
    accrued_interest: Optional[Decimal],
    settlement_currency: Optional[str],
    parser_template: str,
    raw_json: str,
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    side_original_text: Optional[str],
    trade_date_original_text: Optional[str],
    value_date_original_text: Optional[str],
) -> Dict[str, Any]:
    return {
        "internet_message_id": internet_message_id,
        "source_file": source_file,
        "source_type": source_type,
        "broker_name": broker_name,
        "account_number": None,
        "security_name": clean_text(security_name),
        "isin": clean_text(isin),
        "sedol": None,
        "side": side,
        "trade_date": trade_date,
        "value_date": value_date,
        "quantity": quantity,
        "price": price,
        "price_currency": clean_text(price_currency),
        "consideration": consideration,
        "commission": commission,
        "stamp_duty": None,
        "transaction_levy": None,
        "trading_fee": None,
        "afrc_fee": None,
        "net_amount": net_amount,
        "settlement_terms": settlement_terms,
        "counterparty_reference": clean_text(counterparty_reference),
        "validation_status": None,
        "validation_note": None,
        "nominal": nominal,
        "price_in_percentage": price_in_percentage,
        "accrued_interest": accrued_interest,
        "settlement_currency": clean_text(settlement_currency),
        "parser_template": parser_template,
        "raw_json": raw_json,
        "matched_by": None,
        "processing_run_id": processing_run_id,
        "file_id": file_id,
        "email_id": email_id,
        "side_original_text": clean_text(side_original_text),
        "trade_date_original_text": clean_text(trade_date_original_text),
        "value_date_original_text": clean_text(value_date_original_text),
        "our_ssi": None,  # populated later by enrich_cpty_ssi()
    }


def parse_bond_style_pdf_common(
    *,
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
    parser_template: str,
) -> List[Dict[str, Any]]:
    isin = (
        rx(r"ISIN(?:\s+Code)?\s*[:\-]?\s*([A-Z0-9]{10,20})", text)
        or rx(r"Security\s+Nb\.?\s*[:\-]?\s*([A-Z0-9]{10,20})", text)
    )
    security_name = (
        rx(r"Description\s*[:\-]?\s*(.+)", text)
        or rx(r"Security\s+Name\s*[:\-]?\s*(.+)", text)
    )

    direction_phrase = (
        rx(r"(We\s+(?:SOLD|BOUGHT)\s+(?:to|from)\s+you)", text)
        or rx(r"\b(SELLS|BUYS|BUY|SELL)\b", text)
    )

    trade_date_raw = rx(r"Trade\s+Date\s*[:\-]?\s*([0-9A-Za-z\-/\. ]+)", text)
    value_date_raw = rx(r"(?:Settlement|Value)\s+Date\s*[:\-]?\s*([0-9A-Za-z\-/\. ]+)", text)

    nominal_raw = rx(r"Nominal\s+Amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,.\-]+)", text)
    qty_raw = rx(r"Quantity\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,.\-]+)", text)
    price_pct_raw = rx(r"Price\s*[:\-]?\s*([0-9,.\-]+)\s*%", text)
    price_raw = rx(r"Price\s*[:\-]?\s*([0-9,.\-]+)", text)
    principal_amount = rx(r"Principal\s+Amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,.\-]+)", text)
    accrued = rx(r"Accrued\s+Interest\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,.\-]+)", text)
    total_cash = (
        rx(r"Total\s+Cash\s+Settlement\s+Amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,.\-]+)", text)
        or rx(r"Net\s+amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,.\-]+)", text)
    )
    ccy = (
        rx(r"Currency\s*[:\-]?\s*([A-Z]{3})", text)
        or rx(r"\b(USD|EUR|AED|GBP|CHF|HKD)\b", text)
    )
    ref = rx(r"(?:Reference|Confirmation\s+No\.?|Trade\s+Reference)\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)

    side = normalize_side(direction_phrase, parser_template)

    quantity = parse_decimal(qty_raw)
    nominal = parse_decimal(nominal_raw)

    consideration = parse_decimal(principal_amount) or parse_decimal(total_cash)
    price_pct = parse_decimal(price_pct_raw)
    price = parse_decimal(price_raw)

    if not ref:
        ref = build_generic_reference(
            isin=isin,
            side=side,
            trade_date=parse_date_any(trade_date_raw, True),
            value_date=parse_date_any(value_date_raw, True),
            qty=quantity,
            price=price if price is not None else price_pct,
            nominal=nominal,
        )

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=parse_date_any(trade_date_raw, prefer_day_first=True),
        value_date=parse_date_any(value_date_raw, prefer_day_first=True),
        quantity=quantity,
        price=price,
        price_currency=ccy or "USD",
        consideration=consideration,
        commission=None,
        net_amount=parse_decimal(total_cash),
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=nominal,
        price_in_percentage=price_pct,
        accrued_interest=parse_decimal(accrued),
        settlement_currency=ccy or "USD",
        parser_template=parser_template,
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "qty_raw": qty_raw,
            "nominal_raw": nominal_raw,
            "price_raw": price_raw,
            "price_pct_raw": price_pct_raw,
            "principal_amount": principal_amount,
            "accrued": accrued,
            "total_cash": total_cash,
            "currency": ccy,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )

    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []

    return [trade]


# =============================================================================
# EXPLICIT PDF PARSERS PER BROKER
# =============================================================================
def parse_cub_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    # Fix CUB OCR artifact: "USD 1 17,994.43" → "USD 117,994.43"
    # OCR sometimes splits a leading digit from the rest of the number
    text = re.sub(r"(USD|EUR|CHF|GBP|AED)\s+(\d)\s+(\d{2,3},\d{3})", r"\1 \2\3", text)

    isin = (
        rx(r"Security\s+Nb\s*:\s*([A-Z0-9]{10,20})", text)
        or rx(r"ISIN(?:\s+Code)?\s*[:\-]?\s*([A-Z0-9]{10,20})", text)
        or rx(r"\b([A-Z]{2}[A-Z0-9]{9,12})\b", text)
    )

    direction_phrase = (
        rx(r"(We\s+(?:SOLD|BOUGHT)\s+(?:to|from)\s+you(?:\s+on)?)", text)
        or rx(r"\b(SELLS|BUYS|BUY|SELL)\b", text)
    )

    trade_date_raw = rx(r"Trade\s+Date\s*:\s*([0-9A-Za-z\-/\. ]+)", text)
    value_date_raw = rx(r"Value\s+Date\s*:\s*([0-9A-Za-z\-/\. ]+)", text)

    security_name = None
    lines = [clean_text(x) for x in text.splitlines() if clean_text(x)]
    for i, line in enumerate(lines):
        if "WE SOLD TO YOU" in line.upper() or "WE BOUGHT FROM YOU" in line.upper():
            for j in range(i + 1, min(i + 8, len(lines))):
                candidate = lines[j]
                if (
                    "TRADE DATE" in candidate.upper()
                    or "VALUE DATE" in candidate.upper()
                    or "MATURITY" in candidate.upper()
                    or "ISSUE DATE" in candidate.upper()
                    or "NOTIONAL" in candidate.upper()
                    or "PRICE" in candidate.upper()
                    or "PRINCIPAL AMOUNT" in candidate.upper()
                ):
                    continue
                if re.search(r"[A-Z]{2}[A-Z0-9]{9,12}", candidate):
                    continue
                if len(candidate) > 3:
                    security_name = candidate
                    break
            break

    nominal_raw = (
        rx(r"Notional\s+[A-Z]{3}\s+([0-9,.\-]+)", text)
        or rx(r"Notional\s*:\s*(?:[A-Z]{3}\s*)?([0-9,.\-]+)", text)
        or rx(r"Nominal\s+Amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,.\-]+)", text)
    )

    price_pct_raw = rx(r"Price\s*[:\-]?\s*([0-9,.\-]+)\s*%", text)
    price_raw = rx(r"Price\s*[:\-]?\s*([0-9,.\-]+)", text)

    principal_amount = rx(r"Principal\s+Amount\s+(?:[A-Z]{3}\s+)?([0-9,.\-]+)", text)
    accrued = rx(r"Accrued\s+Interest\s+(?:[A-Z]{3}\s+)?([0-9,.\-]+)", text)
    total_cash = rx(r"Net\s+amount\s+(?:[A-Z]{3}\s+)?([0-9,.\-]+)", text)

    ccy = (
        rx(r"Notional\s+([A-Z]{3})\s+[0-9,.\-]+", text)
        or rx(r"Principal\s+Amount\s+([A-Z]{3})\s+[0-9,.\-]+", text)
        or rx(r"Net\s+amount\s+([A-Z]{3})\s+[0-9,.\-]+", text)
        or rx(r"\b(USD|EUR|AED|GBP|CHF|HKD)\b", text)
    )

    ref = (
        rx(r"Our\s+ref\s*:\s*([A-Za-z0-9\-/]+)", text)
        or rx(r"(?:Reference|Confirmation\s+No\.?|Trade\s+Reference)\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
    )

    # Settlement section: "Our EC 23860" or "Our DTC0902 // Custodian Account G 44695"
    our_account_raw = (
        rx(r"Our\s+EC\s+(\d+)", text)
        or rx(r"Our\s+DTC\s*(\d+)", text)
    )

    side = normalize_side(direction_phrase, "CUB_PDF")
    nominal = parse_decimal(nominal_raw)
    price_pct = parse_decimal(price_pct_raw)
    price = parse_decimal(price_raw)
    consideration = parse_decimal(principal_amount) or parse_decimal(total_cash)
    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(value_date_raw, prefer_day_first=True)

    if not ref:
        ref = build_generic_reference(
            isin=isin,
            side=side,
            trade_date=trade_date,
            value_date=value_date,
            qty=None,
            price=price if price is not None else price_pct,
            nominal=nominal,
        )

    # CUB bonds: price is always % of par — use price_in_percentage so
    # exact_score() compares against tab_deals.price_in_percentage (not absolute price)
    price_in_pct = price_pct or price

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=None,
        price=None,
        price_currency=ccy or "USD",
        consideration=consideration,
        commission=None,
        net_amount=parse_decimal(total_cash),
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=nominal,
        price_in_percentage=price_in_pct,
        accrued_interest=parse_decimal(accrued),
        settlement_currency=ccy or "USD",
        parser_template="CUB_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "nominal_raw": nominal_raw,
            "price_raw": price_raw,
            "price_pct_raw": price_pct_raw,
            "principal_amount": principal_amount,
            "accrued": accrued,
            "total_cash": total_cash,
            "currency": ccy,
            "security_name": security_name,
            "our_account_raw": our_account_raw,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )

    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []

    return [trade]


def parse_camcap_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    # Format: "CAMcap Markets Ltd SELLS/BUYS"
    # "ISIN Code XS3096274314" (no colon, no %)
    # "Price 98.5000" — percentage price without % sign
    # "Total Cash Settlement Amount 197,000.00 USD"

    direction_phrase = (
        rx(r"CAMcap\s+Markets\s+Ltd\s+(SELLS|BUYS|SELL|BUY)", text)
        or rx(r"\b(SELLS|BUYS|BUY|SELL)\b", text)
    )
    side = normalize_side(direction_phrase, "CAMCAP_PDF")

    isin = (
        rx(r"ISIN\s+Code\s+([A-Z]{2}[A-Z0-9]{9,12})", text)
        or rx(r"ISIN\s*[:\-]?\s*([A-Z]{2}[A-Z0-9]{9,12})", text)
        or rx(r"\b([A-Z]{2}[A-Z0-9]{9,12})\b", text)
    )
    security_name = rx(r"ISIN\s+Code\s+\S+\s*\n(.+)", text)

    trade_date_raw = rx(r"Trade\s+Date\s+(\d{2}/\d{2}/\d{4})", text)
    value_date_raw = rx(r"Settlement\s+Date\s+(\d{2}/\d{2}/\d{4})", text)

    nominal_raw = rx(r"Nominal\s+Amount\s+([0-9,\.]+)", text)
    ccy = rx(r"Currency\s+([A-Z]{3})", text) or rx(r"\b(USD|EUR|AED|GBP|CHF|HKD)\b", text)

    # CamCap uses bond price in % without "%" sign — treat as price_in_percentage
    price_pct_raw = (
        rx(r"Price\s+([0-9\.]+)\s*%", text)
        or rx(r"Price\s+([0-9\.]+)", text)   # no % sign → still percentage for bonds
    )
    principal_raw = rx(r"Principal\s+Amount\s+([0-9,\.]+)", text)
    accrued_raw = rx(r"Accrued\s+Interest\s+([0-9,\.]+)", text)
    total_cash_raw = (
        rx(r"Total\s+Cash\s+Settlement\s+Amount\s+([0-9,\.]+)", text)
        or rx(r"Net\s+amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,\.]+)", text)
    )
    ref = (
        rx(r"Our\s+Ref\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
        or rx(r"(?:Reference|Confirmation\s+No\.?)\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
    )

    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(value_date_raw, prefer_day_first=True)
    nominal = parse_decimal(nominal_raw)
    price_pct = parse_decimal(price_pct_raw)
    consideration = parse_decimal(principal_raw) or parse_decimal(total_cash_raw)
    net_amount = parse_decimal(total_cash_raw)
    accrued = parse_decimal(accrued_raw)

    if not ref:
        ref = build_generic_reference(isin, side, trade_date, value_date, None, price_pct, nominal)

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=None,
        price=None,
        price_currency=ccy or "USD",
        consideration=consideration,
        commission=None,
        net_amount=net_amount,
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=nominal,
        price_in_percentage=price_pct,
        accrued_interest=accrued,
        settlement_currency=ccy or "USD",
        parser_template="CAMCAP_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "nominal_raw": nominal_raw,
            "price_pct_raw": price_pct_raw,
            "principal_raw": principal_raw,
            "total_cash_raw": total_cash_raw,
            "accrued_raw": accrued_raw,
            "ccy": ccy,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []
    return [trade]


def parse_zarattini_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    isin = (
        rx(r"Security\s+Nb\.?\s*[:\-]?\s*([A-Z0-9]{10,20})", text)
        or rx(r"ISIN(?:\s+Code)?\s*[:\-]?\s*([A-Z0-9]{10,20})", text)
        or rx(r"\b([A-Z]{2}[A-Z0-9]{9,12})\b", text)
    )

    # Zarattini format: company name follows "(Fixed)" on the line after "Foreign Notes"
    # e.g.: "USD 179'000 Foreign Notes\n(Fixed) General Motors Co\n2014-1.4.35 Sr\nISIN:..."
    security_name = (
        rx(r"Description\s*[:\-]?\s*(.+)", text)
        or rx(r"Security\s+Name\s*[:\-]?\s*(.+)", text)
        or rx(r"Foreign\s+Notes\s*\r?\n\s*\([A-Za-z]+\)\s+(.+)", text)
    )

    direction_phrase = (
        rx(r"(YOUR\s+SECURITY\s+SALE)", text)
        or rx(r"(YOUR\s+SECURITY\s+PURCHASE)", text)
        or rx(r"(We\s+(?:SOLD|BOUGHT)\s+(?:to|from)\s+you)", text)
        or rx(r"\b(SELLS|BUYS|BUY|SELL)\b", text)
    )

    side = None
    if direction_phrase:
        dp = direction_phrase.upper()
        if "YOUR SECURITY SALE" in dp:
            side = "SELL"
        elif "YOUR SECURITY PURCHASE" in dp:
            side = "BUY"
        else:
            side = normalize_side(direction_phrase, "ZARATTINI_PDF")

    trade_date_raw = rx(r"Trade\s+Date\s*[:\-]?\s*([0-9A-Za-z\-/\. ]+)", text)
    value_date_raw = rx(r"(?:Settlement|Value)\s+Date\s*[:\-]?\s*([0-9A-Za-z\-/\. ]+)", text)

    # Zarattini nominal format: "USD 200'000 Foreign Notes" (after "Stock Exchange:" line)
    nominal_raw = (
        rx(r"Nominal\s+Amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,'.\-]+)", text)
        or rx(r"USD\s+([0-9,']+)\s+Foreign\s+Notes", text)
        or rx(r"EUR\s+([0-9,']+)\s+Foreign\s+Notes", text)
        or rx(r"CHF\s+([0-9,']+)\s+Foreign\s+Notes", text)
    )
    qty_raw = rx(r"Quantity\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,.\-]+)", text)
    price_pct_raw = rx(r"Price\s*[:\-]?\s*([0-9,'.\-]+)\s*%", text)
    price_raw = rx(r"Price\s*[:\-]?\s*([0-9,'.\-]+)", text)
    principal_amount = rx(r"Principal\s+Amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,'.\-]+)", text)
    accrued = (
        rx(r"Accrued\s+Interests?\s*[:\-]?\s*\d+\s+days?\s+(?:[A-Z]{3}\s*)?([0-9,'.\-]+)", text)
        or rx(r"Accrued\s+Interest\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,'.\-]+)", text)
    )
    total_cash = (
        rx(r"To\s+your\s+(?:DEBIT|CREDIT)\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,'.\-]+)", text)
        or rx(r"Total\s+Cash\s+Settlement\s+Amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,'.\-]+)", text)
        or rx(r"Net\s+amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,'.\-]+)", text)
    )

    ccy = (
        rx(r"Currency\s*[:\-]?\s*([A-Z]{3})", text)
        or rx(r"\b(USD|EUR|AED|GBP|CHF|HKD)\b", text)
    )

    ref = (
        rx(r"Ref\.\s*No\.?\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
        or rx(r"(?:Reference|Confirmation\s+No\.?|Trade\s+Reference)\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
    )

    quantity = parse_decimal(qty_raw)
    nominal = parse_decimal(nominal_raw)
    price_pct = parse_decimal(price_pct_raw)
    price = parse_decimal(price_raw)
    consideration = parse_decimal(principal_amount) or parse_decimal(total_cash)

    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(value_date_raw, prefer_day_first=True)

    if not ref:
        ref = build_generic_reference(
            isin=isin,
            side=side,
            trade_date=trade_date,
            value_date=value_date,
            qty=quantity,
            price=price if price is not None else price_pct,
            nominal=nominal,
        )

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=quantity,
        price=price,
        price_currency=ccy or "USD",
        consideration=consideration,
        commission=None,
        net_amount=parse_decimal(total_cash),
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=nominal,
        price_in_percentage=price_pct,
        accrued_interest=parse_decimal(accrued),
        settlement_currency=ccy or "USD",
        parser_template="ZARATTINI_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "qty_raw": qty_raw,
            "nominal_raw": nominal_raw,
            "price_raw": price_raw,
            "price_pct_raw": price_pct_raw,
            "principal_amount": principal_amount,
            "accrued": accrued,
            "total_cash": total_cash,
            "currency": ccy,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )

    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []

    return [trade]


def parse_ashenden_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    # Format: "YOU SOLD IN USD" / "YOU BOUGHT IN USD"
    # ISIN on line after "TICKER: ISIN: SEDOL: CUSIP:" headers
    # "Trade ref: 127815.01" for reference
    # "Net Amount: 108,705.90 USD" — value after headers on next line

    direction_phrase = (
        rx(r"(YOU\s+SOLD\s+IN\s+[A-Z]{3})", text)
        or rx(r"(YOU\s+BOUGHT\s+IN\s+[A-Z]{3})", text)
    )
    side = normalize_side(direction_phrase, "ASHENDEN_PDF")

    # ISIN: on line after "TICKER: ISIN: SEDOL: CUSIP:" — second token in that line
    # Use strict 12-char pattern ending in digit to avoid false positives (e.g. "NOTIFICATION")
    isin = rx(r"ISIN[: \t]+([A-Z]{2}[A-Z0-9]{9}[0-9])", text)
    if not isin:
        # Table layout: "TICKER:  ISIN:  SEDOL:  CUSIP:\nNOKIA  US654902AC90  ..."
        m = re.search(
            r"TICKER[: \t]+ISIN[: \t]+SEDOL[: \t]+CUSIP[: \t]*\r?\n\S+[ \t]+([A-Z]{2}[A-Z0-9]{9}[0-9])",
            text, re.IGNORECASE
        )
        isin = m.group(1) if m else None
    if not isin:
        # Fallback: strict ISIN — exactly 12 chars, last char must be digit (check digit)
        isin = rx(r"\b([A-Z]{2}[A-Z0-9]{9}[0-9])\b", text)

    # Security name: first non-digit line directly after "Quantity   At the price of:" header.
    # Real PDF layout (all variants):
    #   Quantity At the price of:
    #   BOND NAME [possibly with coupon/maturity]
    #   400,000 97.00            ← qty/price row (starts with digit — excluded)
    _EXCLUDE = ("AM WEALTH", "ASHENDEN", "FOLLOWING TRANSACTION", "YOUR BEHALF", "YOUR ACCOUNT")
    security_name = None
    _m = re.search(
        r"Quantity\s+At\s+the\s+price\s+of\s*[:\-]?\s*\r?\n\s*([^0-9].+)",
        text, re.IGNORECASE
    )
    if _m:
        _cand = _m.group(1).strip()
        if not any(e in _cand.upper() for e in _EXCLUDE):
            security_name = _cand or None

    trade_date_raw = rx(r"Trade\s+Date\s*[:\-]?\s*(\d{2}[-/]\w{3}[-/]\d{2,4})", text)
    value_date_raw = rx(r"Settlement\s+Date\s*[:\-]?\s*(\d{2}[-/]\w{3}[-/]\d{2,4})", text)

    # Quantity and price: on the data row that starts with the bond nominal.
    # pdfplumber merges columns so the row may look like any of:
    #   "400,000 97.00"
    #   "250,000 16-Feb-2031 71.25"           (maturity date merged in)
    #   "555,000 Coupon Share Basket 25-Feb-2031 97.00"  (name continuation merged in)
    #   "200,000 7.200000% 17-Jul-2030 105.75"  (coupon rate merged in)
    # Rule: first large integer on the line = qty; last decimal number = price.
    qty_raw = None
    price_raw = None
    _qty_price_m = re.search(
        r"^([1-9][0-9,]+)\b.*\b([0-9]+\.[0-9]+)\s*$",
        text, re.MULTILINE
    )
    if _qty_price_m:
        qty_raw = _qty_price_m.group(1)
        price_raw = _qty_price_m.group(2)

    # Safer override: if "At the price of: NNN" is explicit, use that as price
    _explicit_price = rx(r"At\s+the\s+price\s+of\s*[:\-]?\s*([0-9,.]+)", text)
    if _explicit_price:
        price_raw = _explicit_price

    # Amounts: "Gross Amount: Net Amount:\n 106,700.00 USD  108,705.90 USD"
    # Net amount is the second value on that data line
    net_amount_raw = None
    gross_amount_raw = None
    m = re.search(r"Gross\s+Amount\s*[:\-]?\s*Net\s+Amount\s*[:\-]?\s*\n\s*([0-9,\.]+)\s*[A-Z]{3}\s+([0-9,\.]+)\s*[A-Z]{3}", text, re.IGNORECASE)
    if m:
        gross_amount_raw = m.group(1)
        net_amount_raw = m.group(2)
    else:
        gross_amount_raw = rx(r"Gross\s+Amount\s*[:\-]?\s*([0-9,\.]+)", text)
        net_amount_raw = rx(r"Net\s+Amount\s*[:\-]?\s*([0-9,\.]+)", text)

    accrued = rx(r"Accrued\s+Interest\s*[/\s]\s*Days\s*\n?\s*([0-9,\.]+)", text)

    ccy = rx(r"YOU\s+(?:SOLD|BOUGHT)\s+IN\s+([A-Z]{3})", text) or rx(r"\b(USD|EUR|AED|GBP|CHF|HKD)\b", text)

    ref = (
        rx(r"Trade\s+ref\s*[:\-]?\s*([A-Za-z0-9\.\-/]+)", text)
        or rx(r"(?:Reference|Confirmation\s+No\.?)\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
    )

    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(value_date_raw, prefer_day_first=True)
    quantity = parse_decimal(qty_raw)
    price = parse_decimal(price_raw)
    consideration = parse_decimal(gross_amount_raw)
    net_amount = parse_decimal(net_amount_raw)
    accrued_dec = parse_decimal(accrued)

    # Derive nominal from consideration / price if quantity not extracted directly.
    # For bonds: nominal = gross_amount / (price/100) → must round to nearest integer.
    nominal = quantity
    if nominal is None and price and consideration and price > 0:
        try:
            derived = (consideration * Decimal("100") / price).quantize(Decimal("1"))
            nominal = derived
            if quantity is None:
                quantity = derived
        except Exception:
            pass

    if not ref:
        ref = build_generic_reference(isin, side, trade_date, value_date, quantity, price)

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=quantity,
        price=price,
        price_currency=ccy or "USD",
        consideration=consideration,
        commission=None,
        net_amount=net_amount,
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=nominal,
        price_in_percentage=price,
        accrued_interest=accrued_dec,
        settlement_currency=ccy or "USD",
        parser_template="ASHENDEN_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "qty_raw": qty_raw,
            "price_raw": price_raw,
            "gross_amount_raw": gross_amount_raw,
            "net_amount_raw": net_amount_raw,
            "accrued": accrued,
            "currency": ccy,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []
    return [trade]


def parse_stonex_fixed_income_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    # Format: "We confirm your BUY/SELL transaction with us"
    # Fields: Our Reference, Trade Date, Settlement Date, ISIN, Quantity,
    #         Gross Price (USD), Accrued Interest (USD), Net Settlement Amount (USD)

    direction_phrase = rx(r"We\s+confirm\s+your\s+(BUY|SELL)\s+transaction", text)
    side = normalize_side(direction_phrase, "STONEX_PDF")

    isin = rx(r"ISIN\s*[:\-]?\s*([A-Z]{2}[A-Z0-9]{9,12})", text)
    security_name = rx(r"Security\s+Description\s*[:\-]?\s*(.+)", text)
    trade_date_raw = rx(r"Trade\s+Date\s*[:\-]?\s*([0-9/\-\.]+)", text)
    value_date_raw = rx(r"Settlement\s+Date\s*[:\-]?\s*([0-9/\-\.]+)", text)
    qty_raw = rx(r"Quantity\s*[:\-]?\s*([0-9,]+)", text)
    # Gross Price: "104.910000USD" or "99.000000 GBP" — price in % of par
    price_pct_raw = rx(r"Gross\s+Price\s*[:\-]?\s*([0-9,.]+?)\s*[A-Z]{3}", text)
    gross_amount_raw = rx(r"Gross\s+Amount\s*[:\-]?\s*([0-9,\.]+?)\s*[A-Z]{3}", text)
    accrued_raw = rx(r"Accrued\s+Interest\s*[:\-]?\s*([0-9,\.]+?)\s*[A-Z]{3}", text)
    net_amount_raw = rx(r"Net\s+Settlement\s+Amount\s*[:\-]?\s*([0-9,\.]+?)\s*[A-Z]{3}", text)
    ref = rx(r"Our\s+Reference\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
    # Currency: from Gross Price field (handles GBP, EUR, etc.) or FX Rate line, fallback USD
    ccy = (
        rx(r"Gross\s+Price\s*[:\-]?\s*[0-9,.]+\s*([A-Z]{3})", text)
        or rx(r"FX\s+Rate\s+\d+\s+([A-Z]{3})/[A-Z]{3}", text)
        or "USD"
    )

    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(value_date_raw, prefer_day_first=True)
    quantity = parse_decimal(qty_raw)
    price_pct = parse_decimal(price_pct_raw)
    consideration = parse_decimal(gross_amount_raw)
    net_amount = parse_decimal(net_amount_raw)
    accrued = parse_decimal(accrued_raw)

    if not ref:
        ref = build_generic_reference(isin, side, trade_date, value_date, quantity, price_pct)

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=quantity,
        price=None,
        price_currency=ccy,
        consideration=consideration,
        commission=None,
        net_amount=net_amount,
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=quantity,
        price_in_percentage=price_pct,
        accrued_interest=accrued,
        settlement_currency=ccy,
        parser_template="STONEX_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "qty_raw": qty_raw,
            "price_pct_raw": price_pct_raw,
            "gross_amount_raw": gross_amount_raw,
            "accrued_raw": accrued_raw,
            "net_amount_raw": net_amount_raw,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []
    return [trade]


def parse_instinet_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    # Format: Instinet settlement confirmation PDF
    # Key fields: "Participant Account No B01824" → CPTY SSI "MS-HK-CCAS/B01824"
    # Also: ISIN, Buy/Sell, Trade Date, Settlement Date, Quantity, Price, Net Amount

    direction_phrase = rx(r"(BUY|SELL)", text, re.IGNORECASE)
    side = normalize_side(direction_phrase, "INSTINET_PDF")

    isin = rx(r"\b([A-Z]{2}[A-Z0-9]{9}[0-9])\b", text)
    security_name = rx(r"Security\s+(?:Description|Name)\s*[:\-]?\s*(.+)", text)
    trade_date_raw = rx(r"Trade\s+Date\s*[:\-]?\s*([0-9]{1,2}[-/][0-9]{1,2}[-/][0-9]{2,4}|[A-Za-z]+\s+\d+,?\s+\d{4})", text)
    value_date_raw = rx(r"Settlement\s+Date\s*[:\-]?\s*([0-9]{1,2}[-/][0-9]{1,2}[-/][0-9]{2,4}|[A-Za-z]+\s+\d+,?\s+\d{4})", text)
    qty_raw = rx(r"(?:Quantity|Shares)\s*[:\-]?\s*([0-9,]+)", text)
    price_raw = rx(r"(?:Price|Trade\s+Price)\s*[:\-]?\s*([0-9,\.]+)", text)
    net_amount_raw = rx(r"Net\s+(?:Settlement\s+)?Amount\s*[:\-]?\s*([0-9,\.]+)", text)
    ccy = rx(r"\b(USD|EUR|GBP|HKD|AED|CHF)\b", text) or "USD"

    # CPTY SSI: "Participant Account No B01824" → "MS-HK-CCAS/B01824"
    participant_account = rx(r"Participant\s+Account\s+No\.?\s+(\w+)", text)
    cpty_ssi = f"MS-HK-CCAS/{participant_account}" if participant_account else None

    ref = (
        rx(r"(?:Confirmation|Reference|Order)\s+(?:No\.?|Number)\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
        or rx(r"Ref(?:erence)?\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
    )

    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(value_date_raw, prefer_day_first=True)
    quantity = parse_decimal(qty_raw)
    price = parse_decimal(price_raw)
    net_amount = parse_decimal(net_amount_raw)

    if not ref:
        ref = build_generic_reference(isin, side, trade_date, value_date, quantity, price)

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=quantity,
        price=price,
        price_currency=ccy,
        consideration=None,
        commission=None,
        net_amount=net_amount,
        settlement_terms="DVP",
        counterparty_reference=cpty_ssi or ref,
        nominal=quantity,
        price_in_percentage=None,
        accrued_interest=None,
        settlement_currency=ccy,
        parser_template="INSTINET_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "qty_raw": qty_raw,
            "price_raw": price_raw,
            "net_amount_raw": net_amount_raw,
            "participant_account": participant_account,
            "cpty_ssi": cpty_ssi,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []
    return [trade]


def parse_enbd_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    # Format: Emirates NBD settlement confirmation PDF
    # "Security Identification Number : ISIN: XS..."
    # "We confirm our Buy order" → ENBD buys from AM Wealth → AM Wealth SELL
    # "We confirm our Sell order" → ENBD sells to AM Wealth → AM Wealth BUY
    # Fields:
    #   Trade Date : 23 March 2026
    #   Settlement Date : 25 March 2026
    #   Nominal/Number of Shares : 400,000.00
    #   Price/Yield : 99.4
    #   Principal : 397,600.00
    #   Accrued Interest : 1,500.00
    #   Total Consideration (Settlement Amount) : 399,100.00
    # Reference: confirmation number at top of PDF (e.g. "143716226")

    isin = (
        rx(r"Security\s+Identification\s+Number\s*[:\-]?\s*ISIN\s*[:\-]?\s*([A-Z]{2}[A-Z0-9]{9,12})", text)
        or rx(r"ISIN\s*[:\-]?\s*([A-Z]{2}[A-Z0-9]{9,12})", text)
        or rx(r"\b([A-Z]{2}[A-Z0-9]{9,12})\b", text)
    )

    # ENBD perspective: "our Buy" = they buy = AM Wealth SELLS; "our Sell" = they sell = AM Wealth BUYS
    direction_phrase = (
        rx(r"(We\s+confirm\s+our\s+(?:Buy|Sell)\s+order)", text, re.IGNORECASE)
        or rx(r"\b(Buy|Sell)\b", text, re.IGNORECASE)
    )
    # Invert: ENBD's Buy = our Sell; ENBD's Sell = our Buy
    if direction_phrase:
        v = direction_phrase.upper()
        if "BUY" in v:
            side = "SELL"
        elif "SELL" in v:
            side = "BUY"
        else:
            side = None
    else:
        side = None

    # Date format: "23 March 2026" — full month name, space-separated
    trade_date_raw = rx(r"Trade\s+Date\s*:\s*(.+)", text)
    value_date_raw = (
        rx(r"Settlement\s+Date\s*:\s*(.+)", text)
        or rx(r"Value\s+Date\s*:\s*(.+)", text)
    )

    # "Nominal/Number of Shares : 400,000.00" — slash in label
    nominal_raw = (
        rx(r"Nominal[^:]*:\s*([0-9,\.]+)", text)
        or rx(r"Face\s+Value\s*:\s*([0-9,\.]+)", text)
    )

    # Price/Yield is a percentage: "Price/Yield : 99.4"
    price_pct_raw = (
        rx(r"Price/Yield\s*:\s*([0-9,\.]+)", text)
        or rx(r"Price\s*:\s*([0-9,\.]+)", text)
    )

    principal_raw = rx(r"Principal\s*:\s*([0-9,\.]+)", text)
    accrued_raw = rx(r"Accrued\s+Interest\s*:\s*([0-9,\.]+)", text)
    # "Total Consideration (Settlement Amount) : 399,100.00" — parens in label
    total_raw = (
        rx(r"Total\s+Consideration[^:]*:\s*([0-9,\.]+)", text)
        or rx(r"Net\s+(?:Settlement\s+)?Amount\s*:\s*([0-9,\.]+)", text)
        or rx(r"Total\s+Amount\s*:\s*([0-9,\.]+)", text)
    )

    ccy = (
        rx(r"(?:Nominal|Principal|Total)\s+(?:Amount\s+)?([A-Z]{3})\s+[0-9,\.]+", text)
        or rx(r"\b(USD|EUR|AED|GBP|CHF|HKD)\b", text)
    )

    ref = (
        rx(r"(?:Confirmation|Reference|Transaction)\s*(?:No\.?|Number|Ref\.?)\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
        or rx(r"Ref\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
        or rx(r"^([0-9]{6,12})\b", text, re.MULTILINE)
    )

    security_name = (
        rx(r"Security\s+(?:Name|Description)\s*[:\-]?\s*(.+)", text)
        or rx(r"Instrument\s*[:\-]?\s*(.+)", text)
    )

    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(value_date_raw, prefer_day_first=True)
    nominal = parse_decimal(nominal_raw)
    price_pct = parse_decimal(price_pct_raw)
    consideration = parse_decimal(principal_raw)
    accrued = parse_decimal(accrued_raw)
    total = parse_decimal(total_raw)

    if not ref:
        ref = build_generic_reference(isin, side, trade_date, value_date, None, price_pct, nominal)

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=None,
        price=None,
        price_currency=ccy or "USD",
        consideration=consideration or total,
        commission=None,
        net_amount=total,
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=nominal,
        price_in_percentage=price_pct,
        accrued_interest=accrued,
        settlement_currency=ccy or "USD",
        parser_template="ENBD_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "nominal_raw": nominal_raw,
            "price_pct_raw": price_pct_raw,
            "principal_raw": principal_raw,
            "accrued_raw": accrued_raw,
            "total_raw": total_raw,
            "currency": ccy,
            "security_name": security_name,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []
    return [trade]


def parse_enbd_securities_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    # Format: ENBD Securities "Order Confirmation Report" (GCM / Sincy Joji)
    # Sample fields:
    #   ISIN              : AEE01657D252
    #   Symbol            : DUBAIRESI
    #   Sell Confirmation / Buy Confirmation  (direct: matches AM Wealth perspective)
    #   Transaction Date  : 01-Apr-2026  (or "April 1, 2026" etc.)
    #   Quantity          : 3,268
    #   Price             : 1.12000
    #   Trading Amount    : 3,660.16
    #   Settlement Date   : 06-Apr-2026
    #   NET               : 3,639.18
    # Direction is NOT inverted — "Sell Confirmation" = AM Wealth SOLD

    logging.warning("ENBD_SECURITIES_DEBUG RAW TEXT START >>>\n%s\n<<< ENBD_SECURITIES_DEBUG RAW TEXT END", text)

    isin = (
        rx(r"ISIN\s*[:\-]?\s*([A-Z]{2}[A-Z0-9]{9,12})", text)
        or rx(r"\b([A-Z]{2}[A-Z0-9]{9,12})\b", text)
    )

    security_name = rx(r"Symbol\s*[:\-]?\s*(\S+)", text)

    # Detect direction from confirmation header
    sell_match = rx(r"(Sell\s+Confirmation)", text, re.IGNORECASE)
    buy_match = rx(r"(Buy\s+Confirmation)", text, re.IGNORECASE)
    if sell_match:
        side = "SELL"
        direction_phrase = sell_match
    elif buy_match:
        side = "BUY"
        direction_phrase = buy_match
    else:
        side = None
        direction_phrase = None

    # The table data row is: OrderNo  DD-Mon-YYYY  Qty  Price  TradingAmt  DD-Mon-YYYY
    # e.g. "4008352012 02-Apr-2026 3,268 1.12000 3,660.16 06-Apr-2026"
    _data_row = re.search(
        r"\b\d{7,12}\s+(\d{2}-[A-Za-z]{3}-\d{4})\s+([\d,]+)\s+([\d.]+)\s+([\d,\.]+)\s+(\d{2}-[A-Za-z]{3}-\d{4})\b",
        text,
    )
    if _data_row:
        trade_date_raw    = _data_row.group(1)
        quantity_raw      = _data_row.group(2)
        price_raw         = _data_row.group(3)
        trading_amount_raw = _data_row.group(4)
        value_date_raw    = _data_row.group(5)
    else:
        trade_date_raw = (
            rx(r"From\s+Date\s+(\d{2}-[A-Za-z]{3}-\d{4})", text)
            or rx(r"Transaction\s+Date\s+(\d{2}-[A-Za-z]{3}-\d{4})", text)
            or rx(r"Trade\s+Date\s*[:\-]?\s*(\S+)", text)
        )
        value_date_raw = rx(r"Settlement\s+Date\s+(\d{2}-[A-Za-z]{3}-\d{4})", text)
        quantity_raw = (
            rx(r"Total\s+([\d,]+)\s+[\d,\.]+", text)
            or rx(r"Quantity\s*[:\-]?\s*([\d,\.]+)", text)
            or rx(r"No\.?\s+of\s+Shares\s*[:\-]?\s*([0-9,\.]+)", text)
        )
        price_raw = (
            rx(r"Gross\s+Avg\s+Price\s+([\d.]+)", text)
            or rx(r"Net\s+Avg\s+Price\s+([\d.]+)", text)
        )
        trading_amount_raw = rx(r"Total\s+[\d,]+\s+([\d,\.]+)", text)
    net_raw = (
        rx(r"\bNET\b\s*[:\-]?\s*([0-9,\.]+)", text)
        or rx(r"Net\s+Amount\s*[:\-]?\s*([0-9,\.]+)", text)
        or rx(r"Net\s+Consideration\s*[:\-]?\s*([0-9,\.]+)", text)
    )

    ccy = rx(r"\b(USD|AED|EUR|GBP|CHF)\b", text)

    ref = (
        rx(r"(?:Order|Confirmation|Reference|Transaction)\s*(?:No\.?|Number|Ref\.?|ID)\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
        or rx(r"^([0-9]{5,12})\b", text, re.MULTILINE)
    )

    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(value_date_raw, prefer_day_first=True)
    quantity = parse_decimal(quantity_raw)
    price = parse_decimal(price_raw)
    trading_amount = parse_decimal(trading_amount_raw)
    net_amount = parse_decimal(net_raw)

    if not ref:
        ref = build_generic_reference(isin, side, trade_date, value_date, quantity, price, None)

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=quantity,
        price=price,
        price_currency=ccy or "AED",
        consideration=trading_amount,
        commission=None,
        net_amount=net_amount or trading_amount,
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=None,
        price_in_percentage=None,
        accrued_interest=None,
        settlement_currency=ccy or "AED",
        parser_template="ENBD_SECURITIES_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "quantity_raw": quantity_raw,
            "price_raw": price_raw,
            "trading_amount_raw": trading_amount_raw,
            "net_raw": net_raw,
            "currency": ccy,
            "security_name": security_name,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []
    return [trade]


def parse_bondpartners_pdf(text, internet_message_id, source_file, email_received_at, processing_run_id, file_id, email_id, broker_name):
    logging.warning("BONDPARTNERS_PDF RAW TEXT START >>>\n%s\n<<< BONDPARTNERS_PDF RAW TEXT END", text)
    # Format: Bondpartners SA contract note
    # Direction: "Transaction type  Your purchase  Our sale"
    # Trade date: "Trade date  Mar. 25, 2026 11:53 CET"
    # Value date embedded in: "Net consideration  To your debit  Value date Mar. 26, 2026  USD 269'732.50"
    # Nominal: "Nominal amount  240'000.00" (Swiss apostrophe thousands separator)
    # Price: "Trade price  110.600000  pct"
    # Accrued: "Accrued interests  101 days  USD  4'292.50"
    # Total: "Total amount  USD  269'732.50"

    direction_phrase = (
        rx(r"Transaction\s+type\s+(Your\s+(?:purchase|sale))", text)
        or rx(r"(Your\s+(?:purchase|sale))", text)
    )

    trade_date_raw = (
        rx(r"Trade\s+date\s+(\w+\.?\s+\d+,\s+\d{4})", text)
        or rx(r"Trade\s+[Dd]ate\s+(\w+[.,]\s*\d+[.,]\s*\d{4})", text)
    )
    value_date_raw = rx(r"Value\s+date\s*(\w+\.?\s+\d+,\s+\d{4})", text)

    ccy = rx(r"Currency\s+([A-Z]{3})\b", text)
    nominal_raw = rx(r"Nominal\s+amount\s+([0-9][0-9',.]*)", text)
    price_pct_raw = rx(r"Trade\s+price\s+([0-9.,]+)\s*\n?\s*pct", text)
    price_raw = rx(r"Trade\s+price\s+([0-9.,]+)", text)

    # Use same proven ISIN regex as other parsers
    isin = (
        rx(r"ISIN(?:\s+Code)?\s*[:\-]?\s*([A-Z0-9]{10,20})", text)
        or rx(r"\bISIN\s+([A-Z]{2}[A-Z0-9]{9,12})", text)
    )

    principal_amount = rx(r"Gross\s+consideration\s+[A-Z]{3}\s+([0-9][0-9',.]*)", text)
    accrued = rx(r"Accrued\s+interests?\s+\d+\s+days?\s+[A-Z]{3}\s+([0-9][0-9',.]*)", text)
    total_cash = rx(r"Total\s+amount\s+[A-Z]{3}\s+([0-9][0-9',.]*)", text)

    logging.info(
        "BONDPARTNERS PDF extract: isin=%s side=%s trade_date=%s value_date=%s nominal=%s price=%s accrued=%s total=%s",
        isin, direction_phrase, trade_date_raw, value_date_raw, nominal_raw, price_raw, accrued, total_cash,
    )

    side = normalize_side(direction_phrase, "BONDPARTNERS_PDF")
    nominal = parse_decimal(nominal_raw)
    price_pct = parse_decimal(price_pct_raw or price_raw)
    price = parse_decimal(price_raw)
    consideration = parse_decimal(principal_amount)
    net_amount = parse_decimal(total_cash)

    ref = build_generic_reference(
        isin=isin,
        side=side,
        trade_date=parse_date_any(trade_date_raw, True),
        value_date=parse_date_any(value_date_raw, True),
        qty=None,
        price=price if price is not None else price_pct,
        nominal=nominal,
    )

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=None,
        isin=isin,
        side=side,
        trade_date=parse_date_any(trade_date_raw, prefer_day_first=True),
        value_date=parse_date_any(value_date_raw, prefer_day_first=True),
        quantity=nominal,
        price=price,
        price_currency=ccy or "USD",
        consideration=consideration,
        commission=None,
        net_amount=net_amount,
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=nominal,
        price_in_percentage=price_pct,
        accrued_interest=parse_decimal(accrued),
        settlement_currency=ccy or "USD",
        parser_template="BONDPARTNERS_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "nominal_raw": nominal_raw,
            "price_raw": price_raw,
            "price_pct_raw": price_pct_raw,
            "principal_amount": principal_amount,
            "accrued": accrued,
            "total_cash": total_cash,
            "currency": ccy,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )

    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []

    return [trade]


def parse_seaport_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    # Format: single data row after 3-line header block
    # Columns: Side  ISIN  Quantity  Price  SettlDate  NetConsideration  Ccy  Accrued  GrossConsideration  TradeDate  InstrumentName
    # Example: "Buy XS0701227075 200000 116.75 17/02/2026 -237548.61 USD 4048.6100 -233500 12/02/2026 AM WEALTH ..."
    pattern = re.compile(
        r"^(Buy|Sell)\s+"
        r"([A-Z]{2}[A-Z0-9]{9,12})\s+"
        r"(-?[\d,]+)\s+"           # quantity
        r"(-?[\d.]+)\s+"           # price
        r"(\d{2}/\d{2}/\d{4})\s+"  # settlement date
        r"(-?[\d,.]+)\s+"          # net consideration
        r"([A-Z]{3})\s+"           # currency
        r"(-?[\d,.]+)\s+"          # accrued interest
        r"(-?[\d,.]+)\s+"          # gross consideration
        r"(\d{2}/\d{2}/\d{4})",    # trade date
        re.IGNORECASE | re.MULTILINE,
    )

    rows: List[Dict[str, Any]] = []
    for m in pattern.finditer(text):
        side_raw, isin, qty_raw, price_raw, settle_raw, net_raw, ccy, accrued_raw, gross_raw, trade_raw = m.groups()

        # Instrument name: rest of the line after trade date
        rest = text[m.end():].split("\n")[0].strip()
        security_name = re.sub(r"^AM\s+WEALTH\s*", "", rest).strip() or None

        trade_date = parse_date_any(trade_raw, prefer_day_first=True)
        value_date = parse_date_any(settle_raw, prefer_day_first=True)
        quantity = parse_decimal(qty_raw)
        price = parse_decimal(price_raw)
        net_amount = parse_decimal(net_raw)
        consideration = parse_decimal(gross_raw)
        accrued = parse_decimal(accrued_raw)
        # Seaport confirmation shows Seaport's action: "Buy" = Seaport buys FROM AM Wealth = AM Wealth SELLS
        _seaport_side = normalize_side(side_raw, "SEAPORT_PDF")
        side = "SELL" if _seaport_side == "BUY" else ("BUY" if _seaport_side == "SELL" else _seaport_side)

        ref = build_generic_reference(isin, side, trade_date, value_date, quantity, price)

        # Extract "OUR SSI : ECLR 75663" → account for enrich_cpty_ssi matching
        our_ssi_raw = rx(r"OUR\s+SSI\s*:\s*(?:ECLR|EUROCLEAR|DTC|CEDE)\s+(\w+)", text)

        trade = build_trade_dict(
            internet_message_id=internet_message_id,
            source_file=source_file,
            source_type="pdf",
            broker_name=broker_name,
            security_name=security_name,
            isin=isin,
            side=side,
            trade_date=trade_date,
            value_date=value_date,
            quantity=quantity,
            price=price,
            price_currency=ccy,
            consideration=consideration,
            commission=None,
            net_amount=net_amount,
            settlement_terms="DVP",
            counterparty_reference=ref,
            nominal=quantity,
            price_in_percentage=price,   # bond price is in % of par
            accrued_interest=accrued,
            settlement_currency=ccy,
            parser_template="SEAPORT_PDF",
            raw_json=json.dumps({
                "side_raw": side_raw, "isin": isin,
                "qty_raw": qty_raw, "price_raw": price_raw,
                "settle_raw": settle_raw, "trade_raw": trade_raw,
                "net_raw": net_raw, "gross_raw": gross_raw,
                "accrued_raw": accrued_raw, "ccy": ccy,
                "our_ssi_raw": our_ssi_raw,
            }, default=str),
            processing_run_id=processing_run_id,
            file_id=file_id,
            email_id=email_id,
            side_original_text=side_raw,
            trade_date_original_text=trade_raw,
            value_date_original_text=settle_raw,
        )
        trade = normalize_trade_signs(trade)
        finalize_trade_validation(trade, email_received_at)
        rows.append(trade)

    logging.info("SEAPORT PDF parsed rows=%s", len(rows))
    return rows


def parse_bridport_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    # Format: "We BOUGHT/SOLD from/to you on DATE for settlement on DATE"
    # ISIN labeled as "Security Nb : USU71878AA76"
    # "Nominal USD 100'000.00" (no "Amount" keyword)
    # "Price 101.42500 %" — with space before %
    # Apostrophe thousands separator throughout

    isin = (
        rx(r"Security\s+Nb\s*[:\-]?\s*([A-Z]{2}[A-Z0-9]{9,12})", text)
        or rx(r"ISIN\s*[:\-]?\s*([A-Z]{2}[A-Z0-9]{9,12})", text)
        or rx(r"\b([A-Z]{2}[A-Z0-9]{9,12})\b", text)
    )

    security_name = rx(r"Security\s+Nb\s*[:\-]?\s*\S+\s*\n(.+)", text)

    # Direction + trade date embedded: "We BOUGHT from you on 18 February 2025"
    direction_phrase = (
        rx(r"(We\s+(?:SOLD|BOUGHT)\s+(?:to|from)\s+you\s+on)", text)
        or rx(r"(We\s+(?:SOLD|BOUGHT)\s+(?:to|from)\s+you)", text)
    )
    side = normalize_side(direction_phrase, "BRIDPORT_PDF")

    # Trade date from direction phrase
    trade_date_raw = rx(
        r"We\s+(?:SOLD|BOUGHT)\s+(?:to|from)\s+you\s+on\s+(\d{1,2}\s+\w+\s+\d{4}|\d{2}[-/]\w{3}[-/]\d{2,4}|\d{2}/\d{2}/\d{4})",
        text,
    )
    # Settlement date from "for settlement on DATE"
    value_date_raw = rx(
        r"for\s+settlement\s+on\s+(\d{1,2}\s+\w+\s+\d{4}|\d{2}[-/]\w{3}[-/]\d{2,4}|\d{2}/\d{2}/\d{4})",
        text,
    )
    if not value_date_raw:
        value_date_raw = rx(r"(?:Settlement|Value)\s+Date\s*[:\-]?\s*([0-9A-Za-z\-/\. ]+)", text)

    # "Nominal USD 100'000.00"
    nominal_raw = (
        rx(r"Nominal\s+[A-Z]{3}\s+([0-9,'\.]+)", text)
        or rx(r"Nominal\s+Amount\s*[:\-]?\s*(?:[A-Z]{3}\s*)?([0-9,'\.]+)", text)
    )
    price_pct_raw = rx(r"Price\s+([0-9,'\.]+)\s*%", text)
    price_raw = rx(r"Price\s*[:\-]?\s*([0-9,'\.]+)", text)
    principal_raw = rx(r"Principal\s+Amount\s+(?:[A-Z]{3}\s+)?([0-9,'\.]+)", text)
    accrued_raw = rx(r"Accrued\s+Interest\s*(?:[A-Z]{3}\s+)?([0-9,'\.]+)", text)
    net_raw = rx(r"Net\s+amount\s+(?:[A-Z]{3}\s+)?([0-9,'\.]+)", text)
    ccy = rx(r"\b(USD|EUR|AED|GBP|CHF|HKD)\b", text)
    ref = (
        rx(r"Our\s+ref\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
        or rx(r"(?:Reference|Confirmation\s+No\.?)\s*[:\-]?\s*([A-Za-z0-9\-/]+)", text)
    )

    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(value_date_raw, prefer_day_first=True)
    nominal = parse_decimal(nominal_raw)
    price_pct = parse_decimal(price_pct_raw)
    price = parse_decimal(price_raw) if not price_pct_raw else None
    consideration = parse_decimal(principal_raw) or parse_decimal(net_raw)
    net_amount = parse_decimal(net_raw)
    accrued = parse_decimal(accrued_raw)

    if not ref:
        ref = build_generic_reference(isin, side, trade_date, value_date, None, price_pct or price, nominal)

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file=source_file,
        source_type="pdf",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=None,
        price=price,
        price_currency=ccy or "USD",
        consideration=consideration,
        commission=None,
        net_amount=net_amount,
        settlement_terms="DVP",
        counterparty_reference=ref,
        nominal=nominal,
        price_in_percentage=price_pct,
        accrued_interest=accrued,
        settlement_currency=ccy or "USD",
        parser_template="BRIDPORT_PDF",
        raw_json=json.dumps({
            "direction_phrase": direction_phrase,
            "trade_date_raw": trade_date_raw,
            "value_date_raw": value_date_raw,
            "nominal_raw": nominal_raw,
            "price_pct_raw": price_pct_raw,
            "principal_raw": principal_raw,
            "accrued_raw": accrued_raw,
            "net_raw": net_raw,
            "ccy": ccy,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=file_id,
        email_id=email_id,
        side_original_text=direction_phrase,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=value_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    if not trade.get("isin") and not trade.get("security_name"):
        return []
    return [trade]


def parse_gtn_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    text = re.sub(r"([A-Z0-9]{8,12})\s*\n\s*([A-Z0-9]{1,4})", r"\1\2", text)

    pattern = re.compile(
        r"""
        (?P<symbol>[A-Z0-9._-]+)\s+
        (?P<side>Buy|Sell)\s+
        (?P<quantity>[\d,]+\.\d+)\s+
        (?P<price>\(?[\d,]+\.\d+\)?)\s+
        (?P<gross>\(?[\d,]+\.\d+\)?)\s+
        (?P<brok_com>[\d,]+\.\d+)\s+
        (?P<mkt_com>[\d,]+\.\d+)\s+
        (?P<net_settle>\(?[\d,]+\.\d+\)?)\s+
        (?P<stl_date>\d{2}/\d{2}/\d{4})\s+
        (?P<isin>[A-Z0-9]{10,14}).*?
        (?P<tr_date>\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}\s+[ap]m)
        """,
        re.IGNORECASE | re.VERBOSE | re.DOTALL,
    )

    for m in pattern.finditer(text):
        row = m.groupdict()

        trade_dt = parse_datetime_any(row["tr_date"], prefer_day_first=False)
        trade_date = trade_dt.date() if trade_dt else parse_date_any(row["tr_date"], prefer_day_first=False)
        value_date = parse_date_any(row["stl_date"], prefer_day_first=False)

        trade = build_trade_dict(
            internet_message_id=internet_message_id,
            source_file=source_file,
            source_type="pdf",
            broker_name=broker_name,
            security_name=row["symbol"],
            isin=row["isin"],
            side=normalize_side(row["side"], "GTN_XLS_PDF"),
            trade_date=trade_date,
            value_date=value_date,
            quantity=parse_decimal(row["quantity"]),
            price=parse_decimal(row["price"]),
            price_currency="USD",
            consideration=parse_decimal(row["gross"]),
            commission=parse_decimal(row["brok_com"]),
            net_amount=parse_decimal(row["net_settle"]),
            settlement_terms=None,
            counterparty_reference=None,
            nominal=None,
            price_in_percentage=None,
            accrued_interest=None,
            settlement_currency="USD",
            parser_template="GTN_XLS_PDF",
            raw_json=json.dumps(row, default=str),
            processing_run_id=processing_run_id,
            file_id=file_id,
            email_id=email_id,
            side_original_text=row["side"],
            trade_date_original_text=row["tr_date"],
            value_date_original_text=row["stl_date"],
        )

        trade = normalize_trade_signs(trade)
        trade["counterparty_reference"] = build_generic_reference(
            trade["isin"], trade["side"], trade["trade_date"], trade["value_date"], trade["quantity"], trade["price"]
        )
        finalize_trade_validation(trade, email_received_at)
        rows.append(trade)

    logging.info("GTN PDF parsed rows=%s", len(rows))
    return rows


def parse_stonex_daily_statement_pdf(
    text: str,
    internet_message_id: str,
    source_file: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    broker_name: str,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    compact = re.sub(r"[ \t]+", " ", text)
    pattern = re.compile(
        r"""
        (?P<symbol>\b[A-Z]{3}\b)\s+
        (?P<qty>[\d,]+\.\d+)\s+
        (?P<price>[\d,]+\.\d+)\s+
        (?P<traded_amount>[\d,]+\.\d+)\s+
        (?P<trade_date>\d{2}/\d{2}/\d{4})\s+
        (?P<settle_date>\d{2}/\d{2}/\d{4})
        """,
        re.VERBOSE,
    )

    for m in pattern.finditer(compact):
        g = m.groupdict()

        side = "BUY"
        qty = parse_decimal(g["qty"])
        price = parse_decimal(g["price"])
        amount = parse_decimal(g["traded_amount"])
        trade_date = parse_date_any(g["trade_date"], prefer_day_first=True)
        value_date = parse_date_any(g["settle_date"], prefer_day_first=True)

        trade = build_trade_dict(
            internet_message_id=internet_message_id,
            source_file=source_file,
            source_type="pdf",
            broker_name=broker_name,
            security_name=g["symbol"],
            isin=g["symbol"],
            side=side,
            trade_date=trade_date,
            value_date=value_date,
            quantity=qty,
            price=price,
            price_currency="EUR",
            consideration=amount,
            commission=None,
            net_amount=amount,
            settlement_terms="FOP",
            counterparty_reference=build_generic_reference(g["symbol"], side, trade_date, value_date, qty, price),
            nominal=qty,
            price_in_percentage=None,
            accrued_interest=None,
            settlement_currency="USD",
            parser_template="STONEX_DAILY_STATEMENT_PDF",
            raw_json=json.dumps(g, default=str),
            processing_run_id=processing_run_id,
            file_id=file_id,
            email_id=email_id,
            side_original_text=side,
            trade_date_original_text=g["trade_date"],
            value_date_original_text=g["settle_date"],
        )
        trade = normalize_trade_signs(trade)
        finalize_trade_validation(trade, email_received_at)
        rows.append(trade)

    logging.info("STONEX daily statement PDF parsed rows=%s", len(rows))
    return rows


# =============================================================================
# EMAIL BODY PARSER FOR REPO
# =============================================================================
def parse_stonex_repo_email_body(
    body_text: str,
    internet_message_id: str,
    sender: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    email_id: Optional[int],
    mapping_by_sender: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    text = clean_text(body_text)
    if not text:
        return []

    broker_name = resolve_broker_name_from_mapping(sender, mapping_by_sender)

    position_id = rx(r"Position\s*ID\s*([A-Za-z0-9\-]+)", text) or rx(r"Position\s*ID\s*[:\-]?\s*([A-Za-z0-9\-]+)", text)
    trade_type = rx(r"Trade\s*Type\s*\(StoneX\)\s*([A-Za-z ]+)", text)
    qty_raw = rx(r"Quantity\s*([0-9,.\-]+)", text)
    price_raw = rx(r"Price\s+in\s+Trade\s+Currency\s+\(including\s+HC\)\s*([0-9,.\-]+)", text)
    traded_amount_raw = rx(r"Traded\s*Amount\s*([0-9,.\-]+)", text)
    isin = rx(r"ISIN\s*([A-Z0-9]{10,20})", text)
    security_name = rx(r"Security\s*Name\s*(.+?)\s+Trade\s*Currency", text)
    trade_currency = rx(r"Trade\s*Currency\s*([A-Z]{3})", text)
    haircut_raw = rx(r"Hair\s*Cut\s*([0-9,.\-]+%)", text)
    trade_date_raw = rx(r"Trade\s*Date\s*([0-9/\-\.]+)", text)
    settle_date_raw = rx(r"Settle\s*Date\s*([0-9/\-\.]+)", text)
    end_date_raw = rx(r"End\s*Date\s*([A-Za-z0-9/\-\.]+)", text)
    benchmark = rx(r"Benchmark\s*(.+?)\s+Rate", text)
    rate_raw = rx(r"Rate\s*([0-9,.\-]+)", text)
    spread_raw = rx(r"Spread\s*([0-9,.\-]+)", text)
    direction_raw = rx(r"Direction\s*(.+?)\s*(?:Comments|$)", text)
    comments = rx(r"Comments\s*(.+)$", text)

    side = normalize_side(direction_raw, "STONEX_REPO_EMAIL")
    qty = parse_decimal(qty_raw)
    price = parse_decimal(price_raw)
    traded_amount = parse_decimal(traded_amount_raw)
    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(settle_date_raw, prefer_day_first=True)

    raw_json = {
        "position_id": position_id,
        "trade_type": trade_type,
        "qty_raw": qty_raw,
        "price_raw": price_raw,
        "traded_amount_raw": traded_amount_raw,
        "isin": isin,
        "security_name": security_name,
        "trade_currency": trade_currency,
        "haircut_raw": haircut_raw,
        "trade_date_raw": trade_date_raw,
        "settle_date_raw": settle_date_raw,
        "end_date_raw": end_date_raw,
        "benchmark": benchmark,
        "rate_raw": rate_raw,
        "spread_raw": spread_raw,
        "direction_raw": direction_raw,
        "comments": comments,
    }

    ref = position_id or build_generic_reference(
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        qty=qty,
        price=price,
        nominal=qty,
    )

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file="EMAIL_BODY",
        source_type="email_body",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=qty,
        price=price,
        price_currency=trade_currency or "USD",
        consideration=traded_amount,
        commission=None,
        net_amount=traded_amount,
        settlement_terms="REPO",
        counterparty_reference=ref,
        nominal=qty,
        price_in_percentage=parse_decimal(haircut_raw),
        accrued_interest=None,
        settlement_currency=trade_currency or "USD",
        parser_template="STONEX_REPO_EMAIL",
        raw_json=json.dumps(raw_json, default=str),
        processing_run_id=processing_run_id,
        file_id=None,
        email_id=email_id,
        side_original_text=direction_raw,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=settle_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)

    return [trade]


def parse_grant_westover_email_body(
    body_text: str,
    internet_message_id: str,
    sender: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    email_id: Optional[int],
    mapping_by_sender: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Grant Westover (grant.westover@stonex.com) sends REPO/settlement emails with fields:
    Security Name, Trading Broker, Amount, ISIN, Transaction Type, Units, CCY,
    Clearing Broker, Trade Date, Settle Date, Loc
    """
    text = clean_text(body_text)
    if not text:
        return []

    broker_name = resolve_broker_name_from_mapping(sender, mapping_by_sender)

    isin = rx(r"ISIN\s*[:\-]?\s*\n?\s*([A-Z]{2}[A-Z0-9]{9,12})", body_text)
    if not isin:
        isin = rx(r"\b([A-Z]{2}[A-Z0-9]{9,12})\b", body_text)

    security_name = rx(r"Security\s+Name\s*[:\-]?\s*\n?\s*(.+)", body_text)
    transaction_type = rx(r"Transaction\s+Type\s*[:\-]?\s*\n?\s*(Buy|Sell|BUY|SELL)", body_text)
    qty_raw = rx(r"Units\s*[:\-]?\s*\n?\s*([0-9,\.]+)", body_text)
    amount_raw = rx(r"Amount\s*[:\-]?\s*\n?\s*\(?([0-9,\.]+)\)?", body_text)
    ccy = rx(r"CCY\s*[:\-]?\s*\n?\s*([A-Z]{3})", body_text) or "USD"
    trade_date_raw = rx(r"Trade\s+Date\s*[:\-]?\s*\n?\s*([0-9/\-\.]+)", body_text)
    settle_date_raw = rx(r"Settle\s+Date\s*[:\-]?\s*\n?\s*([0-9/\-\.]+)", body_text)

    side = normalize_side(transaction_type, "GRANT_WESTOVER_REPO_EMAIL")
    trade_date = parse_date_any(trade_date_raw, prefer_day_first=False)
    value_date = parse_date_any(settle_date_raw, prefer_day_first=False)
    quantity = parse_decimal(qty_raw)
    amount = parse_decimal(amount_raw)

    ref = build_generic_reference(isin, side, trade_date, value_date, quantity, None, quantity)

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file="EMAIL_BODY",
        source_type="email_body",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=quantity,
        price=None,
        price_currency=ccy,
        consideration=amount,
        commission=None,
        net_amount=amount,
        settlement_terms="REPO",
        counterparty_reference=ref,
        nominal=quantity,
        price_in_percentage=None,
        accrued_interest=None,
        settlement_currency=ccy,
        parser_template="GRANT_WESTOVER_REPO_EMAIL",
        raw_json=json.dumps({
            "transaction_type": transaction_type, "isin": isin,
            "security_name": security_name, "qty_raw": qty_raw,
            "amount_raw": amount_raw, "ccy": ccy,
            "trade_date_raw": trade_date_raw, "settle_date_raw": settle_date_raw,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=None,
        email_id=email_id,
        side_original_text=transaction_type,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=settle_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)
    return [trade]


def parse_fab_repo_email_body(
    body_text: str,
    internet_message_id: str,
    sender: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    email_id: Optional[int],
    mapping_by_sender: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    FAB Global Market senders (amna.anwar@bankfab.com, umar.malik@bankfab.com,
    vijuraj.thandalath@bankfab.com).
    Format: "AM Wealth enters Reverse Repo (lends cash/ borrows securities)"
    Fields: ISIN, Description, All in Price, Collateral, Face Amount,
            Start Cash / Cash, Interest, Trade Date, Settlement Date
    Reverse Repo = AM Wealth lends cash, receives (borrows) securities → BUY side.
    """
    text = body_text
    if not text:
        return []

    broker_name = resolve_broker_name_from_mapping(sender, mapping_by_sender)

    # Direction from subject line
    direction_raw = rx(r"(AM\s+Wealth\s+enters\s+Reverse\s+Repo[^\\n]*)", text)
    side = normalize_side(direction_raw or "REVERSE REPO", "FAB_REPO_EMAIL")

    isin = rx(r"ISIN\s*[:\-]?\s*\n?\s*([A-Z]{2}[A-Z0-9]{9,12})", text)
    if not isin:
        isin = rx(r"\b([A-Z]{2}[A-Z0-9]{9,12})\b", text)

    security_name = rx(r"Description\s*[:\-]?\s*\n?\s*(.+)", text)
    price_pct_raw = rx(r"All\s+in\s+Price\s*[:\-]?\s*\n?\s*([0-9,.]+)", text)
    nominal_raw = rx(r"Face\s+Amount\s*[:\-]?\s*\n?\s*([0-9,\.]+)", text)
    start_cash_raw = (
        rx(r"Start\s+Cash\s*[:\-]?\s*\n?\s*(?:USD\s+)?([0-9,\.]+)", text)
        or rx(r"Cash\s*[:\-]?\s*\n?\s*(?:USD\s+)?([0-9,\.]+)", text)
    )
    interest_raw = rx(r"Interest\s*[:\-]?\s*\n?\s*(?:USD\s+)?([0-9,\.]+)", text)
    trade_date_raw = rx(r"Trade\s+Date\s*[:\-]?\s*\n?\s*(.+)", text)
    settle_date_raw = rx(r"Settlement\s+Date\s*[:\-]?\s*\n?\s*(.+)", text)
    rate_raw = (
        rx(r"Fixed\s+Rate\s*[:\-]?\s*\n?\s*([0-9\.]+%?)", text)
        or rx(r"Repo\s+Rate\s*[:\-]?\s*\n?\s*([0-9\.]+%?)", text)
    )

    trade_date = parse_date_any(trade_date_raw, prefer_day_first=True)
    value_date = parse_date_any(settle_date_raw, prefer_day_first=True)
    nominal = parse_decimal(nominal_raw)
    price_pct = parse_decimal(price_pct_raw)
    start_cash = parse_decimal(start_cash_raw)

    ref = build_generic_reference(isin, side, trade_date, value_date, nominal, price_pct, nominal)

    trade = build_trade_dict(
        internet_message_id=internet_message_id,
        source_file="EMAIL_BODY",
        source_type="email_body",
        broker_name=broker_name,
        security_name=security_name,
        isin=isin,
        side=side,
        trade_date=trade_date,
        value_date=value_date,
        quantity=nominal,
        price=None,
        price_currency="USD",
        consideration=start_cash,
        commission=None,
        net_amount=start_cash,
        settlement_terms="REPO",
        counterparty_reference=ref,
        nominal=nominal,
        price_in_percentage=price_pct,
        accrued_interest=parse_decimal(interest_raw),
        settlement_currency="USD",
        parser_template="FAB_REPO_EMAIL",
        raw_json=json.dumps({
            "direction_raw": direction_raw, "isin": isin,
            "security_name": security_name, "price_pct_raw": price_pct_raw,
            "nominal_raw": nominal_raw, "start_cash_raw": start_cash_raw,
            "interest_raw": interest_raw, "trade_date_raw": trade_date_raw,
            "settle_date_raw": settle_date_raw, "rate_raw": rate_raw,
        }, default=str),
        processing_run_id=processing_run_id,
        file_id=None,
        email_id=email_id,
        side_original_text=direction_raw,
        trade_date_original_text=trade_date_raw,
        value_date_original_text=settle_date_raw,
    )
    trade = normalize_trade_signs(trade)
    finalize_trade_validation(trade, email_received_at)
    return [trade]


# =============================================================================
# PDF ROUTER
# =============================================================================
def parse_pdf_file(
    file_bytes: bytes,
    filename: str,
    internet_message_id: str,
    sender: str,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    file_id: Optional[int],
    email_id: Optional[int],
    mapping_by_sender: Dict[str, Dict[str, Any]],
    sender_name: str = "",
) -> List[Dict[str, Any]]:
    text = extract_pdf_text(file_bytes)
    template_code = detect_template_from_mapping(sender, mapping_by_sender)
    broker_name = resolve_broker_name_from_mapping(sender, mapping_by_sender)

    if template_code == "CUB_PDF":
        return parse_cub_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "CAMCAP_PDF":
        return parse_camcap_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "ZARATTINI_PDF":
        return parse_zarattini_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "ASHENDEN_PDF":
        return parse_ashenden_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "STONEX_PDF":
        return parse_stonex_fixed_income_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code in {"BONDPARTNERS_PDF", "BPL_PDF"}:
        return parse_bondpartners_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "SEAPORT_PDF":
        return parse_seaport_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "BRIDPORT_PDF":
        return parse_bridport_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "GTN_XLS_PDF":
        return parse_gtn_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "STONEX_DAILY_STATEMENT_PDF":
        return parse_stonex_daily_statement_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "INSTINET_PDF":
        return parse_instinet_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    if template_code == "ENBD_PDF":
        # ENBD Securities "Order Confirmation Report" (equity, GCM senders @emiratesnbd.com)
        # Only route to securities parser if sender display name contains "GCM"
        has_order_report = "Order Confirmation Report" in text
        has_sell_confo = "Sell Confirmation" in text
        has_buy_confo = "Buy Confirmation" in text
        # GCM filter: block only if sender name is explicitly non-GCM; unknown/empty = allow
        is_gcm = (not sender_name) or ("gcm" in sender_name.lower())
        logging.warning(
            "ENBD_ROUTING sender=%s sender_name=%r file=%s order_report=%s sell_confo=%s buy_confo=%s gcm=%s",
            sender, sender_name, filename, has_order_report, has_sell_confo, has_buy_confo, is_gcm,
        )
        if is_gcm and (has_order_report or has_sell_confo or has_buy_confo):
            return parse_enbd_securities_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)
        return parse_enbd_pdf(text, internet_message_id, filename, email_received_at, processing_run_id, file_id, email_id, broker_name)

    return []


# =============================================================================
# FAB SWIFT MT545 / MT547 SETTLEMENT CONFIRMATION PARSER
# =============================================================================

def _ensure_fab_swift_table(conn) -> None:
    """Create fab_swift_results table if it does not exist yet, and add any new columns."""
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS back_office_auto.fab_swift_results (
                id                      SERIAL PRIMARY KEY,
                email_id                INTEGER,
                source_file             TEXT,
                message_ref             TEXT UNIQUE,
                mt_type                 TEXT,
                isin                    TEXT,
                security_name           TEXT,
                side                    TEXT,
                trade_date              DATE,
                settlement_date         DATE,
                effective_settlement_date DATE,
                face_amount             NUMERIC(20,4),
                settled_amount          NUMERIC(20,4),
                settled_currency        TEXT,
                internal_deal_id        INTEGER,
                match_status            TEXT,
                match_note              TEXT,
                internal_amount         NUMERIC(20,4),
                internal_face_amount    NUMERIC(20,4),
                run_id                  INTEGER,
                created_at              TIMESTAMP WITH TIME ZONE DEFAULT NOW()
            )
        """)
        # Add new columns (idempotent)
        for col_def in [
            "ADD COLUMN IF NOT EXISTS amount_diff         NUMERIC(20,4)",
            "ADD COLUMN IF NOT EXISTS face_amount_match   BOOLEAN",
            "ADD COLUMN IF NOT EXISTS internal_nominal    NUMERIC(20,4)",
            "ADD COLUMN IF NOT EXISTS internal_value_date DATE",
            "ADD COLUMN IF NOT EXISTS linked_mt_type      TEXT",
            "ADD COLUMN IF NOT EXISTS settled_by          TEXT",
            "ADD COLUMN IF NOT EXISTS settled_at          TIMESTAMP WITH TIME ZONE",
        ]:
            cur.execute(f"ALTER TABLE back_office_auto.fab_swift_results {col_def}")
    conn.commit()
    # Add unique constraint on email_id (idempotent via exception handling)
    try:
        with conn.cursor() as cur:
            cur.execute("""
                ALTER TABLE back_office_auto.fab_swift_results
                ADD CONSTRAINT fab_swift_results_email_id_uq UNIQUE (email_id)
            """)
        conn.commit()
    except Exception:
        conn.rollback()


def _upsert_fab_swift_result(conn, result: Dict[str, Any]) -> int:
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO back_office_auto.fab_swift_results
                (email_id, source_file, message_ref, mt_type, isin, security_name,
                 side, trade_date, settlement_date, effective_settlement_date,
                 face_amount, settled_amount, settled_currency, run_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (email_id) DO UPDATE SET
                source_file               = EXCLUDED.source_file,
                message_ref               = EXCLUDED.message_ref,
                mt_type                   = EXCLUDED.mt_type,
                isin                      = EXCLUDED.isin,
                security_name             = EXCLUDED.security_name,
                side                      = EXCLUDED.side,
                trade_date                = EXCLUDED.trade_date,
                settlement_date           = EXCLUDED.settlement_date,
                effective_settlement_date = EXCLUDED.effective_settlement_date,
                face_amount               = EXCLUDED.face_amount,
                settled_amount            = EXCLUDED.settled_amount,
                settled_currency          = EXCLUDED.settled_currency,
                run_id                    = EXCLUDED.run_id
            RETURNING id
        """, (
            result.get("email_id"),
            result.get("source_file"),
            result.get("message_ref"),
            result.get("mt_type"),
            result.get("isin"),
            result.get("security_name"),
            result.get("side"),
            result.get("trade_date"),
            result.get("settlement_date"),
            result.get("effective_settlement_date"),
            result.get("face_amount"),
            result.get("settled_amount"),
            result.get("settled_currency"),
            result.get("run_id"),
        ))
        row = cur.fetchone()
        conn.commit()
        return row[0] if row else None


def parse_fab_swift_pdf(
    text: str,
    filename: str,
    email_id: Optional[int],
    processing_run_id: Optional[int],
) -> Optional[Dict[str, Any]]:
    """Parse FAB SWIFT MT545/MT547 settlement confirmation PDF.
    MT545 = Receive Against Payment = BUY (AM Wealth receives securities)
    MT547 = Deliver Against Payment = SELL (AM Wealth delivers securities)
    """
    # Detect MT type from text
    mt_type = None
    if re.search(r"MT547", text):
        mt_type = "MT547"
    elif re.search(r"MT545", text):
        mt_type = "MT545"
    if not mt_type:
        return None

    # Log first 500 chars around :35B: for debugging ISIN extraction
    _debug_m = re.search(r".{0,100}:35B:.{0,300}", text, re.DOTALL)
    logging.warning("FAB_SWIFT_DEBUG file=%s | 35B_context=%r", filename, _debug_m.group(0) if _debug_m else "NOT_FOUND")

    side = "BUY" if mt_type == "MT545" else "SELL"

    # Message reference: :20C::SEME//<number> (e.g. "//2026040200217623")
    message_ref = rx(r":20C::SEME//(\S+)", text)
    if not message_ref:
        message_ref = rx(r":20C::SEME\s+//(\S+)", text)

    # Dates: e.g. ":98A::TRAD Trade Date/Time 2026-02-18"
    trade_date_raw = rx(r":98A::TRAD\s+[^\n]*?(\d{4}-\d{2}-\d{2})", text)
    settlement_date_raw = rx(r":98A::SETT\s+[^\n]*?(\d{4}-\d{2}-\d{2})", text)
    effective_date_raw = rx(r":98A::ESET\s+[^\n]*?(\d{4}-\d{2}-\d{2})", text)

    # In FAB SWIFT PDFs, the ISIN appears on the line BEFORE ":35B:"
    # Pattern: "{ISIN}\n:35B: Identification of the Financial Instrument"
    # Use known ISO country prefixes to avoid Bloomberg IDs (BBG...) and name fragments
    _ISIN_PREFIXES = (
        "US", "XS", "DE", "GB", "IE", "FR", "NL", "CH", "IT", "ES",
        "AU", "CA", "JP", "HK", "SG", "SE", "NO", "DK", "FI", "AT",
        "BE", "LU", "PT", "GR", "CZ", "PL", "HU", "TR", "ZA", "IN",
        "CN", "KR", "TW", "MX", "BR", "AR", "CL", "CO", "AE", "SA",
        "QA", "KW", "BH", "OM", "JO", "EG", "NG", "KY", "VG", "BM",
    )
    _isin_alt = "|".join(_ISIN_PREFIXES)
    # ISIN is on the line immediately before :35B: (case-sensitive)
    _isin_pat = re.compile(r"\b((?:" + _isin_alt + r")[A-Z0-9]{10})\b\s*\n\s*:35B:")
    _m = _isin_pat.search(text)
    isin = _m.group(1) if _m else None
    # Fallback 1: ISIN anywhere before :35B: within 3 lines
    if not isin:
        _isin_pat2 = re.compile(r"\b((?:" + _isin_alt + r")[A-Z0-9]{10})\b(?:[^\n]*\n){0,3}[^\n]*:35B:")
        _m2 = _isin_pat2.search(text)
        isin = _m2.group(1) if _m2 else None
    # Fallback 2: ISIN on the same line AFTER :35B: (some PDFs put it there)
    if not isin:
        _isin_pat3 = re.compile(r":35B:[^\n]*\b((?:" + _isin_alt + r")[A-Z0-9]{10})\b")
        _m3 = _isin_pat3.search(text)
        isin = _m3.group(1) if _m3 else None

    # Security name: on the line(s) after ":35B: Identification..." (skip /TS/BBG... lines)
    security_name = None
    m35 = re.search(r":35B:[^\n]*\n((?:/[^\n]*\n)*)([^\n/:16]+)", text)
    if m35:
        security_name = m35.group(2).strip() or None

    # Face amount / Quantity: ":36B::ESTT ... Face Amount 347000," or "Unit Number 3268,"
    face_amount_raw = rx(r":36B::ESTT[^\n]*Face Amount\s+([\d,]+)", text)
    if not face_amount_raw:
        face_amount_raw = rx(r":36B::ESTT[^\n]*(?:Unit Number|Quantity)[^\d]*([\d,]+)", text)
    if not face_amount_raw:
        # Generic fallback: any number after :36B::ESTT
        face_amount_raw = rx(r":36B::ESTT[^\n]*([\d][\d,]*)", text)
    if face_amount_raw:
        face_amount_raw = face_amount_raw.rstrip(",").replace(",", "")
    face_amount = parse_decimal(face_amount_raw)

    # Settled amount and currency: ":19A::ESTT Settled Amount USD 479643,20"
    settled_currency = rx(r":19A::ESTT[^\n]*Settled Amount\s+([A-Z]{3})", text)
    settled_amount_raw = rx(r":19A::ESTT[^\n]*Settled Amount\s+[A-Z]{3}\s+([\d,]+(?:\.\d+)?)", text)
    settled_amount = None
    if settled_amount_raw:
        # Strip trailing comma (e.g. "670751,") before decimal detection
        settled_amount_raw = settled_amount_raw.rstrip(",")
        # European decimal comma: "479643,20" → "479643.20"; thousand separator: "1,234,567" → remove commas
        if re.search(r",\d{1,2}$", settled_amount_raw):
            settled_amount = parse_decimal(settled_amount_raw.replace(",", "."))
        else:
            settled_amount = parse_decimal(settled_amount_raw.replace(",", ""))

    return {
        "email_id": email_id,
        "source_file": filename,
        "message_ref": message_ref,
        "mt_type": mt_type,
        "isin": isin,
        "security_name": security_name,
        "side": side,
        "trade_date": parse_date_any(trade_date_raw),
        "settlement_date": parse_date_any(settlement_date_raw),
        "effective_settlement_date": parse_date_any(effective_date_raw),
        "face_amount": face_amount,
        "settled_amount": settled_amount,
        "settled_currency": settled_currency or "USD",
        "run_id": processing_run_id,
    }


def _process_fab_swift_message(
    conn,
    token: str,
    mailbox: str,
    msg: Dict[str, Any],
    internet_message_id: str,
    subject: str,
    received_at,
    processing_run_id: int,
) -> Tuple[str, int]:
    """Handle FAB SWIFT MT545/MT547 emails — parse PDFs and store to fab_swift_results."""
    _ensure_fab_swift_table(conn)

    if email_already_processed(conn, internet_message_id):
        return ("ALREADY_PROCESSED", 0)

    message_id = msg["id"]
    attachments = get_message_attachments(token, mailbox, message_id)

    email_id = insert_settlement_email(
        conn=conn,
        internet_message_id=internet_message_id,
        message_id=message_id,
        sender="noreply@bankfab.com",
        subject=subject,
        received_at=received_at,
        status="RECEIVED",
        note="FAB SWIFT MT545/MT547 received",
        mailbox=mailbox,
        attachment_count=len(attachments),
        parsed_trade_count=0,
        processing_run_id=processing_run_id,
    )

    parsed_count = 0
    for att in attachments:
        if att.get("@odata.type") != "#microsoft.graph.fileAttachment":
            continue
        filename = att.get("name") or "unnamed"
        if not filename.lower().endswith(".pdf"):
            continue
        content_b64 = att.get("contentBytes")
        if not content_b64:
            continue
        file_bytes = base64.b64decode(content_b64)
        text = extract_pdf_text(file_bytes)
        result = parse_fab_swift_pdf(text, filename, email_id, processing_run_id)
        if result and result.get("isin"):
            _upsert_fab_swift_result(conn, result)
            parsed_count += 1

    status = "PARSED" if parsed_count > 0 else "NO_TRADES_FOUND"
    insert_settlement_email(
        conn=conn,
        internet_message_id=internet_message_id,
        message_id=message_id,
        sender="noreply@bankfab.com",
        subject=subject,
        received_at=received_at,
        status=status,
        note=f"FAB SWIFT parsed: {parsed_count}",
        mailbox=mailbox,
        attachment_count=len(attachments),
        parsed_trade_count=parsed_count,
        processing_run_id=processing_run_id,
    )
    return (status, parsed_count)


def run_fab_swift_reconciliation(conn, run_id: Optional[int] = None) -> List[Dict[str, Any]]:
    """Match fab_swift_results against tab_deals (status IN (2,6) = INSTRUCTED or FAILED)
    by ISIN + side + settlement_date. Fallback: match without date if primary fails.
    Compare settled_amount vs net_amount (tolerance ±5 USD) and face_amount vs nominal.
    """
    _ensure_fab_swift_table(conn)

    # Load FAB SWIFT records from last 30 days (or with NULL settlement_date)
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT f.id, f.message_ref, f.mt_type, f.isin, f.security_name, f.side,
                   f.trade_date, f.settlement_date, f.effective_settlement_date,
                   f.face_amount, f.settled_amount, f.settled_currency, f.email_id
            FROM back_office_auto.fab_swift_results f
            LEFT JOIN back_office.tab_deals d ON d.id = f.internal_deal_id
            WHERE f.settled_at IS NULL
              AND (f.settlement_date >= CURRENT_DATE - INTERVAL '30 days'
                   OR f.settlement_date IS NULL)
              AND (d.status IS NULL OR d.status NOT IN (4, 7))
            ORDER BY f.settlement_date DESC NULLS LAST, f.id DESC
        """)
        swift_rows = [dict(r) for r in cur.fetchall()]

    def _find_candidates(isin, action_val, sett_date=None):
        """Query tab_deals; if sett_date given use date filter, else no date filter."""
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if sett_date:
                cur.execute("""
                    SELECT td.id, td.symbol, td.qty, td.nominal, td.transaction_value,
                           td.action, td.status, td.login, td.settle_date_cash, td.value_date_cash,
                           td.value_date_securities, td.settle_date_securities,
                           td.currency_pay,
                           td.transaction_value + (
                               CASE WHEN td.action = 0 THEN -td.execution_cost
                               ELSE td.execution_cost END
                           ) AS net_amount,
                           cp.name AS counterparty
                    FROM back_office.tab_deals td
                    LEFT JOIN back_office.tab_counterparty cp ON td.counterparty_id = cp.id
                    WHERE td.symbol = %s
                      AND td.action = %s
                      AND td.status IN (2, 4, 6)
                      AND td.settle_type = 'external'
                      AND td.reason = 0
                      AND td.login IN (1, 5)
                      AND (
                          td.settle_date_cash = %s
                          OR td.value_date_cash = %s
                          OR td.value_date_securities = %s
                          OR td.settle_date_securities = %s
                      )
                    ORDER BY td.id DESC
                    LIMIT 5
                """, (isin, action_val, sett_date, sett_date, sett_date, sett_date))
            else:
                cur.execute("""
                    SELECT td.id, td.symbol, td.qty, td.nominal, td.transaction_value,
                           td.action, td.status, td.login, td.settle_date_cash, td.value_date_cash,
                           td.value_date_securities, td.settle_date_securities,
                           td.currency_pay,
                           td.transaction_value + (
                               CASE WHEN td.action = 0 THEN -td.execution_cost
                               ELSE td.execution_cost END
                           ) AS net_amount,
                           cp.name AS counterparty
                    FROM back_office.tab_deals td
                    LEFT JOIN back_office.tab_counterparty cp ON td.counterparty_id = cp.id
                    WHERE td.symbol = %s
                      AND td.action = %s
                      AND td.status IN (2, 4, 6)
                      AND td.settle_type = 'external'
                      AND td.reason = 0
                      AND td.login IN (1, 5)
                    ORDER BY td.id DESC
                    LIMIT 5
                """, (isin, action_val))
            return [dict(r) for r in cur.fetchall()]

    def _resolve_instruction_amount(deal_id, isin_val):
        """For CMF (login=5) deals, resolve the actual netted amount from tab_instructions."""
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT i.id AS instruction_id,
                       i.net_settlement_amount,
                       i.quantity,
                       i.value_date AS instr_value_date
                FROM back_office.tab_connect_deal_transfer cdt
                JOIN back_office.tab_transfer tf ON tf.id = cdt.id_transfer
                JOIN back_office.tab_settlements ts ON ts.id = tf.id_settlement
                JOIN back_office.tab_instructions i ON i.id_amwl = ts.id::text
                WHERE cdt.id_deal = %s AND i.isin = %s
                ORDER BY i.id DESC
                LIMIT 1
            """, (deal_id, isin_val))
            row = cur.fetchone()
            return dict(row) if row else None

    # Pre-load set of (ISIN, action) pairs that have at least one deal with login IN (1,5)
    # Used to skip NOT_FOUND rows where no FAB/CMF deals exist for that side
    _fab_isin_actions = set()
    with conn.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT symbol, action FROM back_office.tab_deals
            WHERE login IN (1, 5) AND reason = 0 AND settle_type = 'external'
        """)
        _fab_isin_actions = {(r[0], r[1]) for r in cur.fetchall()}

    results = []
    for sw in swift_rows:
        isin = (sw.get("isin") or "").upper().strip()
        side = (sw.get("side") or "").upper()
        sett_date = sw.get("settlement_date")
        action_val = 0 if side == "BUY" else 1

        # Primary: match with date
        candidates = _find_candidates(isin, action_val, sett_date) if sett_date else []
        date_matched = bool(candidates)

        # Fallback: match without date
        if not candidates:
            candidates = _find_candidates(isin, action_val, sett_date=None)

        sw_amount = sw.get("settled_amount")
        sw_face = sw.get("face_amount")

        if not candidates:
            # Skip ISIN+side combos that have no FAB/CMF deals (login 1 or 5)
            if (isin, action_val) not in _fab_isin_actions:
                continue
            row = dict(sw)
            row["match_status"] = "NOT_FOUND"
            row["match_note"] = "No INSTRUCTED/FAILED deal found for ISIN+side"
            row["internal_deal_id"] = None
            row["internal_amount"] = None
            row["internal_face_amount"] = None
            row["amount_diff"] = None
            row["face_amount_match"] = None
            row["counterparty"] = None
        else:
            best = candidates[0]

            # For CMF trades (login=5), resolve amount from tab_instructions
            if best.get("login") == 5:
                instr = _resolve_instruction_amount(best["id"], isin)
                if instr and instr.get("net_settlement_amount") is not None:
                    int_amount = instr["net_settlement_amount"]
                    int_face = instr.get("quantity") or best.get("nominal") or best.get("qty")
                else:
                    # Fallback: use deal amount if no instruction found
                    int_amount = best.get("net_amount") or best.get("transaction_value")
                    int_face = best.get("nominal") or best.get("qty")
            else:
                int_amount = best.get("net_amount") or best.get("transaction_value")
                int_face = best.get("nominal") or best.get("qty")

            int_value_date = best.get("settle_date_cash") or best.get("value_date_cash")

            # Amount difference
            amount_diff = None
            if sw_amount is not None and int_amount is not None:
                try:
                    amount_diff = round(float(sw_amount) - float(int_amount), 4)
                except Exception:
                    pass

            # Amount match: tolerance ±1 USD
            amount_ok = (sw_amount is not None and int_amount is not None
                         and values_equal_decimal(sw_amount, int_amount, Decimal("1")))

            # Face amount match: skip if PDF has no face amount (equities)
            face_ok = (sw_face is None  # no face in PDF — skip check
                       or (int_face is not None
                           and values_equal_decimal(sw_face, int_face, Decimal("0.0001"))))

            cmf_tag = " [CMF: matched via instruction]" if best.get("login") == 5 else ""

            if not date_matched:
                match_status = "DATE_MISMATCH"
                diff_str = f" Δ={amount_diff:+.2f}" if amount_diff is not None else ""
                match_note = f"Fallback match (date {sett_date} not found); settled={sw_amount} vs internal={int_amount}{diff_str}{cmf_tag}"
            elif amount_ok:
                match_status = "MATCHED"
                match_note = (None if face_ok else f"face_amount mismatch: pdf={sw_face} vs system={int_face}") if not cmf_tag else f"amount OK{cmf_tag}"
            else:
                match_status = "AMOUNT_MISMATCH"
                diff_str = f" Δ={amount_diff:+.2f}" if amount_diff is not None else ""
                match_note = f"settled={sw_amount} vs internal={int_amount}{diff_str}{cmf_tag}"

            row = dict(sw)
            row["match_status"] = match_status
            row["match_note"] = match_note
            row["internal_deal_id"] = best.get("id")
            row["internal_amount"] = int_amount
            row["internal_face_amount"] = int_face
            row["amount_diff"] = amount_diff
            row["face_amount_match"] = face_ok
            row["counterparty"] = best.get("counterparty")
            row["internal_value_date"] = int_value_date
            row["_deal_login"] = best.get("login")

            # Update match result in DB
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE back_office_auto.fab_swift_results
                    SET match_status = %s, match_note = %s,
                        internal_deal_id = %s, internal_amount = %s,
                        internal_face_amount = %s, run_id = %s,
                        amount_diff = %s, face_amount_match = %s,
                        internal_nominal = %s, internal_value_date = %s
                    WHERE id = %s
                """, (match_status, match_note, best.get("id"),
                      int_amount, int_face, run_id,
                      amount_diff, face_ok, int_face, int_value_date,
                      sw["id"]))
            conn.commit()

        results.append(row)

    return results


# =============================================================================
# ATTACHMENT PROCESSING
# =============================================================================
def parse_single_attachment(
    conn,
    internet_message_id: str,
    sender: str,
    filename: str,
    file_bytes: bytes,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    email_id: Optional[int],
    attachment_order: Optional[int] = None,
    parent_zip_file_name: Optional[str] = None,
    mapping_by_sender: Optional[Dict[str, Dict[str, Any]]] = None,
    sender_name: str = "",
) -> int:
    parsed_count = 0
    seen_keys = set()
    file_hash = sha256_bytes(file_bytes)
    file_type = infer_file_type(filename)
    mapping_by_sender = mapping_by_sender or {}

    # Skip if same file content was already parsed from another mailbox (forwarded copy)
    with conn.cursor() as _cur:
        _cur.execute(
            """
            SELECT 1 FROM back_office_auto.settlement_files sf
            WHERE sf.file_hash = %s AND sf.parse_status = 'PARSED'
            LIMIT 1
            """,
            (file_hash,),
        )
        if _cur.fetchone() is not None:
            logging.info("Skipping duplicate file (same hash already parsed): %s", filename)
            return 0

    file_id = insert_settlement_file(
        conn=conn,
        internet_message_id=internet_message_id,
        file_name=filename,
        file_type=file_type,
        file_hash=file_hash,
        attachment_size=len(file_bytes),
        attachment_order=attachment_order,
        parent_zip_file_name=parent_zip_file_name,
        parse_status="RECEIVED",
        parse_note=None,
    )

    template_code = detect_template_from_mapping(sender, mapping_by_sender)

    if file_type in {"xlsx", "xlsm", "xls"}:
        dfs = extract_excel_sheets(file_bytes, filename)
        logging.info("Excel file %s produced %s dataframes", filename, len(dfs))

        for df in dfs:
            trades: List[Dict[str, Any]] = []

            if template_code == "INSTINET_XLSM":
                trades = parse_instinet_excel(
                    df=df,
                    internet_message_id=internet_message_id,
                    source_file=filename,
                    sender=sender,
                    email_received_at=email_received_at,
                    processing_run_id=processing_run_id,
                    file_id=file_id,
                    email_id=email_id,
                    mapping_by_sender=mapping_by_sender,
                )
            elif template_code == "GTN_XLS_PDF":
                trades = parse_gtn_excel(
                    df=df,
                    internet_message_id=internet_message_id,
                    source_file=filename,
                    sender=sender,
                    email_received_at=email_received_at,
                    processing_run_id=processing_run_id,
                    file_id=file_id,
                    email_id=email_id,
                    mapping_by_sender=mapping_by_sender,
                )

            logging.info("Excel file %s parsed trades=%s", filename, len(trades))

            for trade in trades:
                dedup_key = trade_dedup_key(trade)
                if dedup_key in seen_keys:
                    logging.info("Skipping duplicate parsed trade: %s", dedup_key)
                    continue
                seen_keys.add(dedup_key)

                trade_id = upsert_settlement_trade(conn, trade)
                parsed_count += 1
                try:
                    enrich_cpty_ssi(conn, trade_id, trade.get("broker_name", ""), trade.get("raw_json"))
                except Exception as _e:
                    logging.warning("enrich_cpty_ssi skipped for trade %s: %s", trade_id, _e)

    elif file_type == "pdf":
        trades = parse_pdf_file(
            file_bytes=file_bytes,
            filename=filename,
            internet_message_id=internet_message_id,
            sender=sender,
            sender_name=sender_name,
            email_received_at=email_received_at,
            processing_run_id=processing_run_id,
            file_id=file_id,
            email_id=email_id,
            mapping_by_sender=mapping_by_sender,
        )
        logging.info("PDF file %s parsed trades=%s", filename, len(trades))

        for trade in trades:
            dedup_key = trade_dedup_key(trade)
            if dedup_key in seen_keys:
                logging.info("Skipping duplicate parsed trade: %s", dedup_key)
                continue
            seen_keys.add(dedup_key)

            trade_id = upsert_settlement_trade(conn, trade)
            parsed_count += 1
            try:
                enrich_cpty_ssi(conn, trade_id, trade.get("broker_name", ""), trade.get("raw_json"))
            except Exception as _e:
                logging.warning("enrich_cpty_ssi skipped for trade %s: %s", trade_id, _e)

    elif file_type == "zip":
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
                members = [m for m in zf.infolist() if not m.is_dir()]
                logging.info("ZIP file %s contains %s files", filename, len(members))

                for idx, member in enumerate(members, start=1):
                    inner_name = member.filename
                    inner_bytes = zf.read(member)
                    normalized_inner_name = inner_name.split("/")[-1].split("\\")[-1]

                    parsed_count += parse_single_attachment(
                        conn=conn,
                        internet_message_id=internet_message_id,
                        sender=sender,
                        sender_name=sender_name,
                        filename=normalized_inner_name,
                        file_bytes=inner_bytes,
                        email_received_at=email_received_at,
                        processing_run_id=processing_run_id,
                        email_id=email_id,
                        attachment_order=idx,
                        parent_zip_file_name=filename,
                        mapping_by_sender=mapping_by_sender,
                    )
        except zipfile.BadZipFile:
            logging.exception("Bad ZIP file: %s", filename)
        except Exception as e:
            logging.exception("Failed processing ZIP file %s: %s", filename, e)

    with conn.cursor() as cur:
        cur.execute(
            """
            update back_office_auto.settlement_files
               set parse_status = %s,
                   parse_note = %s
             where id = %s
            """,
            (
                "PARSED" if parsed_count > 0 else "NO_TRADES_FOUND",
                f"parsed_trades={parsed_count}",
                file_id,
            ),
        )
    conn.commit()

    return parsed_count


# =============================================================================
# CPTY SSI ENRICHMENT  (separate function — does not affect parsing logic)
# =============================================================================

# Normalize custodian keywords found in PDF/Excel text → canonical form used in counterparty_ssi_mapping
_CUSTODIAN_ALIASES: Dict[str, str] = {
    "euroclear": "EUROCLEAR",
    "eclr":      "EUROCLEAR",
    "clearstream": "CLEARSTREAM",
    "cedel":     "CLEARSTREAM",
    "cede":      "CLEARSTREAM",
    "dtc":       "DTC",
    "crest":     "CREST",
    "fed":       "FEDWIRE",
    "fedwire":   "FEDWIRE",
    "hkscc":     "HKSCC",
    "ccas":      "HKSCC",
    "mshk":      "HKSCC",
    "six":       "SIX SIS",
    "sis":       "SIX SIS",
}

# Regex patterns that extract (custodian_keyword, account) from raw text in broker files
# Examples: "DTC 0067", "ECLR 75663", "Account number : 90439" (with EUROCLEAR context),
#           "OUR SSI : ECLR 75663", "CEDE 83320", "Account: ... (ISNTUS33)\nDTC 0067"
_SSI_PATTERNS = [
    # "OUR SSI : ECLR 75663"  or  "ECLR 75663"
    re.compile(r"\b(ECLR|EUROCLEAR|CLEARSTREAM|CEDEL|CEDE|DTC|CREST|FED|HKSCC|CCAS|MSHK|SIX|SIS)\s*[:\-]?\s*([A-Z0-9/]{3,})", re.IGNORECASE),
    # "Account number : 90439"  (custodian detected separately from context)
    re.compile(r"account\s+(?:number|no\.?)\s*[:\-]?\s*([A-Z0-9/]{3,})", re.IGNORECASE),
    # "Account : Instinet, LLC (ISNTUS33)  DTC 0067"
    re.compile(r"\bDTC\s+(\d{4,})", re.IGNORECASE),
]


def _extract_ssi_hints(text: str) -> List[Dict[str, str]]:
    """
    Extract (custodian, account) pairs from raw text.
    Returns list of dicts with keys 'custodian' and 'account'.
    """
    hints: List[Dict[str, str]] = []
    text_upper = text.upper()

    # Pattern 1: "CUSTODIAN ACCOUNT" pairs
    for m in _SSI_PATTERNS[0].finditer(text):
        kw = m.group(1).lower()
        acc = m.group(2).strip().strip("/").strip("-")
        cust = _CUSTODIAN_ALIASES.get(kw)
        if cust and acc:
            hints.append({"custodian": cust, "account": acc})

    # Pattern 2: bare "DTC XXXX"
    for m in _SSI_PATTERNS[2].finditer(text):
        acc = m.group(1).strip()
        hints.append({"custodian": "DTC", "account": acc})

    # Deduplicate
    seen = set()
    unique: List[Dict[str, str]] = []
    for h in hints:
        key = (h["custodian"], h["account"])
        if key not in seen:
            seen.add(key)
            unique.append(h)
    return unique


# Broker name → fragment of counterparty name in tab_counterparty
# Used when broker_name in settlement_trades doesn't match tab_counterparty directly
_BROKER_NAME_ALIASES: Dict[str, str] = {
    "instinet": "market securities",
}

# instr_type (from Instinet raw_json) → keyword present in ssi_name
_INSTR_TYPE_SSI_KEYWORD: Dict[str, str] = {
    "USE": "DTC",
    "HKE": "HK",
    "LSE": "CREST",
    "EUE": "ECLR",
    "FRE": "FRANCE",
    "GRE": "GERMANY",
    "SWE": "SWISS",
}


def enrich_cpty_ssi(
    conn,
    trade_id: int,
    broker_name: str,
    raw_json_str: Optional[str],
) -> Optional[str]:
    """
    Look up CPTY SSI for a settlement trade using counterparty_ssi_mapping.

    Table structure:
      back_office_auto.counterparty_ssi_mapping (counterparty_id, ssi_id, is_active)
      → joined with back_office.tab_counterparty (id, name, short_name)
      → joined with back_office.tab_standard_settlement_instructions (id, ssi_name, ac, agent_id)
      → joined with back_office.tab_counterparty as agent (id=agent_id, short_name=custodian)

    Strategy:
      1. Extract (custodian_keyword, account) hints from raw_json text.
      2. Find counterparty_id by matching broker_name to tab_counterparty.name
         (with alias fallback via _BROKER_NAME_ALIASES).
      3. a. Match by counterparty_id + account            (best: right broker + right account)
         b. Match by counterparty_id + instr_type keyword (Instinet: USE→DTC, HKE→HK, etc.)
         c. Match by counterparty_id only (single SSI)   (fallback: only one SSI for this broker)
         d. Match by account across all counterparties    (last resort)
      4. If found, UPDATE settlement_trades.our_ssi.

    Returns the matched ssi_name or None.
    """
    if not raw_json_str:
        return None

    # Flatten raw_json to a single searchable string; also keep parsed object
    try:
        raw_obj = json.loads(raw_json_str)
        text = " ".join(str(v) for v in raw_obj.values() if v is not None)
    except Exception:
        raw_obj = {}
        text = raw_json_str

    hints = _extract_ssi_hints(text)
    accounts = [h["account"] for h in hints]

    # Also include explicit our_account_raw if present (e.g. CUB PDF: "Our EC 23860")
    if raw_obj.get("our_account_raw"):
        our_acc = str(raw_obj["our_account_raw"]).strip()
        if our_acc and our_acc not in accounts:
            accounts.append(our_acc)

    # instr_type keyword for market-based SSI matching (e.g. Instinet USE→DTC, HKE→HK)
    instr_type = str(raw_obj.get("instr_type") or "").upper()
    instr_ssi_keyword = _INSTR_TYPE_SSI_KEYWORD.get(instr_type)

    # Currency-based SSI keyword fallback (when instr_type not available, e.g. PDF parsers)
    # Only used if instr_type didn't yield a keyword
    _CURRENCY_SSI_KEYWORD: Dict[str, str] = {
        "USD": "DTC",
        "GBP": "CREST",
        "HKD": "HK",
    }
    currency = str(raw_obj.get("settl_currency") or raw_obj.get("currency") or "").upper()
    currency_ssi_keyword = _CURRENCY_SSI_KEYWORD.get(currency) if not instr_ssi_keyword else None

    ssi_name: Optional[str] = None

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:

        # Find counterparty_id by broker_name (match on name or short_name)
        # Fallback: use _BROKER_NAME_ALIASES for brokers whose file name differs from DB name
        counterparty_id: Optional[int] = None
        if broker_name:
            cur.execute(
                """
                SELECT id FROM back_office.tab_counterparty
                WHERE LOWER(name) = LOWER(%s) OR LOWER(short_name) = LOWER(%s)
                LIMIT 1
                """,
                (broker_name, broker_name),
            )
            cp_row = cur.fetchone()
            if cp_row:
                counterparty_id = cp_row["id"]

        # Alias fallback: broker_name not found directly in tab_counterparty
        if not counterparty_id and broker_name:
            alias_fragment = _BROKER_NAME_ALIASES.get(broker_name.lower())
            if alias_fragment:
                cur.execute(
                    """
                    SELECT id FROM back_office.tab_counterparty
                    WHERE LOWER(name) LIKE %s
                      AND role = 'Counterparty'
                    LIMIT 1
                    """,
                    (f"%{alias_fragment}%",),
                )
                cp_row = cur.fetchone()
                if cp_row:
                    counterparty_id = cp_row["id"]

        # ILIKE fallback: broker_name is a substring of the full counterparty name
        # (e.g. "Banca Zarattini" matches "Banca Zarattini & Co. SA")
        if not counterparty_id and broker_name and len(broker_name) >= 5:
            cur.execute(
                """
                SELECT id FROM back_office.tab_counterparty
                WHERE LOWER(name) LIKE LOWER(%s)
                  AND role = 'Counterparty'
                LIMIT 1
                """,
                (f"%{broker_name}%",),
            )
            cp_row = cur.fetchone()
            if cp_row:
                counterparty_id = cp_row["id"]

        # ── a. counterparty_id + account (ac or agent_ac) ────────────────────
        if not ssi_name and counterparty_id and accounts:
            accs_upper = [a.upper() for a in accounts]
            cur.execute(
                """
                SELECT ti.ssi_name
                FROM back_office_auto.counterparty_ssi_mapping csm
                JOIN back_office.tab_standard_settlement_instructions ti ON csm.ssi_id = ti.id
                WHERE csm.counterparty_id = %s
                  AND csm.is_active = true
                  AND (UPPER(ti.ac) = ANY(%s) OR UPPER(ti.agent_ac) = ANY(%s))
                LIMIT 1
                """,
                (counterparty_id, accs_upper, accs_upper),
            )
            row = cur.fetchone()
            if row:
                ssi_name = row["ssi_name"]

        # ── b. counterparty_id + market keyword (instr_type or currency) ────────
        # instr_type takes priority (Instinet: USE→DTC, HKE→HK); currency is fallback
        market_keyword = instr_ssi_keyword or currency_ssi_keyword
        if not ssi_name and counterparty_id and market_keyword:
            cur.execute(
                """
                SELECT ti.ssi_name
                FROM back_office_auto.counterparty_ssi_mapping csm
                JOIN back_office.tab_standard_settlement_instructions ti ON csm.ssi_id = ti.id
                WHERE csm.counterparty_id = %s
                  AND csm.is_active = true
                  AND ti.ssi_name ILIKE %s
                LIMIT 1
                """,
                (counterparty_id, f"%{market_keyword}%"),
            )
            row = cur.fetchone()
            if row:
                ssi_name = row["ssi_name"]

        # ── c. counterparty_id only — single SSI ─────────────────────────────
        if not ssi_name and counterparty_id:
            cur.execute(
                """
                SELECT ti.ssi_name
                FROM back_office_auto.counterparty_ssi_mapping csm
                JOIN back_office.tab_standard_settlement_instructions ti ON csm.ssi_id = ti.id
                WHERE csm.counterparty_id = %s
                  AND csm.is_active = true
                """,
                (counterparty_id,),
            )
            rows = cur.fetchall()
            if len(rows) == 1:
                ssi_name = rows[0]["ssi_name"]

        # ── d. account across all counterparties ──────────────────────────────
        if not ssi_name and accounts:
            cur.execute(
                """
                SELECT ti.ssi_name
                FROM back_office_auto.counterparty_ssi_mapping csm
                JOIN back_office.tab_standard_settlement_instructions ti ON csm.ssi_id = ti.id
                JOIN back_office.tab_counterparty tc ON csm.counterparty_id = tc.id
                WHERE csm.is_active = true
                  AND UPPER(ti.ac) = ANY(%s)
                  AND tc.role = 'Counterparty'
                LIMIT 1
                """,
                ([a.upper() for a in accounts],),
            )
            row = cur.fetchone()
            if row:
                ssi_name = row["ssi_name"]

    if ssi_name:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE back_office_auto.settlement_trades SET our_ssi = %s WHERE id = %s",
                (ssi_name, trade_id),
            )
        conn.commit()
        logging.info("enrich_cpty_ssi: trade %s → %s", trade_id, ssi_name)

    return ssi_name


# =============================================================================
# DEBUG / TEST PARSING WITHOUT WRITING
# =============================================================================
def parse_single_attachment_dry_run(
    internet_message_id: str,
    sender: str,
    filename: str,
    file_bytes: bytes,
    email_received_at: Optional[datetime],
    processing_run_id: Optional[int],
    mapping_by_sender: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    parsed_count = 0
    seen_keys = set()
    file_type = infer_file_type(filename)
    mapping_by_sender = mapping_by_sender or {}
    template_code = detect_template_from_mapping(sender, mapping_by_sender)

    result = {
        "file_name": filename,
        "file_type": file_type,
        "template_code": template_code,
        "parsed_trade_count": 0,
        "parsed_trades": [],
        "status": "NO_TRADES_FOUND",
        "error": None,
    }

    try:
        if file_type in {"xlsx", "xlsm", "xls"}:
            dfs = extract_excel_sheets(file_bytes, filename)
            all_trades = []

            for df in dfs:
                trades: List[Dict[str, Any]] = []

                if template_code == "INSTINET_XLSM":
                    trades = parse_instinet_excel(
                        df=df,
                        internet_message_id=internet_message_id,
                        source_file=filename,
                        sender=sender,
                        email_received_at=email_received_at,
                        processing_run_id=processing_run_id,
                        file_id=None,
                        email_id=None,
                        mapping_by_sender=mapping_by_sender,
                    )
                elif template_code == "GTN_XLS_PDF":
                    trades = parse_gtn_excel(
                        df=df,
                        internet_message_id=internet_message_id,
                        source_file=filename,
                        sender=sender,
                        email_received_at=email_received_at,
                        processing_run_id=processing_run_id,
                        file_id=None,
                        email_id=None,
                        mapping_by_sender=mapping_by_sender,
                    )

                for trade in trades:
                    dedup_key = trade_dedup_key(trade)
                    if dedup_key in seen_keys:
                        continue
                    seen_keys.add(dedup_key)
                    parsed_count += 1
                    all_trades.append(trade)

            result["parsed_trades"] = all_trades

        elif file_type == "pdf":
            trades = parse_pdf_file(
                file_bytes=file_bytes,
                filename=filename,
                internet_message_id=internet_message_id,
                sender=sender,
                email_received_at=email_received_at,
                processing_run_id=processing_run_id,
                file_id=None,
                email_id=None,
                mapping_by_sender=mapping_by_sender,
            )

            all_trades = []
            for trade in trades:
                dedup_key = trade_dedup_key(trade)
                if dedup_key in seen_keys:
                    continue
                seen_keys.add(dedup_key)
                parsed_count += 1
                all_trades.append(trade)

            result["parsed_trades"] = all_trades

        elif file_type == "zip":
            zip_results = []
            with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
                members = [m for m in zf.infolist() if not m.is_dir()]
                for member in members:
                    inner_name = member.filename
                    inner_bytes = zf.read(member)
                    normalized_inner_name = inner_name.split("/")[-1].split("\\")[-1]

                    child_result = parse_single_attachment_dry_run(
                        internet_message_id=internet_message_id,
                        sender=sender,
                        filename=normalized_inner_name,
                        file_bytes=inner_bytes,
                        email_received_at=email_received_at,
                        processing_run_id=processing_run_id,
                        mapping_by_sender=mapping_by_sender,
                    )
                    zip_results.append(child_result)
                    parsed_count += child_result.get("parsed_trade_count", 0)

            result["zip_children"] = zip_results
            result["parsed_trades"] = []

        else:
            result["status"] = "UNSUPPORTED_FILE_TYPE"

        result["parsed_trade_count"] = parsed_count
        result["status"] = "PARSED" if parsed_count > 0 else result["status"]

    except Exception as e:
        logging.exception("Dry-run parsing failed for file %s: %s", filename, e)
        result["status"] = "ERROR"
        result["error"] = str(e)

    return result


def process_message_for_debug(
    conn,
    token: str,
    mailbox: str,
    msg: Dict[str, Any],
    mapping_by_sender: Dict[str, Dict[str, Any]],
    processing_run_id: int,
    dry_run: bool = True,
) -> Dict[str, Any]:
    message_id = msg["id"]
    internet_message_id = clean_text(msg.get("internetMessageId"))
    subject = clean_text(msg.get("subject"))
    received_at_raw = msg.get("receivedDateTime")
    sender = normalize_email_address(msg.get("from", {}))

    if isinstance(received_at_raw, str):
        try:
            received_at = datetime.fromisoformat(received_at_raw.replace("Z", "+00:00"))
        except Exception:
            received_at = None
    else:
        received_at = received_at_raw

    result = {
        "message_id": message_id,
        "internet_message_id": internet_message_id,
        "sender": sender,
        "subject": subject,
        "received_at": str(received_at) if received_at else None,
        "template_code": detect_template_from_mapping(sender, mapping_by_sender),
        "has_mapping": sender in mapping_by_sender,
        "body_parsed_trade_count": 0,
        "attachments_total": 0,
        "attachments_processed": [],
        "total_parsed_trade_count": 0,
        "status": "SKIPPED",
        "error": None,
    }

    if not internet_message_id:
        result["status"] = "SKIPPED_NO_INTERNET_MESSAGE_ID"
        return result

    if sender not in mapping_by_sender and _fallback_by_domain(sender) is None:
        result["status"] = "SKIPPED_NO_MAPPING"
        return result

    if sender == "bo.tdsm@zarattinibank.ch" and "confirmation" not in subject.lower():
        logging.warning("ZARATTINI SUBJECT FILTER HIT DEBUG | subject=%s", subject)
        result["status"] = "SKIPPED_SUBJECT_FILTER"
        return result

    try:
        full_msg = get_message_full(token, mailbox, message_id)
        body_html = (full_msg.get("body", {}) or {}).get("content") or ""
        body_text = strip_html_tags(body_html)
        attachments = get_message_attachments(token, mailbox, message_id)
        result["attachments_total"] = len(attachments)

        template_code = detect_template_from_mapping(sender, mapping_by_sender)

        _EMAIL_BODY_PARSERS = {
            "STONEX_REPO_EMAIL": parse_stonex_repo_email_body,
            "GRANT_WESTOVER_REPO_EMAIL": parse_grant_westover_email_body,
            "FAB_REPO_EMAIL": parse_fab_repo_email_body,
        }

        if template_code in _EMAIL_BODY_PARSERS:
            repo_trades = _EMAIL_BODY_PARSERS[template_code](
                body_text=body_text,
                internet_message_id=internet_message_id,
                sender=sender,
                email_received_at=received_at,
                processing_run_id=processing_run_id,
                email_id=None,
                mapping_by_sender=mapping_by_sender,
            )
            result["body_parsed_trade_count"] = len(repo_trades)
            result["total_parsed_trade_count"] += len(repo_trades)

            if not dry_run:
                email_id = insert_settlement_email(
                    conn=conn,
                    internet_message_id=internet_message_id,
                    message_id=message_id,
                    sender=sender,
                    subject=subject,
                    received_at=received_at,
                    status="DEBUG_REPARSE",
                    note="Debug reparse body",
                    mailbox=mailbox,
                    attachment_count=len(attachments),
                    parsed_trade_count=0,
                    processing_run_id=processing_run_id,
                )
                for trade in repo_trades:
                    trade["email_id"] = email_id
                    upsert_settlement_trade(conn, trade)

        for idx, att in enumerate(attachments, start=1):
            odata_type = att.get("@odata.type")
            filename = att.get("name") or "unnamed"
            attachment_id = att.get("id")

            if odata_type != "#microsoft.graph.fileAttachment":
                result["attachments_processed"].append({
                    "file_name": filename,
                    "status": "SKIPPED_NON_FILE_ATTACHMENT",
                })
                continue

            content_bytes_b64 = att.get("contentBytes")
            if content_bytes_b64:
                file_bytes = base64.b64decode(content_bytes_b64)
            else:
                if not attachment_id:
                    result["attachments_processed"].append({
                        "file_name": filename,
                        "status": "SKIPPED_NO_ATTACHMENT_ID",
                    })
                    continue

                file_bytes = get_attachment_content_bytes(
                    token=token,
                    mailbox=mailbox,
                    message_id=message_id,
                    attachment_id=attachment_id,
                )

            if dry_run:
                att_result = parse_single_attachment_dry_run(
                    internet_message_id=internet_message_id,
                    sender=sender,
                    filename=filename,
                    file_bytes=file_bytes,
                    email_received_at=received_at,
                    processing_run_id=processing_run_id,
                    mapping_by_sender=mapping_by_sender,
                )
                result["attachments_processed"].append(att_result)
                result["total_parsed_trade_count"] += att_result.get("parsed_trade_count", 0)
            else:
                parsed_count = parse_single_attachment(
                    conn=conn,
                    internet_message_id=internet_message_id,
                    sender=sender,
                    filename=filename,
                    file_bytes=file_bytes,
                    email_received_at=received_at,
                    processing_run_id=processing_run_id,
                    email_id=None,
                    attachment_order=idx,
                    parent_zip_file_name=None,
                    mapping_by_sender=mapping_by_sender,
                )
                result["attachments_processed"].append({
                    "file_name": filename,
                    "status": "PARSED" if parsed_count > 0 else "NO_TRADES_FOUND",
                    "parsed_trade_count": parsed_count,
                })
                result["total_parsed_trade_count"] += parsed_count

        result["status"] = "PARSED" if result["total_parsed_trade_count"] > 0 else "NO_TRADES_FOUND"
        return result

    except Exception as e:
        logging.exception("Debug processing failed for message %s: %s", internet_message_id, e)
        result["status"] = "ERROR"
        result["error"] = str(e)
        return result


def debug_test_last_messages_parsing(
    conn,
    token: str,
    mailbox: str,
    senders: List[str],
    processing_run_id: int,
    top_n: int = 3,
    days_back: int = 120,
    dry_run: bool = True,
) -> Dict[str, Any]:
    mapping_by_sender = load_mapping(conn)
    since_dt = now_utc() - timedelta(days=days_back)

    summary = {
        "run_id": processing_run_id,
        "mailbox": mailbox,
        "top_n_per_sender": top_n,
        "days_back": days_back,
        "dry_run": dry_run,
        "total_senders": len(senders),
        "total_messages_found": 0,
        "total_messages_parsed": 0,
        "total_messages_no_trades": 0,
        "total_messages_errors": 0,
        "total_messages_skipped_subject_filter": 0,
        "total_trades_found": 0,
        "senders": [],
    }

    for sender in senders:
        sender = (sender or "").strip().lower()
        sender_result = {
            "sender": sender,
            "template_code": detect_template_from_mapping(sender, mapping_by_sender),
            "in_mapping": sender in mapping_by_sender,
            "messages_found": 0,
            "messages": [],
        }

        # For file-based parsers: require PDF/Excel attachment.
        # Fetch up to top_n*5 candidates so we can skip image-only emails
        # (e.g. image001.png logo in signature) and still find top_n real ones.
        # Email-body parsers need no attachment.
        template_for_sender = detect_template_from_mapping(sender, mapping_by_sender)
        needs_attachment = template_for_sender not in EMAIL_BODY_TEMPLATES

        try:
            candidates = list_recent_messages_by_sender_python_filter(
                token=token,
                mailbox=mailbox,
                sender_email=sender,
                since_dt=since_dt,
                top_n=top_n * 5 if needs_attachment else top_n,
                require_attachments=needs_attachment,
            )

            msgs = []
            if needs_attachment:
                # Process candidates until we collect top_n with real PDF/Excel
                for candidate in candidates:
                    if len(msgs) >= top_n:
                        break
                    msg_result = process_message_for_debug(
                        conn=conn,
                        token=token,
                        mailbox=mailbox,
                        msg=candidate,
                        mapping_by_sender=mapping_by_sender,
                        processing_run_id=processing_run_id,
                        dry_run=dry_run,
                    )
                    # Skip emails where every attachment was an unsupported type (image, etc.)
                    all_unsupported = (
                        msg_result.get("attachments_total", 0) > 0
                        and msg_result.get("total_parsed_trade_count", 0) == 0
                        and all(
                            a.get("status") == "UNSUPPORTED_FILE_TYPE"
                            for a in msg_result.get("attachments_processed", [])
                        )
                    )
                    if all_unsupported:
                        logging.info(
                            "Skipping image-only email from %s: %s",
                            sender, candidate.get("subject", "")[:60],
                        )
                        continue
                    msgs.append(candidate)
                    sender_result["messages"].append(msg_result)
                    summary["total_trades_found"] += msg_result.get("total_parsed_trade_count", 0)
                    if msg_result["status"] == "PARSED":
                        summary["total_messages_parsed"] += 1
                    elif msg_result["status"] == "NO_TRADES_FOUND":
                        summary["total_messages_no_trades"] += 1
                    elif msg_result["status"] == "ERROR":
                        summary["total_messages_errors"] += 1
                    elif msg_result["status"] == "SKIPPED_SUBJECT_FILTER":
                        summary["total_messages_skipped_subject_filter"] += 1
            else:
                msgs = candidates
                for msg in msgs:
                    msg_result = process_message_for_debug(
                        conn=conn,
                        token=token,
                        mailbox=mailbox,
                        msg=msg,
                        mapping_by_sender=mapping_by_sender,
                        processing_run_id=processing_run_id,
                        dry_run=dry_run,
                    )
                    sender_result["messages"].append(msg_result)
                    summary["total_trades_found"] += msg_result.get("total_parsed_trade_count", 0)
                    if msg_result["status"] == "PARSED":
                        summary["total_messages_parsed"] += 1
                    elif msg_result["status"] == "NO_TRADES_FOUND":
                        summary["total_messages_no_trades"] += 1
                    elif msg_result["status"] == "ERROR":
                        summary["total_messages_errors"] += 1
                    elif msg_result["status"] == "SKIPPED_SUBJECT_FILTER":
                        summary["total_messages_skipped_subject_filter"] += 1

            sender_result["messages_found"] = len(msgs)
            summary["total_messages_found"] += len(msgs)

        except Exception as e:
            logging.exception("Debug test failed for sender %s: %s", sender, e)
            sender_result["error"] = str(e)
            summary["total_messages_errors"] += 1

        summary["senders"].append(sender_result)

    return summary


# =============================================================================
# MESSAGE PROCESSING
# =============================================================================
def process_message(
    conn,
    token: str,
    mailbox: str,
    msg: Dict[str, Any],
    mapping_by_sender: Dict[str, Dict[str, Any]],
    processing_run_id: int,
) -> Tuple[str, int]:
    message_id = msg["id"]
    internet_message_id = clean_text(msg.get("internetMessageId"))
    subject = clean_text(msg.get("subject"))
    received_at_raw = msg.get("receivedDateTime")
    sender = normalize_email_address(msg.get("from", {}))
    sender_name = ((msg.get("from") or {}).get("emailAddress") or {}).get("name") or ""

    if isinstance(received_at_raw, str):
        try:
            received_at = datetime.fromisoformat(received_at_raw.replace("Z", "+00:00"))
        except Exception:
            received_at = None
    else:
        received_at = received_at_raw

    if not internet_message_id:
        return ("SKIPPED", 0)

    # CMF emails from trading desk: route to dedicated handler
    subj_lower = (subject or "").lower()
    if sender.endswith("@amwealth.ae") and "cash management facilit" in subj_lower:
        return _process_cmf_message(
            conn=conn, token=token, mailbox=mailbox, msg=msg,
            sender=sender,
            internet_message_id=internet_message_id, subject=subject,
            received_at=received_at, processing_run_id=processing_run_id,
        )

    if sender not in mapping_by_sender and _fallback_by_domain(sender) is None:
        insert_settlement_email(
            conn=conn,
            internet_message_id=internet_message_id,
            message_id=message_id,
            sender=sender,
            subject=subject,
            received_at=received_at,
            status="SKIPPED",
            note="Sender not found in counterparty_email_mapping",
            mailbox=mailbox,
            attachment_count=0,
            parsed_trade_count=0,
            processing_run_id=processing_run_id,
        )
        return ("SKIPPED", 0)

    # FAB SWIFT MT545/MT547: route to dedicated handler, skip normal pipeline
    if sender == "noreply@bankfab.com":
        subj_upper = (subject or "").upper()
        if "MT545" in subj_upper or "MT547" in subj_upper:
            return _process_fab_swift_message(
                conn=conn, token=token, mailbox=mailbox, msg=msg,
                internet_message_id=internet_message_id, subject=subject,
                received_at=received_at, processing_run_id=processing_run_id,
            )
        if "MT566" in subj_upper:
            return _process_mt566_message(
                conn=conn, token=token, mailbox=mailbox, msg=msg,
                internet_message_id=internet_message_id, subject=subject,
                received_at=received_at, processing_run_id=processing_run_id,
            )
        return ("SKIPPED", 0)

    if sender == "bo.tdsm@zarattinibank.ch" and "confirmation" not in subject.lower():
        logging.warning("ZARATTINI SUBJECT FILTER HIT MAIN | subject=%s", subject)
        insert_settlement_email(
            conn=conn,
            internet_message_id=internet_message_id,
            message_id=message_id,
            sender=sender,
            subject=subject,
            received_at=received_at,
            status="SKIPPED",
            note="Zarattini subject filter: 'confirmation' not in subject",
            mailbox=mailbox,
            attachment_count=0,
            parsed_trade_count=0,
            processing_run_id=processing_run_id,
        )
        return ("SKIPPED", 0)

    if email_already_processed(conn, internet_message_id):
        return ("ALREADY_PROCESSED", 0)

    full_msg = get_message_full(token, mailbox, message_id)
    body_html = (full_msg.get("body", {}) or {}).get("content") or ""
    body_text = strip_html_tags(body_html)
    attachments = get_message_attachments(token, mailbox, message_id)

    email_id = insert_settlement_email(
        conn=conn,
        internet_message_id=internet_message_id,
        message_id=message_id,
        sender=sender,
        subject=subject,
        received_at=received_at,
        status="RECEIVED",
        note="Message received for parsing",
        mailbox=mailbox,
        attachment_count=len(attachments),
        parsed_trade_count=0,
        processing_run_id=processing_run_id,
    )

    parsed_count = 0

    template_code = detect_template_from_mapping(sender, mapping_by_sender)
    _EMAIL_BODY_PARSERS = {
        "STONEX_REPO_EMAIL": parse_stonex_repo_email_body,
        "GRANT_WESTOVER_REPO_EMAIL": parse_grant_westover_email_body,
        "FAB_REPO_EMAIL": parse_fab_repo_email_body,
    }
    if template_code in _EMAIL_BODY_PARSERS:
        repo_trades = _EMAIL_BODY_PARSERS[template_code](
            body_text=body_text,
            internet_message_id=internet_message_id,
            sender=sender,
            email_received_at=received_at,
            processing_run_id=processing_run_id,
            email_id=email_id,
            mapping_by_sender=mapping_by_sender,
        )
        for trade in repo_trades:
            upsert_settlement_trade(conn, trade)
            parsed_count += 1

    for idx, att in enumerate(attachments, start=1):
        odata_type = att.get("@odata.type")
        filename = att.get("name") or "unnamed"
        attachment_id = att.get("id")

        if odata_type != "#microsoft.graph.fileAttachment":
            continue

        content_bytes_b64 = att.get("contentBytes")
        if content_bytes_b64:
            file_bytes = base64.b64decode(content_bytes_b64)
        else:
            if not attachment_id:
                continue
            try:
                file_bytes = get_attachment_content_bytes(
                    token=token,
                    mailbox=mailbox,
                    message_id=message_id,
                    attachment_id=attachment_id,
                )
            except Exception as e:
                logging.exception("Failed to load attachment body for %s: %s", filename, e)
                continue

        parsed_count += parse_single_attachment(
            conn=conn,
            internet_message_id=internet_message_id,
            sender=sender,
            sender_name=sender_name,
            filename=filename,
            file_bytes=file_bytes,
            email_received_at=received_at,
            processing_run_id=processing_run_id,
            email_id=email_id,
            attachment_order=idx,
            parent_zip_file_name=None,
            mapping_by_sender=mapping_by_sender,
        )

    status = "PARSED" if parsed_count > 0 else "NO_TRADES_FOUND"
    note = f"Parsed trades: {parsed_count}"

    insert_settlement_email(
        conn=conn,
        internet_message_id=internet_message_id,
        message_id=message_id,
        sender=sender,
        subject=subject,
        received_at=received_at,
        status=status,
        note=note,
        mailbox=mailbox,
        attachment_count=len(attachments),
        parsed_trade_count=parsed_count,
        processing_run_id=processing_run_id,
    )

    return (status, parsed_count)


# =============================================================================
# TIMERS
# =============================================================================
def settlement_email_parser_timer(mytimer=None) -> None:
    logging.info("### settlement_email_parser_timer started ###")

    run_id = None
    conn = None

    try:
        token = get_graph_token()
        conn = get_conn()
        run_id = start_agent_run(conn, "settlement_email_parser_timer")

        mapping_by_sender = load_mapping(conn)
        allowed_senders = get_allowed_senders(mapping_by_sender)
        since_dt = now_utc() - timedelta(hours=LOOKBACK_HOURS)

        total = 0
        parsed_messages = 0
        parsed_trades = 0
        skipped = 0

        for mailbox in GRAPH_MAILBOXES:
            messages = list_recent_messages(token, mailbox, since_dt)
            logging.info("EMAIL_PARSER mailbox=%s messages_count=%d since=%s", mailbox, len(messages), since_dt.isoformat())
            for msg in messages:
                total += 1
                sender = normalize_email_address(msg.get("from", {}))
                # Allow internal @amwealth.ae senders (for CMF emails)
                if not is_sender_allowed(sender, allowed_senders) and not sender.endswith("@amwealth.ae"):
                    skipped += 1
                    continue

                status, count = process_message(
                    conn=conn,
                    token=token,
                    mailbox=mailbox,
                    msg=msg,
                    mapping_by_sender=mapping_by_sender,
                    processing_run_id=run_id,
                )

                if status in {"PARSED", "NO_TRADES_FOUND"}:
                    parsed_messages += 1
                elif status in {"SKIPPED", "ALREADY_PROCESSED"}:
                    skipped += 1

                parsed_trades += count

        finish_agent_run(
            conn,
            run_id,
            "SUCCESS",
            f"Checked messages={total}, processed_messages={parsed_messages}, parsed_trades={parsed_trades}, skipped={skipped}",
        )

        logging.info("Settlement parsing completed successfully")

    except Exception as e:
        logging.exception("Settlement parsing failed: %s", e)
        try:
            if conn and run_id:
                finish_agent_run(conn, run_id, "FAILED", str(e))
        except Exception:
            logging.exception("Failed to mark run as FAILED")
        raise

    finally:
        if conn:
            conn.close()


# =============================================================================
# HTTP DEBUG FUNCTION TO TEST EXISTING EMAILS
# =============================================================================
@app.function_name(name="debug_settlement_sender_test")
@app.route(route="debug/settlement-sender-test", auth_level=func.AuthLevel.FUNCTION, methods=["GET", "POST"])
def debug_settlement_sender_test(req: func.HttpRequest) -> func.HttpResponse:
    conn = None
    run_id = None

    try:
        body = {}
        try:
            body = req.get_json()
        except Exception:
            body = {}

        senders_param = body.get("senders") or req.params.get("senders")
        top_n_param = body.get("top_n") or req.params.get("top_n")
        days_back_param = body.get("days_back") or req.params.get("days_back")
        dry_run_param = body.get("dry_run") if "dry_run" in body else req.params.get("dry_run")
        mailbox_param = body.get("mailbox") or req.params.get("mailbox")

        if senders_param:
            if isinstance(senders_param, list):
                senders = [str(x).strip().lower() for x in senders_param if str(x).strip()]
            else:
                senders = [x.strip().lower() for x in str(senders_param).split(",") if x.strip()]
        else:
            senders = TEST_SENDERS_DEFAULT

        top_n = int(top_n_param) if top_n_param else 3
        days_back = int(days_back_param) if days_back_param else 120

        if isinstance(dry_run_param, bool):
            dry_run = dry_run_param
        else:
            dry_run = str(dry_run_param or "true").strip().lower() in {"true", "1", "yes", "y"}

        mailbox_to_use = (mailbox_param or GRAPH_MAILBOX).strip()

        token = get_graph_token()
        conn = get_conn()
        run_id = start_agent_run(conn, "debug_settlement_sender_test")

        result = debug_test_last_messages_parsing(
            conn=conn,
            token=token,
            mailbox=mailbox_to_use,
            senders=senders,
            processing_run_id=run_id,
            top_n=top_n,
            days_back=days_back,
            dry_run=dry_run,
        )

        result["code_version"] = "2026-03-16-zarattini-mailbox-fix-v1"
        result["generated_at_utc"] = str(now_utc())
        logging.warning("DEBUG MAILBOX USED: %s", mailbox_to_use)

        finish_agent_run(
            conn,
            run_id,
            "SUCCESS",
            (
                f"senders={len(senders)}, "
                f"messages_found={result['total_messages_found']}, "
                f"messages_parsed={result['total_messages_parsed']}, "
                f"messages_no_trades={result['total_messages_no_trades']}, "
                f"errors={result['total_messages_errors']}, "
                f"trades_found={result['total_trades_found']}, "
                f"dry_run={dry_run}"
            ),
        )

        return func.HttpResponse(
            json.dumps(result, default=str, indent=2),
            mimetype="application/json",
            status_code=200,
        )

    except Exception as e:
        logging.exception("debug_settlement_sender_test failed: %s", e)
        try:
            if conn and run_id:
                finish_agent_run(conn, run_id, "FAILED", str(e))
        except Exception:
            logging.exception("Failed to mark debug run as FAILED")

        return func.HttpResponse(
            json.dumps({"status": "ERROR", "error": str(e)}, indent=2),
            mimetype="application/json",
            status_code=500,
        )

    finally:
        if conn:
            conn.close()


# =============================================================================
# SETTLEMENT RECONCILIATION
# =============================================================================
# =============================================================================
# RECONCILIATION HELPERS
# =============================================================================
def values_equal_decimal(a, b, tolerance: Decimal = Decimal("0.0001")) -> bool:
    da = parse_decimal(a)
    db = parse_decimal(b)

    if da is None and db is None:
        return True
    if da is None or db is None:
        return False

    return abs(da - db) <= tolerance


def build_reconciliation_key(st: Dict[str, Any]) -> str:
    return "|".join([
        clean_text(str(st.get("isin") or "")),
        clean_text(str(st.get("side") or "")),
        clean_text(str(st.get("trade_date") or "")),
        clean_text(str(st.get("value_date") or "")),
        clean_text(str(st.get("quantity") or st.get("nominal") or "")),
        clean_text(str(st.get("price") or st.get("price_in_percentage") or "")),
    ])


def load_settlement_trades_for_reconciliation(
    conn,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    value_date_from: Optional[date] = None,
) -> List[Dict[str, Any]]:
    """
    Loads settlement trades (parsed from counterparty emails) for reconciliation.
    date_from / date_to filter by trade_date (inclusive).
    value_date_from: exclude confos that already settled before this date (value_date < value_date_from).
    """
    where = "where validation_status = 'PARSED'"
    params: list = []
    if date_from is not None:
        where += " AND trade_date >= %s"
        params.append(date_from)
    if date_to is not None:
        where += " AND trade_date <= %s"
        params.append(date_to)
    if value_date_from is not None:
        where += " AND (value_date IS NULL OR value_date >= %s)"
        params.append(value_date_from)

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            f"""
            select
                id,
                internet_message_id,
                source_file,
                source_type,
                broker_name,
                security_name,
                isin,
                side,
                trade_date,
                value_date,
                quantity,
                price,
                price_currency,
                consideration,
                commission,
                net_amount,
                nominal,
                price_in_percentage,
                accrued_interest,
                settlement_currency,
                parser_template,
                validation_status,
                validation_note,
                counterparty_reference,
                our_ssi,
                created_at
            from back_office_auto.settlement_trades
            {where}
            order by created_at desc, id desc
            """,
            params or None,
        )
        return [dict(r) for r in cur.fetchall()]


def aggregate_settlement_trades(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Group settlement_trades rows from the same source file with the same
    ISIN + Side + Trade Date + Value Date and sum Qty / Net Amount.
    This handles cases like Instinet CSV where multiple partial fills
    for the same security arrive in one file and must be netted before
    matching against a single internal deal.
    """
    groups: Dict[tuple, Dict[str, Any]] = {}
    for row in rows:
        key = (
            row.get("internet_message_id") or "",
            (row.get("isin") or "").upper().strip(),
            (row.get("side") or "").upper().strip(),
            row.get("trade_date"),
            row.get("value_date"),
        )
        if key not in groups:
            groups[key] = dict(row)
        else:
            agg = groups[key]
            for field in ("quantity", "net_amount", "consideration", "nominal"):
                if row.get(field) is not None:
                    prev = agg.get(field)
                    agg[field] = (prev or Decimal("0")) + row[field]
    return list(groups.values())


def load_strict_deals_to_process(
    conn,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    include_settled: bool = False,
) -> List[Dict[str, Any]]:
    """
    Loads internal deals (tab_deals) eligible for reconciliation comparison.
    date_from / date_to filter by trade_date (inclusive).
    include_settled=True includes status=4 (SETTLED) deals (for Table A matching).
    """
    extra = ""
    params: list = []
    if date_from is not None:
        extra += " AND trades.trade_date >= %s"
        params.append(date_from)
    if date_to is not None:
        extra += " AND trades.trade_date <= %s"
        params.append(date_to)

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            f"""
            SELECT
                trades.id,
                trades.id AS back_id,
                trades.deal,
                trades.comment,
                CASE
                    WHEN trades.action = 0 THEN 'BUY'
                    WHEN trades.action = 2 THEN 'BALANCE'
                    WHEN trades.action = 1 THEN 'SELL'
                END AS direction,
                trades.symbol,
                trades.qty,
                trades.price,
                trades.price_in_percentage,
                trades.currency_price,
                trades.transaction_value,
                trades.currency_pay,
                trades.login,
                cp.name AS counterparty,
                trades.trade_date,
                trades.value_date_cash,
                trades.value_date_securities,
                trades.settle_date_cash,
                trades.settle_date_securities,
                trades.type_calculations,
                trades.settle_type,
                s.description AS status,
                trades.reason,
                trades.commission,
                trades.commission_fee,
                trades.dealer,
                trades.lot,
                trades.nominal,
                trades.accrued,
                trades.external_id,
                trades.order_id,
                trades.time,
                trades.transaction_value + (CASE WHEN trades.action = 0 THEN -trades.execution_cost ELSE trades.execution_cost END) AS net_amount,
                (
                    SELECT sss.ssi_name
                    FROM back_office.tab_connect_deal_transfer cdt2
                    JOIN back_office.tab_transfer tf2 ON cdt2.id_transfer = tf2.id
                    JOIN back_office.tab_standard_settlement_instructions sss ON tf2.ssi_id = sss.id
                    WHERE cdt2.id_deal = trades.id
                    LIMIT 1
                ) AS ssi_name
            FROM back_office.tab_deals trades
            LEFT JOIN back_office.tab_counterparty cp
                ON trades.counterparty_id = cp.id
            LEFT JOIN back_office.tab_status s
                ON trades.status = s.id
            WHERE trades.reason = 0
              AND trades.status NOT IN (4, 6, 7)
              AND trades.login = 1
              AND trades.type_deal <> 2
              AND trades.settle_type = 'external'
              {extra}
            ORDER BY trades.trade_date DESC, trades.time DESC
            """,
            params or None,
        )
        return [dict(r) for r in cur.fetchall()]


def load_broad_trade_search(conn) -> List[Dict[str, Any]]:
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            """
            SELECT
                trades.id,
                trades.id AS back_id,
                CASE
                    WHEN trades.action = 0 THEN 'BUY'
                    WHEN trades.action = 2 THEN 'BALANCE'
                    WHEN trades.action = 1 THEN 'SELL'
                END AS direction,
                trades.symbol,
                trades.qty,
                trades.price,
                trades.price_in_percentage,
                trades.transaction_value,
                trades.currency_pay,
                cp.name AS counterparty,
                trades.trade_date,
                trades.value_date_cash,
                trades.settle_date_cash,
                trades.settle_type,
                trades.reason,
                trades.status,
                trades.login,
                trades.order_id,
                trades.time,
                trades.nominal,
                trades.accrued,
                trades.transaction_value + (CASE WHEN trades.action = 0 THEN -trades.execution_cost ELSE trades.execution_cost END) AS net_amount,
                (
                    SELECT sss.ssi_name
                    FROM back_office.tab_connect_deal_transfer cdt2
                    JOIN back_office.tab_transfer tf2 ON cdt2.id_transfer = tf2.id
                    JOIN back_office.tab_standard_settlement_instructions sss ON tf2.ssi_id = sss.id
                    WHERE cdt2.id_deal = trades.id
                    LIMIT 1
                ) AS ssi_name
            FROM back_office.tab_deals trades
            LEFT JOIN back_office.tab_counterparty cp
                ON trades.counterparty_id = cp.id
            WHERE trades.reason = 0
              AND trades.status NOT IN (4, 6, 7)
              AND trades.login <> 1007
              AND trades.type_deal <> 2
              AND trades.settle_type = 'external'
            ORDER BY trades.trade_date DESC, trades.time DESC
            """
        )
        return [dict(r) for r in cur.fetchall()]


def load_unconfirmed_deals(
    conn,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> List[Dict[str, Any]]:
    """
    Loads deals that will never receive an automated confirmation email:
      - FX trades (type_deal = 2)
    These are shown in Table C as UNCONFIRMED for manual review.
    Note: ENBD is now handled via ENBD_PDF parser and goes through normal reconciliation (Tables A/B).
    """
    extra = ""
    params: list = []
    if date_from is not None:
        extra += " AND trades.trade_date >= %s"
        params.append(date_from)
    if date_to is not None:
        extra += " AND trades.trade_date <= %s"
        params.append(date_to)

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            f"""
            SELECT
                trades.id,
                trades.id AS back_id,
                CASE
                    WHEN trades.action = 0 THEN 'BUY'
                    WHEN trades.action = 2 THEN 'BALANCE'
                    WHEN trades.action = 1 THEN 'SELL'
                END AS direction,
                trades.symbol,
                trades.qty,
                trades.nominal,
                trades.price,
                trades.price_in_percentage,
                trades.transaction_value,
                trades.transaction_value + (CASE WHEN trades.action = 0 THEN -trades.execution_cost ELSE trades.execution_cost END) AS net_amount,
                trades.currency_pay,
                trades.login,
                cp.name AS counterparty,
                trades.trade_date,
                trades.value_date_cash,
                trades.value_date_securities,
                trades.settle_date_cash,
                trades.settle_date_securities,
                trades.type_deal,
                trades.settle_type,
                s.description AS status
            FROM back_office.tab_deals trades
            LEFT JOIN back_office.tab_counterparty cp
                ON trades.counterparty_id = cp.id
            LEFT JOIN back_office.tab_status s
                ON trades.status = s.id
            WHERE trades.reason = 0
              AND trades.status NOT IN (4, 6, 7)
              AND trades.login = 1
              AND trades.type_deal = 2
              {extra}
            ORDER BY trades.trade_date DESC, trades.id DESC
            """,
            params or None,
        )
        return [dict(r) for r in cur.fetchall()]


# =============================================================================
# T0 / T1 DATE HELPERS
# =============================================================================
# Dubai timezone is UTC+4.
# T0 = trade date (today Dubai).
# T1 = next business day (tomorrow Dubai, or Monday if Friday).
# The rules define WHEN reconciliation window opens per currency:
#   AED, HKD, EUR, GBP, CHF  → compare window opens at 17:00 T0 Dubai
#   USD                       → compare window opens at 08:00 T1 Dubai
#   Other                     → compare window opens at 08:00 T1 Dubai
#
# For DATA FILTERING in reconciliation we include:
#   - settlement_trades with trade_date in [t1_date .. t0_date] (last 2 business days)
#   - tab_deals with trade_date in [t1_date .. t0_date]
# =============================================================================

DUBAI_TZ = timezone(timedelta(hours=4))

# Currencies settled same day (T+0 — compare at T0 5pm)
T0_CURRENCIES = frozenset({"AED", "HKD", "EUR", "GBP", "CHF"})

# Hour thresholds (Dubai time) after which the comparison window is open
T0_COMPARE_HOUR = 17   # 5pm Dubai  — for T0 currencies
T1_COMPARE_HOUR = 8    # 8am Dubai  — for T1 currencies (USD, Other)


def prev_business_day(d: date) -> date:
    """Return the previous calendar business day (Mon–Fri), skipping weekends."""
    prev = d - timedelta(days=1)
    while prev.weekday() >= 5:   # 5=Sat, 6=Sun
        prev -= timedelta(days=1)
    return prev


def next_business_day(d: date) -> date:
    """Return the next calendar business day (Mon–Fri), skipping weekends."""
    nxt = d + timedelta(days=1)
    while nxt.weekday() >= 5:   # 5=Sat, 6=Sun
        nxt += timedelta(days=1)
    return nxt


def n_prev_business_days(d: date, n: int) -> date:
    """Return the date n business days before d."""
    result = d
    for _ in range(n):
        result = prev_business_day(result)
    return result


def get_t0_t1_dates() -> tuple[date, date, date]:
    """
    Returns (t0_date, t1_date, t_next_date) in Dubai timezone:
      t0_date      = today (Dubai)
      t1_date      = 5 business days ago — covers holiday gaps
      t_next_date  = next business day (Dubai)
    """
    t0 = datetime.now(DUBAI_TZ).date()
    t1 = n_prev_business_days(t0, 5)
    t_next = next_business_day(t0)
    return t0, t1, t_next


def is_reconciliation_window_open(currency: str | None) -> bool:
    """
    Returns True if the reconciliation compare window is currently open
    for the given currency based on Dubai time.
    """
    now_dubai = datetime.now(DUBAI_TZ)
    ccy = (currency or "").upper()
    if ccy in T0_CURRENCIES:
        return now_dubai.hour >= T0_COMPARE_HOUR
    else:
        return now_dubai.hour >= T1_COMPARE_HOUR


# =============================================================================
# RECONCILIATION FIELD MAPPING
# =============================================================================
# Показывает какие поля из settlement_trades сравниваются с полями tab_deals.
#
# ЖЁСТКИЙ ФИЛЬТР (все три обязательны — без совпадения кандидат не рассматривается):
#
#   settlement_trades.isin          ==  tab_deals.symbol
#   settlement_trades.side          ==  tab_deals.direction   (action: 0→BUY, 1→SELL)
#   settlement_trades.trade_date    ==  tab_deals.trade_date
#
# SCORE (нужно набрать ≥ 90 из максимальных 90 баллов):
#
#   +20  settlement_trades.value_date      ==  tab_deals.settle_date_cash  (или value_date_cash если settle_date_cash пустой)
#   +30  settlement_trades.quantity        ≈   tab_deals.qty               (допуск ±0.0001)
#   +20  settlement_trades.price           ≈   tab_deals.price             (допуск ±0.0001)
#   +20  settlement_trades.consideration   ≈   tab_deals.transaction_value (допуск ±0.01)
#
# АГРЕГИРОВАННЫЙ МАТЧИНГ (если точного нет):
#   SUM(tab_deals.qty)                ≈   settlement_trades.quantity      (допуск ±0.0001)
#   SUM(tab_deals.transaction_value)  ≈   settlement_trades.consideration (допуск ±0.01)
#   группировка по tab_deals.settle_date_cash / value_date_cash
#
# ПОХОЖИЕ (SIMILAR): те же кандидаты из broad search по isin в пределах ±3 дней от trade_date
# =============================================================================

def exact_score(st: Dict[str, Any], td: Dict[str, Any]) -> Tuple[int, List[str]]:
    score = 0
    notes = []

    td_value_date = td.get("settle_date_cash") or td.get("value_date_cash")

    if st.get("value_date") == td_value_date:
        score += 20
    else:
        notes.append("value_date_mismatch")

    # Qty: for bonds tab_deals.nominal = face value; for equities nominal is NULL/0 → fall back to qty
    ext_qty = st.get("quantity") or st.get("nominal")
    int_qty = td.get("nominal") if td.get("nominal") else td.get("qty")
    if ext_qty is not None and int_qty is not None:
        if values_equal_decimal(ext_qty, int_qty):
            score += 30
        else:
            notes.append("quantity_mismatch")

    # Price: prefer price_in_percentage (bond % of par) over absolute price.
    # If external has price_in_percentage, compare both sides on price_in_percentage.
    if st.get("price_in_percentage") is not None:
        ext_price = st.get("price_in_percentage")
        int_price = td.get("price_in_percentage") or td.get("price")
    elif st.get("price") is not None:
        ext_price = st.get("price")
        int_price = td.get("price") or td.get("price_in_percentage")
    else:
        ext_price = None
        int_price = None
    if ext_price is not None and int_price is not None:
        if values_equal_decimal(ext_price, int_price, Decimal("0.5")):
            score += 20
        else:
            notes.append("price_mismatch")

    ext_amount = st.get("consideration") if st.get("consideration") is not None else st.get("net_amount")
    int_amount = td.get("transaction_value") if td.get("transaction_value") else td.get("net_amount")
    if ext_amount is not None and int_amount is not None:
        if values_equal_decimal(ext_amount, int_amount, Decimal("0.5")):
            score += 20
        else:
            notes.append("amount_mismatch")

    return score, notes


def find_strict_candidates(st: Dict[str, Any], deals: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    candidates = []

    for td in deals:
        if clean_text(st.get("isin")) != clean_text(td.get("symbol")):
            continue
        if clean_text(st.get("side")) != clean_text(td.get("direction")):
            continue
        if st.get("trade_date") != td.get("trade_date"):
            continue
        candidates.append(td)

    return candidates


def try_exact_single_match(
    st: Dict[str, Any],
    candidates: List[Dict[str, Any]]
) -> Tuple[Optional[Dict[str, Any]], Optional[str], Optional[List[str]]]:
    if not candidates:
        return None, None, None

    scored = []
    for td in candidates:
        score, notes = exact_score(st, td)
        scored.append((score, notes, td))

    scored.sort(key=lambda x: x[0], reverse=True)
    best_score, best_notes, best_td = scored[0]

    if best_score >= 90:
        return best_td, "MATCHED", best_notes

    return best_td, None, best_notes


def try_aggregate_match(
    st: Dict[str, Any],
    candidates: List[Dict[str, Any]]
) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str], Optional[str]]:
    if not candidates:
        return None, None, None

    groups: Dict[str, List[Dict[str, Any]]] = {}
    for td in candidates:
        key = str(td.get("settle_date_cash") or td.get("value_date_cash") or "")
        groups.setdefault(key, []).append(td)

    for _, rows in groups.items():
        summed_qty = sum(
            float(r.get("nominal") or r.get("qty") or 0) if r.get("nominal") else float(r.get("qty") or 0)
            for r in rows
        )
        summed_amount = sum(float(r.get("net_amount") or r.get("transaction_value") or 0) for r in rows)

        ext_qty = st.get("quantity") or st.get("nominal")
        ext_amount = st.get("net_amount") or st.get("consideration")

        qty_ok = ext_qty is not None and values_equal_decimal(ext_qty, summed_qty)
        amount_ok = ext_amount is not None and values_equal_decimal(
            ext_amount, summed_amount, Decimal("0.5")
        )

        if qty_ok and amount_ok:
            # Build note with any field mismatches vs confo
            issues = []
            if st.get("value_date") and rows[0].get("settle_date_cash") or rows[0].get("value_date_cash"):
                int_vd = rows[0].get("settle_date_cash") or rows[0].get("value_date_cash")
                if st.get("value_date") != int_vd:
                    issues.append(f"value_date: confo={st.get('value_date')} vs BO={int_vd}")
            note = "; ".join(issues) if issues else f"matched against {len(rows)} internal rows"
            return rows, "MATCHED_AGGREGATED", note

    return None, None, None


def find_similar_broad_rows(st: Dict[str, Any], broad_deals: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    similar = []
    st_trade_date = st.get("trade_date")
    date_window = 3

    for td in broad_deals:
        if clean_text(st.get("isin")) != clean_text(td.get("symbol")):
            continue

        td_trade_date = td.get("trade_date")
        if st_trade_date and td_trade_date:
            date_diff = abs((st_trade_date - td_trade_date).days)
            if date_diff > date_window:
                continue

        similar.append(td)

    similar.sort(key=lambda x: (x.get("trade_date") or date.min, x.get("id") or 0), reverse=True)
    return similar[:5]


def upsert_reconciliation_result(
    conn,
    settlement_trade_id: int,
    internal_order_id: Optional[int],
    match_status: str,
    match_note: Optional[str],
    reconciliation_key: str,
    counterparty: Optional[str],
    isin: Optional[str],
    direction: Optional[str],
    trade_date: Optional[date],
    value_date: Optional[date],
    external_qty,
    internal_qty,
    external_amount,
    internal_amount,
    compare_side: Optional[str],
    run_id: Optional[int],
    matched_internal_ids: Optional[str] = None,
    matched_internal_count: Optional[int] = None,
    mismatch_json: Optional[str] = None,
    external_price=None,
    internal_price=None,
    external_value_date: Optional[date] = None,
    internal_value_date: Optional[date] = None,
):
    with conn.cursor() as cur:
        cur.execute(
            """
            insert into back_office_auto.settlement_reconciliation
            (
                settlement_trade_id,
                internal_order_id,
                match_status,
                match_note,
                created_at,
                reconciliation_key,
                counterparty,
                isin,
                direction,
                trade_date,
                value_date,
                external_qty,
                internal_qty,
                external_amount,
                internal_amount,
                compare_side,
                run_id,
                matched_internal_ids,
                matched_internal_count,
                mismatch_json,
                external_price,
                internal_price,
                external_value_date,
                internal_value_date
            )
            values
            (%s, %s, %s, %s, now(), %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s, %s, %s, %s)
            on conflict (settlement_trade_id, reconciliation_key)
            do update set
                internal_order_id = excluded.internal_order_id,
                match_status = excluded.match_status,
                match_note = excluded.match_note,
                counterparty = excluded.counterparty,
                isin = excluded.isin,
                direction = excluded.direction,
                trade_date = excluded.trade_date,
                value_date = excluded.value_date,
                external_qty = excluded.external_qty,
                internal_qty = excluded.internal_qty,
                external_amount = excluded.external_amount,
                internal_amount = excluded.internal_amount,
                compare_side = excluded.compare_side,
                run_id = excluded.run_id,
                matched_internal_ids = excluded.matched_internal_ids,
                matched_internal_count = excluded.matched_internal_count,
                mismatch_json = excluded.mismatch_json,
                external_price = excluded.external_price,
                internal_price = excluded.internal_price,
                external_value_date = excluded.external_value_date,
                internal_value_date = excluded.internal_value_date,
                created_at = now()
            """,
            (
                settlement_trade_id,
                internal_order_id,
                match_status,
                match_note,
                reconciliation_key,
                counterparty,
                isin,
                direction,
                trade_date,
                value_date,
                external_qty,
                internal_qty,
                external_amount,
                internal_amount,
                compare_side,
                run_id,
                matched_internal_ids,
                matched_internal_count,
                mismatch_json,
                external_price,
                internal_price,
                external_value_date,
                internal_value_date,
            ),
        )
    conn.commit()


def _detail_row(
    st: Dict[str, Any],
    status: str,
    td: Optional[Dict[str, Any]] = None,
    agg_rows: Optional[List[Dict[str, Any]]] = None,
    agg_note: Optional[str] = None,
    notes: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Build a flat dict for the human-readable reconciliation report."""
    internal_qty = None
    internal_amount = None
    internal_price = None
    internal_value_date = None
    internal_ids = None
    counterparty = None

    if td is not None:
        internal_qty = td.get("nominal") if td.get("nominal") else td.get("qty")
        internal_amount = td.get("net_amount") if td.get("net_amount") else td.get("transaction_value")
        internal_price = td.get("price_in_percentage") if td.get("price_in_percentage") else td.get("price")
        internal_value_date = td.get("settle_date_cash") or td.get("value_date_cash")
        internal_ids = str(td["id"])
        counterparty = td.get("counterparty")
    elif agg_rows:
        internal_qty = sum(float(r.get("nominal") or r.get("qty") or 0) if r.get("nominal") else float(r.get("qty") or 0) for r in agg_rows)
        internal_amount = sum(float(r.get("net_amount") or r.get("transaction_value") or 0) for r in agg_rows)
        internal_value_date = agg_rows[0].get("settle_date_cash") or agg_rows[0].get("value_date_cash")
        internal_ids = ",".join(str(r["id"]) for r in agg_rows)
        counterparty = agg_rows[0].get("counterparty")

    # CPTY SSI: use our_ssi populated by enrich_cpty_ssi() after parsing
    cpty_ssi = st.get("our_ssi") or ""
    if td is not None:
        int_ssi = td.get("ssi_name") or ""
    elif agg_rows:
        int_ssi = agg_rows[0].get("ssi_name") or ""
    else:
        int_ssi = ""
    ssi_note = "ssi_mismatch" if (cpty_ssi and int_ssi and cpty_ssi != int_ssi) else ""

    all_notes = list(notes or [])
    if ssi_note:
        all_notes.append(ssi_note)

    return {
        "status": status,
        "isin": st.get("isin"),
        "security_name": st.get("security_name"),
        "side": st.get("side"),
        "trade_date": st.get("trade_date"),
        "value_date": st.get("value_date"),
        "broker": st.get("broker_name"),
        "source_file": st.get("source_file"),
        "counterparty": counterparty,
        # confo (external) fields
        "ext_qty": st.get("quantity") or st.get("nominal"),
        "ext_price": st.get("price_in_percentage") or st.get("price"),
        "ext_amount": st.get("net_amount") or st.get("consideration"),
        "ext_value_date": st.get("value_date"),
        # internal (tab_deals) fields
        "int_qty": internal_qty,
        "int_price": internal_price,
        "int_amount": internal_amount,
        "int_value_date": internal_value_date,
        "int_ids": internal_ids,
        # CPTY SSI
        "cpty_ssi": cpty_ssi,
        "int_ssi": int_ssi,
        # mismatch details
        "notes": ", ".join(all_notes) if all_notes else (agg_note or ""),
    }


def _fmt_num(v) -> str:
    if v is None:
        return ""
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return str(v)


def _fmt_amount(v) -> str:
    if v is None:
        return ""
    try:
        return f"{float(v):,.2f}"
    except Exception:
        return str(v)


def _fmt_price(v) -> str:
    if v is None:
        return ""
    try:
        return f"{float(v):,.6f}".rstrip("0").rstrip(".")
    except Exception:
        return str(v)


def _fmt_date(v) -> str:
    if v is None:
        return ""
    return str(v)[:10]


def build_reconciliation_excel(result: dict, date_from, date_to) -> bytes:
    """Build an Excel workbook with reconciliation results. Returns bytes."""
    STATUS_COLOR = {
        "MATCHED":                            "C6EFCE",   # green
        "MATCHED_AGGREGATED":                 "C6EFCE",
        "PARTIAL":                            "FFEB9C",   # yellow
        "SIMILAR_FOUND_OUTSIDE_STRICT_SCOPE": "F4B942",   # orange
        "NOT_FOUND":                          "FFC7CE",   # red
        "NO_CONFO":                           "FCE4D6",   # light orange
    }

    wb = Workbook()

    # ── Sheet 1: Confo vs Internal ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Confo vs Internal"

    hdr = ["Status", "ISIN", "Side", "Trade Date", "Value Date", "Broker",
           "Ext Qty", "Int Qty", "Ext Price", "Int Price", "Ext Amount", "Int Amount", "Δ Amount", "CPTY SSI", "Notes"]
    ws1.append(hdr)
    for col_idx, _ in enumerate(hdr, 1):
        cell = ws1.cell(1, col_idx)
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for row in result.get("detail_rows", []):
        status = row.get("status", "")
        if status not in ("MATCHED", "PARTIAL", "MATCHED_AGGREGATED", "NETTING"):
            continue
        color = STATUS_COLOR.get(status, "FFFFFF")
        data = [
            status,
            row.get("isin") or "",
            row.get("side") or "",
            _fmt_date(row.get("trade_date")),
            _fmt_date(row.get("value_date")),
            row.get("broker") or "",
            _fmt_num(row.get("ext_qty")),
            _fmt_num(row.get("int_qty")),
            _fmt_price(row.get("ext_price")),
            _fmt_price(row.get("int_price")),
            _fmt_amount(row.get("ext_amount")),
            _fmt_amount(row.get("int_amount")),
            round(float(row["ext_amount"]) - float(row["int_amount"]), 2) if row.get("ext_amount") is not None and row.get("int_amount") is not None else "",
            row.get("cpty_ssi") or "",
            row.get("notes") or "",
        ]
        ws1.append(data)
        fill = PatternFill("solid", fgColor=color)
        for col_idx in range(1, len(data) + 1):
            ws1.cell(ws1.max_row, col_idx).fill = fill

    # Auto-width
    for col_idx, _ in enumerate(hdr, 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 16
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["F"].width = 22
    ws1.column_dimensions["K"].width = 30

    # ── Sheet 2: Internal without Confo ───────────────────────────────────────
    ws2 = wb.create_sheet("Internal - No Confo")
    hdr2 = ["Back ID", "ISIN", "Side", "Trade Date", "Value Date", "Counterparty", "Qty", "Amount", "GL Account"]
    ws2.append(hdr2)
    for col_idx, _ in enumerate(hdr2, 1):
        cell = ws2.cell(1, col_idx)
        cell.fill = PatternFill("solid", fgColor="843C0C")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    fill_red = PatternFill("solid", fgColor="FCE4D6")
    for td in result.get("unmatched_internal", []):
        data2 = [
            td.get("id") or "",
            td.get("symbol") or "",
            td.get("direction") or "",
            _fmt_date(td.get("trade_date")),
            _fmt_date(td.get("settle_date_cash") or td.get("value_date_cash")),
            td.get("counterparty") or "",
            _fmt_num(td.get("qty")),
            _fmt_amount(td.get("net_amount") or td.get("transaction_value")),
            compute_gl_account(td.get("symbol", "")),
        ]
        ws2.append(data2)
        for col_idx in range(1, len(data2) + 1):
            ws2.cell(ws2.max_row, col_idx).fill = fill_red

    for col_idx, _ in enumerate(hdr2, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── Sheet 3: Unconfirmed (FX) ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Unconfirmed FX")
    hdr3 = ["Back ID", "ISIN / CCY", "Type", "Side", "Trade Date", "Value Date", "Counterparty", "Qty / Nominal", "Amount"]
    ws3.append(hdr3)
    for col_idx, _ in enumerate(hdr3, 1):
        cell = ws3.cell(1, col_idx)
        cell.fill = PatternFill("solid", fgColor="17375E")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    fill_blue = PatternFill("solid", fgColor="DAEEF3")
    for td in result.get("unconfirmed_deals", []):
        type_label = "FX"
        qty_val = td.get("nominal") if td.get("nominal") else td.get("qty")
        data3 = [
            td.get("id") or "",
            td.get("symbol") or "",
            type_label,
            td.get("direction") or "",
            _fmt_date(td.get("trade_date")),
            _fmt_date(td.get("settle_date_cash") or td.get("value_date_cash")),
            td.get("counterparty") or "",
            _fmt_num(qty_val),
            _fmt_amount(td.get("net_amount") or td.get("transaction_value")),
        ]
        ws3.append(data3)
        for col_idx in range(1, len(data3) + 1):
            ws3.cell(ws3.max_row, col_idx).fill = fill_blue

    for col_idx, _ in enumerate(hdr3, 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── Sheet 4: FAB SWIFT MT545/MT547 ────────────────────────────────────────
    fab_swift_rows = result.get("fab_swift_rows", [])
    if fab_swift_rows:
        ws4 = wb.create_sheet("FAB SWIFT MT545-MT547")
        hdr4 = [
            "ISIN", "Side", "Match", "Internal Deal",
            "MT", "Sett. Date", "Eff. Sett. (PDF)",
            "Settled Amt", "Net Amt (Sys)", "Δ Amount", "Note",
        ]
        ws4.append(hdr4)
        for col_idx, _ in enumerate(hdr4, 1):
            cell = ws4.cell(1, col_idx)
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        STATUS_COLOR_SWIFT = {
            "MATCHED":        "E2EFDA",
            "AMOUNT_MISMATCH":"FCE4D6",
            "DATE_MISMATCH":  "FFF2CC",
            "NOT_FOUND":      "FFC7CE",
        }
        for sw in fab_swift_rows:
            st = sw.get("match_status") or "NOT_FOUND"
            color = STATUS_COLOR_SWIFT.get(st, "FFFFFF")
            amount_diff = sw.get("amount_diff")
            diff_val = round(float(amount_diff), 2) if amount_diff is not None else ""
            data4 = [
                sw.get("isin") or "",
                sw.get("side") or "",
                st,
                sw.get("internal_deal_id") or "",
                sw.get("mt_type") or "",
                _fmt_date(sw.get("settlement_date")),
                _fmt_date(sw.get("effective_settlement_date")),
                _fmt_amount(sw.get("settled_amount")),
                _fmt_amount(sw.get("internal_amount")),
                diff_val,
                sw.get("match_note") or "",
            ]
            ws4.append(data4)
            fill4 = PatternFill("solid", fgColor=color)
            for col_idx in range(1, len(data4) + 1):
                ws4.cell(ws4.max_row, col_idx).fill = fill4

        for col_idx, _ in enumerate(hdr4, 1):
            ws4.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def compute_gl_account(symbol: str) -> str:
    """Return GL Account based on ISIN/symbol.
    Currencies (2-3 uppercase letters only) → FAB-Clients-{symbol}
    Securities (ISIN format) → FAB_EC_Clients
    """
    s = (symbol or "").strip()
    if s and re.match(r'^[A-Z]{2,3}$', s):
        return f"FAB-Clients-{s}"
    return "FAB_EC_Clients"


def build_reconciliation_html(result: dict, date_from, date_to) -> str:
    """Build HTML email body for reconciliation report."""
    # Count only non-instructed rows for summary
    _active_rows = [
        r for r in result.get("detail_rows", [])
        if not r.get("instructed")
    ]
    matched = sum(1 for r in _active_rows if r.get("status") in ("MATCHED", "MATCHED_AGGREGATED"))
    partial = sum(1 for r in _active_rows if r.get("status") == "PARTIAL")
    no_confo = len(result.get("unmatched_internal", []))
    unconfirmed_count = len(result.get("unconfirmed_deals", []))

    date_label = str(date_to or date_from or "all dates")

    STATUS_BG = {
        "MATCHED":                            "#C6EFCE",
        "MATCHED_AGGREGATED":                 "#C6EFCE",
        "PARTIAL":                            "#FFEB9C",
        "SIMILAR_FOUND_OUTSIDE_STRICT_SCOPE": "#F4B942",
        "NOT_FOUND":                          "#FFC7CE",
    }
    STATUS_LABEL = {
        "MATCHED":                            "✅ MATCHED",
        "MATCHED_AGGREGATED":                 "✅ MATCHED~",
        "PARTIAL":                            "⚠️ PARTIAL",
        "SIMILAR_FOUND_OUTSIDE_STRICT_SCOPE": "🟠 SIMILAR",
        "NOT_FOUND":                          "❌ NOT FOUND",
    }

    style = """
    body { font-family: Calibri, Arial, sans-serif; font-size: 13px; color: #222; }
    h2 { color: #2F5496; margin-bottom: 4px; }
    .summary { display: flex; gap: 12px; margin: 16px 0; }
    .badge { padding: 12px 20px; border-radius: 6px; font-size: 15px; font-weight: bold; min-width: 120px; text-align: center; }
    .green  { background: #C6EFCE; color: #276221; }
    .yellow { background: #FFEB9C; color: #7D6608; }
    .red    { background: #FFC7CE; color: #9C0006; }
    table { border-collapse: collapse; width: 100%; margin-bottom: 24px; font-size: 12px; }
    th { background: #2F5496; color: white; padding: 6px 8px; text-align: center; }
    td { padding: 5px 8px; border: 1px solid #D0D0D0; white-space: nowrap; }
    .sect { color: #2F5496; font-size: 14px; font-weight: bold; margin: 20px 0 6px 0; border-bottom: 2px solid #2F5496; padding-bottom: 3px; }
    .mismatch { color: #C00000; font-weight: bold; }
    """

    def td_pair(ext_val, int_val, is_amount=False, is_price=False):
        """Render two cells; highlight if they differ significantly."""
        fmt = _fmt_price if is_price else (_fmt_amount if is_amount else _fmt_num)
        ext_s = fmt(ext_val) if ext_val is not None else "&nbsp;"
        int_s = fmt(int_val) if int_val is not None else "&nbsp;"
        threshold = 1.0 if is_amount else (0.01 if not is_price else 0.5)
        mismatch = (
            ext_val is not None and int_val is not None
            and abs(float(ext_val or 0) - float(int_val or 0)) > threshold
        )
        cls = ' class="mismatch"' if mismatch else ""
        return f"<td{cls}>{ext_s}</td><td{cls}>{int_s}</td>"

    # ── Summary badges ─────────────────────────────────────────────────────────
    html = f"""<html><head><style>{style}</style></head><body>
<h2>Settlement Reconciliation Report</h2>
<div style="color:#555; margin-bottom:8px;">Trade dates: <b>{date_label}</b></div>
<div class="summary">
  <div class="badge green">✅ Matched<br><span style="font-size:22px">{matched}</span></div>
  <div class="badge yellow">⚠️ Needs Review<br><span style="font-size:22px">{partial}</span></div>
  <div class="badge red">❌ No Confo<br><span style="font-size:22px">{no_confo}</span></div>
  <div class="badge" style="background:#DAEEF3; color:#1F497D;">🔔 Unconfirmed<br><span style="font-size:22px">{unconfirmed_count}</span></div>
</div>
<div style="font-size:12px; color:#555; margin-bottom:16px;">
  Exact match: {sum(1 for r in _active_rows if r.get('status') == 'MATCHED')} &nbsp;|&nbsp;
  Netting: {sum(1 for r in _active_rows if r.get('status') == 'MATCHED_AGGREGATED')} &nbsp;|&nbsp;
  Partial (mismatch): {partial} &nbsp;|&nbsp;
  Internal no confo: {no_confo} &nbsp;|&nbsp;
  Unconfirmed (FX): {unconfirmed_count}
</div>
"""

    # ── Table A: Confo vs Internal ─────────────────────────────────────────────
    _table_a_rows = [
        r for r in result.get("detail_rows", [])
        if r.get("status") in ("MATCHED", "PARTIAL", "MATCHED_AGGREGATED", "NETTING")
        and r.get("validation_status") != "FAILED"
        and not r.get("instructed")
    ]
    html += f"""<div class="sect">A. Confo vs Internal — {len(_table_a_rows)} trades to process</div>
<table>
<tr>
  <th>Status</th><th>ISIN</th><th>Side</th><th>Trade Date</th><th>Value Date</th>
  <th>Broker</th><th>Ext Qty</th><th>Int Qty</th><th>Ext Price</th><th>Int Price</th><th>Ext Amount</th><th>Int Amount</th><th>Δ Amount</th><th>CPTY SSI</th><th>Notes</th>
</tr>
"""
    for row in _table_a_rows:
        status = row.get("status", "")
        if status not in ("MATCHED", "PARTIAL", "MATCHED_AGGREGATED", "NETTING"):
            continue
        is_instructed = row.get("instructed", False)
        bg = STATUS_BG.get(status, "#FFFFFF")
        row_style = f'background:{bg};opacity:0.6' if is_instructed else f'background:{bg}'
        label = STATUS_LABEL.get(status, status)
        if is_instructed:
            label = f"{label} ✅ INSTRUCTED"
        cpty_ssi = row.get("cpty_ssi") or ""
        int_ssi = row.get("int_ssi") or ""
        ssi_mismatch = cpty_ssi and int_ssi and cpty_ssi != int_ssi
        ssi_cls = ' class="mismatch"' if ssi_mismatch else ""
        ssi_display = cpty_ssi or "&nbsp;"
        html += f'<tr style="{row_style}">'
        html += f"<td><b>{label}</b></td>"
        html += f"<td>{row.get('isin') or ''}</td>"
        html += f"<td>{row.get('side') or ''}</td>"
        html += f"<td>{_fmt_date(row.get('trade_date'))}</td>"
        html += f"<td>{_fmt_date(row.get('value_date'))}</td>"
        html += f"<td>{row.get('broker') or ''}</td>"
        html += td_pair(row.get("ext_qty"), row.get("int_qty"))
        html += td_pair(row.get("ext_price"), row.get("int_price"), is_price=True)
        html += td_pair(row.get("ext_amount"), row.get("int_amount"), is_amount=True)
        _ea = row.get("ext_amount")
        _ia = row.get("int_amount")
        if _ea is not None and _ia is not None:
            try:
                _diff = float(_ea) - float(_ia)
                _diff_str = f"{_diff:+,.2f}"
                _diff_color = ' style="color:green"' if abs(_diff) < 1 else ' style="color:red;font-weight:bold"'
            except Exception:
                _diff_str, _diff_color = "", ""
        else:
            _diff_str, _diff_color = "", ""
        html += f"<td{_diff_color}>{_diff_str}</td>"
        html += f"<td{ssi_cls}>{ssi_display}</td>"
        html += f"<td>{row.get('notes') or ''}</td>"
        html += "</tr>\n"
    html += "</table>\n"

    # ── Table B: Internal without Confo ───────────────────────────────────────
    unmatched = result.get("unmatched_internal", [])
    if unmatched:
        html += """<div class="sect">B. Internal Deals — No Confo Received</div>
<table>
<tr>
  <th>Back ID</th><th>ISIN</th><th>Side</th><th>Trade Date</th><th>Value Date</th>
  <th>Counterparty</th><th>Qty</th><th>Amount</th><th>GL Account</th><th>Note</th>
</tr>
"""
        for td in unmatched:
            html += '<tr style="background:#FCE4D6">'
            html += f"<td>{td.get('id') or ''}</td>"
            html += f"<td>{td.get('symbol') or ''}</td>"
            html += f"<td>{td.get('direction') or ''}</td>"
            html += f"<td>{_fmt_date(td.get('trade_date'))}</td>"
            html += f"<td>{_fmt_date(td.get('settle_date_cash') or td.get('value_date_cash'))}</td>"
            html += f"<td>{td.get('counterparty') or ''}</td>"
            html += f"<td>{_fmt_num(td.get('qty'))}</td>"
            html += f"<td>{_fmt_amount(td.get('net_amount') or td.get('transaction_value'))}</td>"
            html += f"<td>{compute_gl_account(td.get('symbol', ''))}</td>"
            html += f"<td>{td.get('_similar_confo_note') or ''}</td>"
            html += "</tr>\n"
        html += "</table>\n"

    # ── Table C: Unconfirmed (FX) ──────────────────────────────────────────────
    unconfirmed = result.get("unconfirmed_deals", [])
    if unconfirmed:
        html += """<div class="sect">C. Unconfirmed Trades (FX — no auto-confo expected)</div>
<table>
<tr>
  <th>Back ID</th><th>ISIN / CCY</th><th>Type</th><th>Side</th><th>Trade Date</th><th>Value Date</th>
  <th>Counterparty</th><th>Qty / Nominal</th><th>Amount</th>
</tr>
"""
        for td in unconfirmed:
            type_label = "FX"
            qty_val = td.get("nominal") if td.get("nominal") else td.get("qty")
            html += '<tr style="background:#DAEEF3">'
            html += f"<td>{td.get('id') or ''}</td>"
            html += f"<td>{td.get('symbol') or ''}</td>"
            html += f"<td><b>{type_label}</b></td>"
            html += f"<td>{td.get('direction') or ''}</td>"
            html += f"<td>{_fmt_date(td.get('trade_date'))}</td>"
            html += f"<td>{_fmt_date(td.get('settle_date_cash') or td.get('value_date_cash'))}</td>"
            html += f"<td>{td.get('counterparty') or ''}</td>"
            html += f"<td>{_fmt_num(qty_val)}</td>"
            html += f"<td>{_fmt_amount(td.get('net_amount') or td.get('transaction_value'))}</td>"
            html += "</tr>\n"
        html += "</table>\n"

    # ── Table D: FAB SWIFT MT545/MT547 Settlement Confirmations ──────────────
    fab_swift_rows = result.get("fab_swift_rows", [])
    if fab_swift_rows:
        html += f"""<div class="sect">D. FAB SWIFT Settlement Confirmations — MT545/MT547 ({len(fab_swift_rows)})</div>
<table>
<tr>
  <th>ISIN</th><th>Side</th><th>Match</th><th>Internal Deal</th>
  <th>MT</th><th>Sett. Date</th><th>Eff. Sett. (PDF)</th>
  <th>Settled Amt</th><th>Net Amt (Sys)</th><th>Δ Amount</th>
  <th>Note</th>
</tr>
"""
        for sw in fab_swift_rows:
            status = sw.get("match_status") or ""
            if status == "MATCHED":
                color = "#E2EFDA"
            elif status == "AMOUNT_MISMATCH":
                color = "#FCE4D6"
            elif status == "DATE_MISMATCH":
                color = "#FFF2CC"
            else:
                color = "#FFC7CE"
            ccy = sw.get("settled_currency") or ""
            amount_diff = sw.get("amount_diff")
            diff_str = f"{float(amount_diff):+,.2f}" if amount_diff is not None else ""
            diff_color = ' style="color:green"' if (amount_diff is not None and abs(float(amount_diff)) < 5) else (' style="color:red;font-weight:bold"' if amount_diff is not None else "")
            html += f'<tr style="background:{color}">'
            html += f"<td>{sw.get('isin') or ''}</td>"
            html += f"<td>{sw.get('side') or ''}</td>"
            html += f"<td><b>{status}</b></td>"
            html += f"<td>{sw.get('internal_deal_id') or '—'}</td>"
            html += f"<td>{sw.get('mt_type') or ''}</td>"
            html += f"<td>{_fmt_date(sw.get('settlement_date'))}</td>"
            html += f"<td>{_fmt_date(sw.get('effective_settlement_date'))}</td>"
            html += f"<td>{_fmt_amount(sw.get('settled_amount'))} {ccy}</td>"
            html += f"<td>{_fmt_amount(sw.get('internal_amount'))}</td>"
            html += f"<td{diff_color}>{diff_str}</td>"
            html += f"<td>{sw.get('match_note') or ''}</td>"
            html += "</tr>\n"
        html += "</table>\n"

    html += "<div style='color:#888; font-size:11px; margin-top:20px'>Generated by AM Wealth Settlement Agent</div>"
    html += "</body></html>"
    return html


def send_reconciliation_email(token: str, result: dict, date_from, date_to) -> None:
    """Send reconciliation report email with Excel attachment via Graph API."""
    # Count only non-instructed rows for subject line
    _active = [r for r in result.get("detail_rows", []) if not r.get("instructed")]
    matched = sum(1 for r in _active if r.get("status") in ("MATCHED", "MATCHED_AGGREGATED"))
    partial = sum(1 for r in _active if r.get("status") == "PARTIAL")
    no_confo = len(result.get("unmatched_internal", []))
    unconfirmed_count = len(result.get("unconfirmed_deals", []))

    date_label = str(date_to or date_from or "")
    subject = (
        f"Settlement Reconciliation {date_label}"
        f" | ✅ {matched} matched"
        f" | ⚠️ {partial} review"
        f" | ❌ {no_confo} missing"
        f" | 🔔 {unconfirmed_count} unconfirmed"
    )

    html_body = build_reconciliation_html(result, date_from, date_to)

    # Build Excel attachment
    xlsx_bytes = build_reconciliation_excel(result, date_from, date_to)
    xlsx_b64 = base64.b64encode(xlsx_bytes).decode("utf-8")
    attach_name = f"reconciliation_{date_label or 'report'}.xlsx"

    payload = {
        "message": {
            "subject": subject,
            "body": {"contentType": "HTML", "content": html_body},
            "toRecipients": (
                [{"emailAddress": {"address": TEST_EMAIL}}]
                if TEST_MODE
                else [
                    {"emailAddress": {"address": REPORT_TO}},
                    {"emailAddress": {"address": "Back.office@amwealth.ae"}},
                ]
            ),
            "attachments": [
                {
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": attach_name,
                    "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "contentBytes": xlsx_b64,
                }
            ],
        },
        "saveToSentItems": False,
    }

    url = f"{GRAPH_BASE}/users/{GRAPH_MAILBOX}/sendMail"
    r = requests.post(url, json=payload, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    r.raise_for_status()
    logging.info("Reconciliation email sent to %s", REPORT_TO)


def run_settlement_reconciliation(
    conn,
    run_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    deal_date_from: Optional[date] = None,
    deal_date_to: Optional[date] = None,
    value_date_from: Optional[date] = None,
) -> Dict[str, Any]:
    """
    Runs settlement reconciliation.
    date_from / date_to       — confo window by trade_date (last 2 business days)
    deal_date_from / deal_date_to — internal deals window by trade_date (last 5 business days)
    value_date_from           — exclude confos already settled before this date
    Returns a summary dict plus:
      detail_rows          – list of per-confo comparison results
      unmatched_internal   – internal deals with no matching confo in the wider window
    """
    settlement_trades = aggregate_settlement_trades(
        load_settlement_trades_for_reconciliation(
            conn, date_from=date_from, date_to=date_to, value_date_from=value_date_from
        )
    )
    strict_deals = load_strict_deals_to_process(
        conn,
        date_from=deal_date_from if deal_date_from is not None else date_from,
        date_to=deal_date_to if deal_date_to is not None else date_to,
    )
    broad_deals = load_broad_trade_search(conn)
    unconfirmed_deals = load_unconfirmed_deals(
        conn,
        date_from=deal_date_from if deal_date_from is not None else date_from,
        date_to=deal_date_to if deal_date_to is not None else date_to,
    )

    # Build set of (isin, side, trade_date) for ALL confos ever received
    # (no date filter) — used to exclude already-confirmed deals from Table B
    all_confo_trades = aggregate_settlement_trades(
        load_settlement_trades_for_reconciliation(conn)
    )
    all_confo_keys = {
        (clean_text(st.get("isin")), clean_text(st.get("side")), st.get("trade_date"))
        for st in all_confo_trades
    }

    comparison_rows = 0
    matched_count = 0
    matched_aggregated_count = 0
    partial_count = 0
    not_found_count = 0
    similar_found_count = 0

    detail_rows: List[Dict[str, Any]] = []   # per-confo row for readable report
    matched_internal_ids: set = set()         # track which internal deals were matched

    if run_id is not None:
        clear_reconciliation_run_rows(conn, run_id)

    for st in settlement_trades:
        comparison_rows += 1
        strict_candidates = find_strict_candidates(st, strict_deals)

        td, match_status, best_notes = try_exact_single_match(st, strict_candidates)
        if td is not None and match_status == "MATCHED":
            matched_count += 1
            matched_internal_ids.add(td["id"])

            upsert_reconciliation_result(
                conn=conn,
                settlement_trade_id=st["id"],
                internal_order_id=td["id"],
                match_status="MATCHED",
                match_note=None,
                reconciliation_key=build_reconciliation_key(st),
                counterparty=td.get("counterparty"),
                isin=st.get("isin"),
                direction=st.get("side"),
                trade_date=st.get("trade_date"),
                value_date=st.get("value_date"),
                external_qty=st.get("quantity"),
                internal_qty=td.get("qty"),
                external_amount=st.get("consideration"),
                internal_amount=td.get("transaction_value"),
                compare_side=st.get("side"),
                run_id=run_id,
                matched_internal_ids=str(td["id"]),
                matched_internal_count=1,
                mismatch_json=None,
                external_price=st.get("price"),
                internal_price=td.get("price"),
                external_value_date=st.get("value_date"),
                internal_value_date=td.get("settle_date_cash") or td.get("value_date_cash"),
            )
            detail_rows.append(_detail_row(st, "MATCHED", td=td))
            continue

        # Only try aggregate match if exact match found NO candidate at all.
        # If td is not None, we already have a candidate — go to PARTIAL with mismatch notes.
        rows, agg_status, agg_note = (None, None, None) if td is not None else try_aggregate_match(st, strict_candidates)
        if rows is not None and agg_status is not None:
            matched_aggregated_count += 1
            internal_ids = ",".join([str(r["id"]) for r in rows])
            summed_qty = sum([float(r.get("qty") or 0) for r in rows])
            summed_amount = sum([float(r.get("transaction_value") or 0) for r in rows])
            for r in rows:
                matched_internal_ids.add(r["id"])

            upsert_reconciliation_result(
                conn=conn,
                settlement_trade_id=st["id"],
                internal_order_id=None,
                match_status=agg_status,
                match_note=agg_note,
                reconciliation_key=build_reconciliation_key(st),
                counterparty=rows[0].get("counterparty") if rows else None,
                isin=st.get("isin"),
                direction=st.get("side"),
                trade_date=st.get("trade_date"),
                value_date=st.get("value_date"),
                external_qty=st.get("quantity"),
                internal_qty=summed_qty,
                external_amount=st.get("consideration"),
                internal_amount=summed_amount,
                compare_side=st.get("side"),
                run_id=run_id,
                matched_internal_ids=internal_ids,
                matched_internal_count=len(rows),
                mismatch_json=None,
                external_price=st.get("price"),
                internal_price=None,
                external_value_date=st.get("value_date"),
                internal_value_date=rows[0].get("settle_date_cash") or rows[0].get("value_date_cash"),
            )
            detail_rows.append(_detail_row(st, agg_status, agg_rows=rows, agg_note=agg_note))
            continue

        if td is not None:
            partial_count += 1
            matched_internal_ids.add(td["id"])
            mismatch_json = json.dumps({
                "external_qty": str(st.get("quantity")),
                "internal_qty": str(td.get("qty")),
                "external_amount": str(st.get("consideration")),
                "internal_amount": str(td.get("transaction_value")),
                "external_price": str(st.get("price")),
                "internal_price": str(td.get("price")),
                "notes": best_notes or [],
            }, default=str)

            upsert_reconciliation_result(
                conn=conn,
                settlement_trade_id=st["id"],
                internal_order_id=td["id"],
                match_status="PARTIAL",
                match_note=", ".join(best_notes or []) if best_notes else "partial_match",
                reconciliation_key=build_reconciliation_key(st),
                counterparty=td.get("counterparty"),
                isin=st.get("isin"),
                direction=st.get("side"),
                trade_date=st.get("trade_date"),
                value_date=st.get("value_date"),
                external_qty=st.get("quantity"),
                internal_qty=td.get("qty"),
                external_amount=st.get("consideration"),
                internal_amount=td.get("transaction_value"),
                compare_side=st.get("side"),
                run_id=run_id,
                matched_internal_ids=str(td["id"]),
                matched_internal_count=1,
                mismatch_json=mismatch_json,
                external_price=st.get("price"),
                internal_price=td.get("price"),
                external_value_date=st.get("value_date"),
                internal_value_date=td.get("settle_date_cash") or td.get("value_date_cash"),
            )
            detail_rows.append(_detail_row(st, "PARTIAL", td=td, notes=best_notes))
            continue

        similar_rows = find_similar_broad_rows(st, broad_deals)
        if similar_rows:
            # Try to score the broad matches — if we find a candidate with same ISIN+side+date,
            # show as MATCHED or PARTIAL depending on score (covers settled internal deals).
            broad_strict_candidates = find_strict_candidates(st, similar_rows)
            broad_td, broad_status, broad_notes = try_exact_single_match(st, broad_strict_candidates)
            if broad_td is not None:
                matched_internal_ids.add(broad_td["id"])
                if broad_status == "MATCHED":
                    matched_count += 1
                    upsert_reconciliation_result(
                        conn=conn,
                        settlement_trade_id=st["id"],
                        internal_order_id=broad_td["id"],
                        match_status="MATCHED",
                        match_note=None,
                        reconciliation_key=build_reconciliation_key(st),
                        counterparty=broad_td.get("counterparty"),
                        isin=st.get("isin"),
                        direction=st.get("side"),
                        trade_date=st.get("trade_date"),
                        value_date=st.get("value_date"),
                        external_qty=st.get("quantity"),
                        internal_qty=broad_td.get("qty"),
                        external_amount=st.get("consideration"),
                        internal_amount=broad_td.get("transaction_value"),
                        compare_side=st.get("side"),
                        run_id=run_id,
                        matched_internal_ids=str(broad_td["id"]),
                        matched_internal_count=1,
                        mismatch_json=None,
                        external_price=st.get("price"),
                        internal_price=broad_td.get("price"),
                        external_value_date=st.get("value_date"),
                        internal_value_date=broad_td.get("settle_date_cash") or broad_td.get("value_date_cash"),
                    )
                    detail_rows.append(_detail_row(st, "MATCHED", td=broad_td))
                else:
                    # Score < 90 — show as PARTIAL with internal data
                    partial_count += 1
                    mismatch_json = json.dumps({
                        "external_qty": str(st.get("quantity")),
                        "internal_qty": str(broad_td.get("qty")),
                        "external_amount": str(st.get("consideration")),
                        "internal_amount": str(broad_td.get("transaction_value")),
                        "external_price": str(st.get("price")),
                        "internal_price": str(broad_td.get("price")),
                        "notes": (broad_notes or []) + ["internal deal outside strict date/status window"],
                    }, default=str)
                    upsert_reconciliation_result(
                        conn=conn,
                        settlement_trade_id=st["id"],
                        internal_order_id=broad_td["id"],
                        match_status="PARTIAL",
                        match_note=", ".join(broad_notes or []) + "; outside strict window",
                        reconciliation_key=build_reconciliation_key(st),
                        counterparty=broad_td.get("counterparty"),
                        isin=st.get("isin"),
                        direction=st.get("side"),
                        trade_date=st.get("trade_date"),
                        value_date=st.get("value_date"),
                        external_qty=st.get("quantity"),
                        internal_qty=broad_td.get("qty"),
                        external_amount=st.get("consideration"),
                        internal_amount=broad_td.get("transaction_value"),
                        compare_side=st.get("side"),
                        run_id=run_id,
                        matched_internal_ids=str(broad_td["id"]),
                        matched_internal_count=1,
                        mismatch_json=mismatch_json,
                        external_price=st.get("price"),
                        internal_price=broad_td.get("price"),
                        external_value_date=st.get("value_date"),
                        internal_value_date=broad_td.get("settle_date_cash") or broad_td.get("value_date_cash"),
                    )
                    detail_rows.append(_detail_row(st, "PARTIAL", td=broad_td, notes=broad_notes))
                continue

            similar_found_count += 1
            upsert_reconciliation_result(
                conn=conn,
                settlement_trade_id=st["id"],
                internal_order_id=None,
                match_status="SIMILAR_FOUND_OUTSIDE_STRICT_SCOPE",
                match_note=f"Found {len(similar_rows)} similar rows for same ISIN outside strict match",
                reconciliation_key=build_reconciliation_key(st),
                counterparty=None,
                isin=st.get("isin"),
                direction=st.get("side"),
                trade_date=st.get("trade_date"),
                value_date=st.get("value_date"),
                external_qty=st.get("quantity"),
                internal_qty=None,
                external_amount=st.get("consideration"),
                internal_amount=None,
                compare_side=st.get("side"),
                run_id=run_id,
                matched_internal_ids=",".join([str(r["id"]) for r in similar_rows]),
                matched_internal_count=len(similar_rows),
                mismatch_json=json.dumps(similar_rows, default=str),
                external_price=st.get("price"),
                internal_price=None,
                external_value_date=st.get("value_date"),
                internal_value_date=None,
            )
            detail_rows.append(_detail_row(st, "SIMILAR_FOUND_OUTSIDE_STRICT_SCOPE"))
            continue

        not_found_count += 1
        upsert_reconciliation_result(
            conn=conn,
            settlement_trade_id=st["id"],
            internal_order_id=None,
            match_status="NOT_FOUND",
            match_note="No candidate matched by isin + side + trade_date and no similar ISIN rows found",
            reconciliation_key=build_reconciliation_key(st),
            counterparty=None,
            isin=st.get("isin"),
            direction=st.get("side"),
            trade_date=st.get("trade_date"),
            value_date=st.get("value_date"),
            external_qty=st.get("quantity"),
            internal_qty=None,
            external_amount=st.get("consideration"),
            internal_amount=None,
            compare_side=st.get("side"),
            run_id=run_id,
            matched_internal_ids=None,
            matched_internal_count=None,
            mismatch_json=None,
            external_price=st.get("price"),
            internal_price=None,
            external_value_date=st.get("value_date"),
            internal_value_date=None,
        )
        detail_rows.append(_detail_row(st, "NOT_FOUND"))

    # Reverse check: internal deals that have no matching confo in this window.
    # Exclude deals whose confo exists (even if filtered by value_date) — already settled in prior run.
    # Build a set of (isin, trade_date) from all confos — for side-mismatch detection.
    # Build lookup (isin, trade_date) → list of confos for side-mismatch detection
    all_confo_by_isin_date: Dict[tuple, list] = {}
    for st in all_confo_trades:
        key = (clean_text(st.get("isin")), st.get("trade_date"))
        all_confo_by_isin_date.setdefault(key, []).append(st)

    unmatched_internal = []
    for td in strict_deals:
        if td["id"] in matched_internal_ids:
            continue
        isin_key = clean_text(td.get("symbol"))
        side_key = clean_text(td.get("direction"))
        date_key = td.get("trade_date")
        if (isin_key, side_key, date_key) in all_confo_keys:
            continue
        row = dict(td)
        # Check if a confo exists for same ISIN+date but different side or mismatched fields
        similar_confos = all_confo_by_isin_date.get((isin_key, date_key), [])
        if similar_confos:
            st = similar_confos[0]
            issues = []
            # Side mismatch
            if clean_text(st.get("side")) != side_key:
                issues.append(f"direction: confo={st.get('side')} vs BO={td.get('direction')}")
            # Qty/nominal mismatch
            ext_qty = st.get("quantity") or st.get("nominal")
            int_qty = td.get("nominal") if td.get("nominal") else td.get("qty")
            if ext_qty is not None and int_qty is not None and not values_equal_decimal(ext_qty, int_qty):
                issues.append(f"qty: confo={ext_qty} vs BO={int_qty}")
            # Amount mismatch
            ext_amt = st.get("net_amount") or st.get("consideration")
            int_amt = td.get("net_amount") or td.get("transaction_value")
            if ext_amt is not None and int_amt is not None and not values_equal_decimal(ext_amt, int_amt, Decimal("1")):
                issues.append(f"amount: confo={ext_amt} vs BO={int_amt}")
            # Price mismatch
            ext_price = st.get("price_in_percentage") or st.get("price")
            int_price = td.get("price_in_percentage") or td.get("price")
            if ext_price is not None and int_price is not None and not values_equal_decimal(ext_price, int_price, Decimal("0.5")):
                issues.append(f"price: confo={ext_price} vs BO={int_price}")
            if issues:
                row["_similar_confo_note"] = "⚠️ Similar confo found — " + "; ".join(issues)
            else:
                row["_similar_confo_note"] = "⚠️ Similar confo found — check BO entry"
        unmatched_internal.append(row)

    fab_swift_rows = run_fab_swift_reconciliation(conn, run_id=run_id)

    # Mark detail_rows as INSTRUCTED if tab_deals.status=2 for the matched deal
    instructed_deal_ids: set = set()
    if run_id is not None:
        with conn.cursor() as _cur:
            _cur.execute("""
                SELECT DISTINCT td.id
                FROM back_office_auto.settlement_reconciliation r
                JOIN back_office.tab_deals td ON td.id = r.internal_order_id
                WHERE r.run_id = %s AND td.status = 2
                UNION
                SELECT DISTINCT TRIM(v)::int
                FROM back_office_auto.settlement_reconciliation r
                CROSS JOIN LATERAL unnest(string_to_array(r.matched_internal_ids, ',')) AS v
                WHERE r.run_id = %s
                  AND r.matched_internal_ids IS NOT NULL
                  AND TRIM(v) ~ '^\\d+$'
                  AND EXISTS (
                      SELECT 1 FROM back_office.tab_deals td2
                      WHERE td2.id = TRIM(v)::int AND td2.status = 2
                  )
            """, (run_id, run_id))
            instructed_deal_ids = {row[0] for row in _cur.fetchall() if row[0]}

    for row in detail_rows:
        ids_str = row.get("int_ids") or ""
        row_ids = {int(x.strip()) for x in ids_str.split(",") if x.strip().isdigit()}
        if row_ids and row_ids.issubset(instructed_deal_ids):
            row["instructed"] = True

    return {
        "comparison_rows": comparison_rows,
        "matched_count": matched_count,
        "matched_aggregated_count": matched_aggregated_count,
        "partial_count": partial_count,
        "not_found_count": not_found_count,
        "similar_found_count": similar_found_count,
        "detail_rows": detail_rows,
        "unmatched_internal": unmatched_internal,
        "unconfirmed_deals": unconfirmed_deals,
        "fab_swift_rows": fab_swift_rows,
    }


def settlement_reconciliation_timer(mytimer=None) -> None:
    logging.warning("### settlement_reconciliation_timer started ###")

    run_id = None
    conn = None

    try:
        conn = get_conn()
        run_id = start_agent_run(conn, "settlement_reconciliation_timer")

        t0_date, _t1_date, _t_next_date = get_t0_t1_dates()
        confo_from = n_prev_business_days(t0_date, 10)  # confo: last 10 business days (~2 weeks)
        result = run_settlement_reconciliation(
            conn,
            run_id=run_id,
            date_from=confo_from,     # confo window: last 10 business days
            date_to=t0_date,
            deal_date_from=None,      # internal deals: no date restriction (all active)
            deal_date_to=None,
            # no value_date_from: show all confos in window including already-settled
        )

        finish_agent_run(
            conn,
            run_id,
            "SUCCESS",
            (
                f"comparison_rows={result['comparison_rows']}, "
                f"matched_count={result['matched_count']}, "
                f"matched_aggregated_count={result['matched_aggregated_count']}, "
                f"partial_count={result['partial_count']}, "
                f"not_found_count={result['not_found_count']}, "
                f"similar_found_count={result['similar_found_count']}"
            ),
        )

        has_data = (
            result.get("comparison_rows", 0) > 0
            or len(result.get("unmatched_internal", [])) > 0
        )
        if has_data:
            try:
                token = get_graph_token()
                send_reconciliation_email(token, result, confo_from, t0_date)
            except Exception as email_err:
                logging.exception("Failed to send reconciliation email: %s", email_err)
        else:
            logging.info("settlement_reconciliation_timer: no trades in window, skipping email")

        logging.warning(
            "### settlement_reconciliation_timer finished | result=%s ###",
            result
        )

    except Exception as e:
        logging.exception("Settlement reconciliation failed: %s", e)

        try:
            if conn:
                conn.rollback()
        except Exception:
            logging.exception("Rollback failed")

        try:
            if conn and run_id:
                finish_agent_run(conn, run_id, "FAILED", str(e))
        except Exception:
            logging.exception("Failed to mark reconciliation run as FAILED")

        raise

    finally:
        if conn:
            conn.close()


# =============================================================================
# HTTP TRIGGER: reparse_recent
# POST /api/reparse-recent
# Body (JSON, all optional):
#   { "n": 20, "dry_run": false, "senders": ["bo.tdsm@zarattinibank.ch"] }
# =============================================================================
@app.function_name(name="reparse_recent")
@app.route(route="reparse-recent", auth_level=func.AuthLevel.FUNCTION, methods=["POST"])
def reparse_recent_http(req: func.HttpRequest) -> func.HttpResponse:
    conn = None
    run_id = None
    try:
        body = {}
        try:
            body = req.get_json()
        except Exception:
            body = {}

        n = int(body.get("n") or req.params.get("n") or 3)
        dry_run_param = body.get("dry_run") if "dry_run" in body else req.params.get("dry_run")
        if isinstance(dry_run_param, bool):
            dry_run = dry_run_param
        else:
            dry_run = str(dry_run_param or "false").strip().lower() in {"true", "1", "yes", "y"}

        filter_senders_param = body.get("senders") or req.params.get("senders")
        if filter_senders_param:
            if isinstance(filter_senders_param, list):
                filter_senders = [s.strip().lower() for s in filter_senders_param if s.strip()]
            else:
                filter_senders = [s.strip().lower() for s in str(filter_senders_param).split(",") if s.strip()]
        else:
            filter_senders = None

        conn = get_conn()
        token = get_graph_token()
        mapping_by_sender = load_mapping(conn)

        # Build candidate email list
        with conn.cursor() as cur:
            cur.execute(
                """
                select
                    internet_message_id, message_id, sender,
                    subject, received_at, status, attachment_count
                from (
                    select *,
                        row_number() over (
                            partition by sender
                            order by received_at desc nulls last
                        ) as rn
                    from back_office_auto.settlement_emails
                    where status <> 'SKIPPED'
                ) ranked
                where rn <= %s
                order by sender, received_at desc
                """,
                (n,),
            )
            cols = [d[0] for d in cur.description]
            emails = [dict(zip(cols, row)) for row in cur.fetchall()]

        if filter_senders:
            emails = [e for e in emails if e["sender"].lower() in filter_senders]

        # Filter: skip no-attachment emails for PDF/Excel parsers
        filtered = []
        for e in emails:
            template = detect_template_from_mapping(e["sender"], mapping_by_sender)
            if template not in EMAIL_BODY_TEMPLATES and (e.get("attachment_count") or 0) == 0:
                continue
            filtered.append(e)

        if not filtered:
            return func.HttpResponse(
                json.dumps({"dry_run": dry_run, "n": n, "emails_found": 0, "message": "Nothing to reparse"}),
                mimetype="application/json", status_code=200,
            )

        results = {"ok": 0, "failed": 0, "skipped": 0, "trades": []}

        if not dry_run:
            run_id = start_agent_run(conn, "reparse_recent_http")

        for e in filtered:
            imid = e["internet_message_id"]
            message_id = e["message_id"]
            sender = e["sender"]

            if dry_run:
                results["ok"] += 1
                results["trades"].append({
                    "sender": sender,
                    "internet_message_id": imid,
                    "subject": e.get("subject"),
                    "dry_run": True,
                })
                continue

            # 1. Delete from DB
            try:
                with conn.cursor() as cur:
                    cur.execute(
                        "delete from back_office_auto.settlement_reconciliation"
                        " where settlement_trade_id in ("
                        "  select id from back_office_auto.settlement_trades"
                        "  where internet_message_id = %s)", (imid,))
                    cur.execute(
                        "delete from back_office_auto.settlement_trades where internet_message_id = %s", (imid,))
                    cur.execute(
                        "delete from back_office_auto.settlement_files where internet_message_id = %s", (imid,))
                    cur.execute(
                        "delete from back_office_auto.settlement_emails where internet_message_id = %s", (imid,))
                conn.commit()
            except Exception as ex:
                logging.error("reparse_recent: delete failed for %s: %s", imid, ex)
                conn.rollback()
                results["failed"] += 1
                continue

            # 2. Fetch message stub from Graph
            try:
                url = (
                    f"{GRAPH_BASE}/users/{GRAPH_MAILBOX}/messages/{message_id}"
                    f"?$select=id,internetMessageId,subject,receivedDateTime,from,hasAttachments"
                )
                r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
                r.raise_for_status()
                msg_stub = r.json()
            except Exception as ex:
                logging.error("reparse_recent: graph fetch failed for %s: %s", imid, ex)
                results["failed"] += 1
                continue

            # 3. Reparse
            try:
                status, count = process_message(
                    conn=conn, token=token, mailbox=GRAPH_MAILBOX,
                    msg=msg_stub, mapping_by_sender=mapping_by_sender,
                    processing_run_id=run_id,
                )
                results["trades"].append({"sender": sender, "internet_message_id": imid, "status": status, "trades": count})
                if status in ("PARSED", "NO_TRADES_FOUND"):
                    results["ok"] += 1
                else:
                    results["skipped"] += 1
            except Exception as ex:
                logging.error("reparse_recent: parse failed for %s: %s", imid, ex)
                conn.rollback()
                results["failed"] += 1

        if not dry_run and run_id:
            finish_agent_run(
                conn, run_id,
                "SUCCESS" if results["failed"] == 0 else "PARTIAL",
                f"ok={results['ok']} skipped={results['skipped']} failed={results['failed']}",
            )

        return func.HttpResponse(
            json.dumps({
                "dry_run": dry_run,
                "n": n,
                "emails_found": len(filtered),
                "ok": results["ok"],
                "skipped": results["skipped"],
                "failed": results["failed"],
                "trades": results["trades"],
            }, default=str),
            mimetype="application/json", status_code=200,
        )

    except Exception as ex:
        logging.exception("reparse_recent_http failed: %s", ex)
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
            if run_id:
                try:
                    finish_agent_run(conn, run_id, "FAILED", str(ex))
                except Exception:
                    pass
        return func.HttpResponse(
            json.dumps({"error": str(ex)}),
            mimetype="application/json", status_code=500,
        )
    finally:
        if conn:
            conn.close()


# =============================================================================
# HTTP TRIGGER: run_reconciliation
# POST /api/run-reconciliation
# =============================================================================
@app.function_name(name="run_reconciliation")
@app.route(route="run-reconciliation", auth_level=func.AuthLevel.FUNCTION, methods=["POST"])
def run_reconciliation_http(req: func.HttpRequest) -> func.HttpResponse:
    conn = None
    run_id = None
    try:
        conn = get_conn()
        run_id = start_agent_run(conn, "run_reconciliation_http")
        t0_date, _t1_date, _t_next_date = get_t0_t1_dates()
        confo_from = n_prev_business_days(t0_date, 10)
        result = run_settlement_reconciliation(
            conn, run_id=run_id,
            date_from=confo_from, date_to=t0_date,
            deal_date_from=None, deal_date_to=None,
        )
        finish_agent_run(conn, run_id, "SUCCESS",
            f"comparison_rows={result['comparison_rows']}, matched={result['matched_count']}")
        has_data = (result.get("comparison_rows", 0) > 0
                    or len(result.get("unmatched_internal", [])) > 0)
        if has_data:
            token = get_graph_token()
            send_reconciliation_email(token, result, confo_from, t0_date)
        return func.HttpResponse(
            json.dumps({
                "ok": True,
                "comparison_rows": result["comparison_rows"],
                "matched": result["matched_count"],
                "matched_aggregated_count": result.get("matched_aggregated_count", 0),
                "partial_count": result.get("partial_count", 0),
                "run_id": run_id,
                # Full data for web display (mirrors what is sent in the email)
                "detail_rows": result.get("detail_rows", []),
                "unmatched_internal": result.get("unmatched_internal", []),
                "unconfirmed_deals": result.get("unconfirmed_deals", []),
                "fab_swift_rows": result.get("fab_swift_rows", []),
            }, default=str),
            mimetype="application/json", status_code=200,
        )
    except Exception as ex:
        logging.exception("run_reconciliation_http failed: %s", ex)
        if conn:
            try: conn.rollback()
            except Exception: pass
        if conn and run_id:
            try: finish_agent_run(conn, run_id, "FAILED", str(ex))
            except Exception: pass
        return func.HttpResponse(
            json.dumps({"error": str(ex)}),
            mimetype="application/json", status_code=500,
        )
    finally:
        if conn:
            conn.close()


# =============================================================================
# HTTP TRIGGER: run_email_parser
# POST /api/run-email-parser
# =============================================================================
@app.function_name(name="run_email_parser")
@app.route(route="run-email-parser", auth_level=func.AuthLevel.FUNCTION, methods=["POST"])
def run_email_parser_http(req: func.HttpRequest) -> func.HttpResponse:
    conn = None
    run_id = None
    try:
        token = get_graph_token()
        conn = get_conn()
        run_id = start_agent_run(conn, "run_email_parser_http")
        mapping_by_sender = load_mapping(conn)
        allowed_senders = get_allowed_senders(mapping_by_sender)
        since_hours_param = req.params.get("since_hours")
        lookback = int(since_hours_param) if since_hours_param and since_hours_param.isdigit() else LOOKBACK_HOURS
        since_dt = now_utc() - timedelta(hours=lookback)
        total = parsed_messages = parsed_trades = skipped = 0
        for mailbox in GRAPH_MAILBOXES:
            messages = list_recent_messages(token, mailbox, since_dt)
            logging.warning("EMAIL_PARSER: fetched %d messages from %s since %s", len(messages), mailbox, since_dt.isoformat())
            for msg in messages:
                total += 1
                sender = normalize_email_address(msg.get("from", {}))
                if not is_sender_allowed(sender, allowed_senders):
                    skipped += 1
                    continue
                status, count = process_message(
                    conn=conn,
                    token=token,
                    mailbox=mailbox,
                    msg=msg,
                    mapping_by_sender=mapping_by_sender,
                    processing_run_id=run_id,
                )
                if status in {"PARSED", "NO_TRADES_FOUND"}:
                    parsed_messages += 1
                elif status in {"SKIPPED", "ALREADY_PROCESSED"}:
                    skipped += 1
                parsed_trades += count
        finish_agent_run(conn, run_id, "SUCCESS",
            f"total={total}, parsed={parsed_messages}, trades={parsed_trades}, skipped={skipped}")
        return func.HttpResponse(
            json.dumps({"ok": True, "total": total, "parsed_messages": parsed_messages,
                        "parsed_trades": parsed_trades, "skipped": skipped, "run_id": run_id}),
            mimetype="application/json", status_code=200,
        )
    except Exception as ex:
        logging.exception("run_email_parser_http failed: %s", ex)
        if conn:
            try: conn.rollback()
            except Exception: pass
        if conn and run_id:
            try: finish_agent_run(conn, run_id, "FAILED", str(ex))
            except Exception: pass
        return func.HttpResponse(
            json.dumps({"error": str(ex)}),
            mimetype="application/json", status_code=500,
        )
    finally:
        if conn:
            conn.close()

# =============================================================================
# TIMER TRIGGER: daily_email_parser
# Runs every weekday at 08:30 UAE time (04:30 UTC)
# NCRONTAB: 0 30 4 * * 1-5
# =============================================================================
@app.function_name(name="daily_email_parser")
@app.timer_trigger(schedule="0 30 4 * * 1-5", arg_name="timer", run_on_startup=False, use_monitor=False)
def daily_email_parser(timer: func.TimerRequest) -> None:
    conn = None
    run_id = None
    try:
        conn = get_conn()
        run_id = start_agent_run(conn, "daily_email_parser")
        token = get_graph_token()
        mapping_by_sender = load_mapping(conn)
        allowed_senders = get_allowed_senders(mapping_by_sender)
        lookback = LOOKBACK_HOURS
        since_dt = now_utc() - timedelta(hours=lookback)
        total = parsed_messages = parsed_trades = skipped = 0
        for mailbox in GRAPH_MAILBOXES:
            messages = list_recent_messages(token, mailbox, since_dt)
            logging.warning("EMAIL_PARSER: fetched %d messages from %s since %s", len(messages), mailbox, since_dt.isoformat())
            for msg in messages:
                total += 1
                sender = normalize_email_address(msg.get("from", {}))
                if not is_sender_allowed(sender, allowed_senders):
                    skipped += 1
                    continue
                status, count = process_message(
                    conn=conn,
                    token=token,
                    mailbox=mailbox,
                    msg=msg,
                    mapping_by_sender=mapping_by_sender,
                    processing_run_id=run_id,
                )
                if status in {"PARSED", "NO_TRADES_FOUND"}:
                    parsed_messages += 1
                elif status in {"SKIPPED", "ALREADY_PROCESSED"}:
                    skipped += 1
                parsed_trades += count
        finish_agent_run(conn, run_id, "SUCCESS",
            f"total={total}, parsed={parsed_messages}, trades={parsed_trades}, skipped={skipped}")
    except Exception as ex:
        logging.exception("daily_email_parser failed: %s", ex)
        if conn:
            try: conn.rollback()
            except Exception: pass
        if conn and run_id:
            try: finish_agent_run(conn, run_id, "FAILED", str(ex))
            except Exception: pass
    finally:
        if conn:
            conn.close()


# =============================================================================
# TIMER TRIGGER: daily_reconciliation
# Runs every weekday at 09:00 UAE time (05:00 UTC)
# NCRONTAB: 0 0 5 * * 1-5
# =============================================================================
@app.function_name(name="daily_reconciliation")
@app.timer_trigger(schedule="0 0 5 * * 1-5", arg_name="timer", run_on_startup=False, use_monitor=False)
def daily_reconciliation(timer: func.TimerRequest) -> None:
    conn = None
    run_id = None
    try:
        conn = get_conn()
        run_id = start_agent_run(conn, "daily_reconciliation")
        t0_date, _t1_date, _t_next_date = get_t0_t1_dates()
        confo_from = n_prev_business_days(t0_date, 10)
        result = run_settlement_reconciliation(
            conn, run_id=run_id,
            date_from=confo_from, date_to=t0_date,
            deal_date_from=None, deal_date_to=None,
        )
        finish_agent_run(conn, run_id, "SUCCESS",
            f"comparison_rows={result['comparison_rows']}, matched={result['matched_count']}")
        has_data = (result.get("comparison_rows", 0) > 0
                    or len(result.get("unmatched_internal", [])) > 0)
        if has_data:
            token = get_graph_token()
            send_reconciliation_email(token, result, confo_from, t0_date)
    except Exception as ex:
        logging.exception("daily_reconciliation failed: %s", ex)
        if conn:
            try: conn.rollback()
            except Exception: pass
        if conn and run_id:
            try: finish_agent_run(conn, run_id, "FAILED", str(ex))
            except Exception: pass
    finally:
        if conn:
            conn.close()


# =============================================================================
# FAB MT566 CORPORATE ACTIONS PARSER
# =============================================================================

def parse_mt566_pdf(text: str, filename: str) -> Optional[Dict[str, Any]]:
    """Parse FAB MT566 corporate action PDF.
    FAB SWIFT PDFs render fields with text labels, e.g.:
      :19B::NETT  Net Cash Amount       USD  144,65
      :19B::GRSS  Gross Amount           USD  206,65
      :19B::TAXR  Withholding Tax Amount USD  62,
      :93B::ELIG  Total Eligible ...     Unit Number  1850,
      :98A::RDTE  Record Date/Time       2026-04-01
      :69A::INPE  Interest Period        2025-10-03/2026-04-03
      :98A::VALU  Value Date/Time        2026-04-07
    """
    text_upper = text.upper()

    # Log first 2000 chars for debugging
    logging.warning("MT566_DEBUG file=%s | text_start=%r", filename, text[:2000])

    # Detect action type via :22F::CAEV field
    action_type = None
    caev = rx(r":22F::CAEV[^\n]*(Interest Payment|Cash Dividend|Full Redemption|"
              r"Partial Redemption|Call Redemption)", text)
    if caev:
        caev_upper = caev.upper()
        if "INTEREST" in caev_upper:
            action_type = "COUPON"
        elif "DIVIDEND" in caev_upper:
            action_type = "DIVIDEND"
        elif "FULL" in caev_upper or "CALL" in caev_upper:
            action_type = "FULL_REDEMPTION"
        elif "PARTIAL" in caev_upper:
            action_type = "PARTIAL_REDEMPTION"

    # Fallback: keyword search
    if not action_type:
        if re.search(r"FULL\s+REDEMPTION|CALL\s+REDEMPTION", text_upper):
            action_type = "FULL_REDEMPTION"
        elif re.search(r"PARTIAL\s+REDEMPTION", text_upper):
            action_type = "PARTIAL_REDEMPTION"
        elif re.search(r"INTEREST\s+PAYMENT", text_upper):
            action_type = "COUPON"
        elif re.search(r"CASH\s+DIVIDEND", text_upper):
            action_type = "DIVIDEND"

    if not action_type:
        logging.warning("MT566_PARSER: cannot detect action_type in %s", filename)
        return None

    # SEME — dedup key: :20C::SEME//<ref> or :20C::SEME //<ref>
    seme = rx(r":20C::SEME\s*//(\S+)", text)
    # Fallback: :20C::PREV (Previous Message Reference)
    if not seme:
        seme = rx(r":20C::PREV\s*[^\n]*?(\d{10,})", text)

    # ── ISIN ──
    _ISIN_PREFIXES = (
        "US", "XS", "DE", "GB", "IE", "FR", "NL", "CH", "IT", "ES",
        "AU", "CA", "JP", "HK", "SG", "SE", "NO", "DK", "FI", "AT",
        "BE", "LU", "PT", "GR", "CZ", "PL", "HU", "TR", "ZA", "IN",
        "CN", "KR", "TW", "MX", "BR", "AR", "CL", "CO", "AE", "SA",
        "QA", "KW", "BH", "OM", "JO", "EG", "NG", "KY", "VG", "BM",
    )
    _isin_alt = "|".join(_ISIN_PREFIXES)
    isin = None
    # Pattern 1: ISIN on line before :35B:
    _m = re.search(r"\b((?:" + _isin_alt + r")[A-Z0-9]{10})\b\s*\n\s*:35B:", text)
    if _m:
        isin = _m.group(1)
    # Pattern 2: ISIN within 3 lines before :35B:
    if not isin:
        _m2 = re.search(r"\b((?:" + _isin_alt + r")[A-Z0-9]{10})\b(?:[^\n]*\n){0,3}[^\n]*:35B:", text)
        if _m2:
            isin = _m2.group(1)
    # Pattern 3: after :97A::SAFE (Safekeeping Account section)
    if not isin:
        _m3 = re.search(r":97A::SAFE[^\n]*\n(?:[^\n]*\n){0,3}[^\n]*\b((?:" + _isin_alt + r")[A-Z0-9]{10})\b", text)
        if _m3:
            isin = _m3.group(1)
    # Pattern 4: any ISIN in text
    if not isin:
        _m5 = re.search(r"\b((?:" + _isin_alt + r")[A-Z0-9]{10})\b", text)
        if _m5:
            isin = _m5.group(1)

    # ── Helper: parse :19B:: amounts ──
    def _parse_19b(tag: str) -> Tuple[Optional[str], Optional[Decimal]]:
        # :19B::TAG//CCY12345,67
        m = re.search(rf":19B::{tag}//([A-Z]{{3}})([\d,\.]+)", text, re.IGNORECASE)
        if not m:
            # FAB PDF: :19B::TAG ... CCY 12345,67
            m = re.search(rf":19B::{tag}\b[^\n]*?([A-Z]{{3}})\s+([\d,\.]+)", text, re.IGNORECASE)
        if m:
            ccy = m.group(1).upper()
            raw = m.group(2).rstrip(",")
            if re.search(r",\d{1,2}$", raw):
                raw = raw.replace(",", ".")
            else:
                raw = raw.replace(",", "")
            return ccy, parse_decimal(raw)
        return None, None

    # ── Amount: NETT first (what Sharly enters), fallback to GRSS ──
    currency, cash_amount = _parse_19b("NETT")
    if cash_amount is None:
        currency, cash_amount = _parse_19b("GRSS")

    # ── Tax: :19B::TAXR (Withholding Tax Amount), 0 if absent ──
    _, tax_amount = _parse_19b("TAXR")
    if tax_amount is None:
        _, tax_amount = _parse_19b("WITL")
    if tax_amount is None:
        tax_amount = Decimal("0")

    # ── Charges: :19B::CHAR, 0 if absent ──
    _, charges_amount = _parse_19b("CHAR")
    if charges_amount is None:
        charges_amount = Decimal("0")

    # ── Nominal: :93B::ELIG (Total Eligible for Corporate Action Balance) ──
    # FAB PDF: ":93B::ELIG ... Unit Number 1850,"  or  "Face Amount 200000,"
    nominal_raw = rx(r":93B::ELIG[^\n]*?(?:Unit Number|Face Amount|Quantity)\s+([\d,\.]+)", text)
    if not nominal_raw:
        nominal_raw = rx(r":93B::ELIG\b[^\n]*([\d][\d,\.]*)", text)
    if not nominal_raw:
        nominal_raw = rx(r":93B::CONB[^\n]*?(?:Unit Number|Face Amount|Quantity)\s+([\d,\.]+)", text)
    nominal = None
    if nominal_raw:
        nominal_raw = nominal_raw.rstrip(",")
        if re.search(r",\d{1,2}$", nominal_raw):
            nominal_raw = nominal_raw.replace(",", ".")
        else:
            nominal_raw = nominal_raw.replace(",", "")
        nominal = parse_decimal(nominal_raw)

    # ── Helper: parse :98A:: dates ──
    def _parse_98a(tag: str) -> Optional[date]:
        raw = rx(rf":98A::{tag}\s*//(\d{{8}})", text)
        if raw:
            raw = f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
            return parse_date_any(raw, prefer_day_first=False)
        raw = rx(rf":98A::{tag}\b[^\n]*?(\d{{4}}-\d{{2}}-\d{{2}})", text)
        if raw:
            return parse_date_any(raw, prefer_day_first=False)
        return None

    # ── Value date: :98A::VALU or :98A::PAYD ──
    payment_date = _parse_98a("VALU")
    if not payment_date:
        payment_date = _parse_98a("PAYD")

    # ── Trade date ──
    trade_date = None
    if action_type == "COUPON":
        # Interest Period :69A::INPE — take SECOND (latest) date
        # FAB PDF: ":69A::INPE Interest Period 2025-10-03/2026-04-03"
        inpe_raw = rx(r":69A::INPE\b[^\n]*?(\d{4}-\d{2}-\d{2})/(\d{4}-\d{2}-\d{2})", text)
        if not inpe_raw:
            # Try full line with both dates
            m_inpe = re.search(r":69A::INPE\b[^\n]*?(\d{4}-\d{2}-\d{2})[^\d]+(\d{4}-\d{2}-\d{2})", text)
            if m_inpe:
                trade_date = parse_date_any(m_inpe.group(2), prefer_day_first=False)
        else:
            # rx returns first group only; re-search for both dates
            m_inpe2 = re.search(r":69A::INPE\b[^\n]*?(\d{4}-\d{2}-\d{2})/(\d{4}-\d{2}-\d{2})", text)
            if m_inpe2:
                trade_date = parse_date_any(m_inpe2.group(2), prefer_day_first=False)
    if not trade_date and action_type == "DIVIDEND":
        # Record Date :98A::RDTE
        trade_date = _parse_98a("RDTE")
    if not trade_date:
        # Fallback: RDTE for any type
        trade_date = _parse_98a("RDTE")

    # ── Cash account IBAN ──
    cash_account_iban = None
    m_iban = re.search(r":97[A-Z]::CASH[^\n]*?(?:IBAN[/\s]*)?\b([A-Z]{2}\d{2}[0-9A-Z]{8,30})\b", text)
    if m_iban:
        cash_account_iban = m_iban.group(1)
    if not cash_account_iban:
        m_iban2 = re.search(r":97[A-Z]::CASH[^\n]*\n\s*(?:IBAN[/\s]*)?\b([A-Z]{2}\d{2}[0-9A-Z]{8,30})\b", text)
        if m_iban2:
            cash_account_iban = m_iban2.group(1)
    if not cash_account_iban:
        m_iban3 = re.search(r"\b(AE\d{21})\b", text)
        if m_iban3:
            cash_account_iban = m_iban3.group(1)

    account_number_key = cash_account_iban[-16:] if cash_account_iban and len(cash_account_iban) >= 16 else None

    logging.warning(
        "MT566_PARSED file=%s | type=%s isin=%s amount=%s ccy=%s tax=%s charges=%s nominal=%s "
        "trade_date=%s value_date=%s iban=%s seme=%s",
        filename, action_type, isin, cash_amount, currency, tax_amount, charges_amount,
        nominal, trade_date, payment_date, cash_account_iban, seme,
    )

    # Auto-generate comment
    amount_str = f"{cash_amount:,.2f} {currency}" if cash_amount and currency else "N/A"
    comment_parts = {
        "DIVIDEND": f"Cash Dividend | ISIN {isin or 'N/A'} | {amount_str} | Pay Date {payment_date or 'N/A'}",
        "COUPON": f"Interest Payment | ISIN {isin or 'N/A'} | {amount_str} | Pay Date {payment_date or 'N/A'}",
        "PARTIAL_REDEMPTION": f"Partial Redemption | ISIN {isin or 'N/A'} | Nominal {nominal or 'N/A'} | {amount_str} | Date {payment_date or 'N/A'}",
        "FULL_REDEMPTION": f"Full Redemption | ISIN {isin or 'N/A'} | Nominal {nominal or 'N/A'} | {amount_str} | Date {payment_date or 'N/A'}",
    }
    comment = comment_parts.get(action_type, "")

    return {
        "pdf_filename": filename,
        "seme": seme,
        "action_type": action_type,
        "isin": isin,
        "cash_amount": cash_amount,
        "currency": currency,
        "payment_date": payment_date,
        "trade_date": trade_date,
        "tax_amount": tax_amount,
        "charges_amount": charges_amount,
        "nominal": nominal,
        "cash_account_iban": cash_account_iban,
        "account_number_key": account_number_key,
        "comment": comment,
    }


def _lookup_gl_account(conn, account_number_key: Optional[str]) -> Optional[str]:
    """Look up gl_account_name from back_office.tab_gl_account by last 16 chars of account_number."""
    if not account_number_key:
        return None
    with conn.cursor() as cur:
        cur.execute(
            "SELECT account_name FROM back_office.tab_gl_account "
            "WHERE RIGHT(account_number::text, 16) = %s LIMIT 1",
            (account_number_key,)
        )
        row = cur.fetchone()
        return row[0] if row else None


def _insert_mt566_parsed(conn, data: Dict[str, Any]) -> Optional[int]:
    """Insert or update MT566 parsed record. UPSERT by seme (dedup key)."""
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO back_office_auto.tab_mt566_parsed
                (received_at, pdf_filename, seme, action_type, isin,
                 cash_amount, currency, payment_date, trade_date,
                 tax_amount, charges_amount, nominal,
                 cash_account_iban, account_number_key, gl_account_name, comment, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (seme) DO UPDATE SET
                pdf_filename = EXCLUDED.pdf_filename,
                action_type = EXCLUDED.action_type,
                isin = EXCLUDED.isin,
                cash_amount = EXCLUDED.cash_amount,
                currency = EXCLUDED.currency,
                payment_date = EXCLUDED.payment_date,
                trade_date = EXCLUDED.trade_date,
                tax_amount = EXCLUDED.tax_amount,
                charges_amount = EXCLUDED.charges_amount,
                nominal = EXCLUDED.nominal,
                cash_account_iban = EXCLUDED.cash_account_iban,
                account_number_key = EXCLUDED.account_number_key,
                gl_account_name = EXCLUDED.gl_account_name,
                comment = EXCLUDED.comment,
                status = EXCLUDED.status
            WHERE back_office_auto.tab_mt566_parsed.status IN ('pending', 'review_required')
            RETURNING id
        """, (
            data["received_at"],
            data.get("pdf_filename"),
            data.get("seme"),
            data["action_type"],
            data.get("isin"),
            data.get("cash_amount"),
            data.get("currency"),
            data.get("payment_date"),
            data.get("trade_date"),
            data.get("tax_amount"),
            data.get("charges_amount"),
            data.get("nominal"),
            data.get("cash_account_iban"),
            data.get("account_number_key"),
            data.get("gl_account_name"),
            data.get("comment"),
            data.get("status", "pending"),
        ))
        row = cur.fetchone()
        conn.commit()
        return row[0] if row else None


def _process_mt566_message(
    conn,
    token: str,
    mailbox: str,
    msg: Dict[str, Any],
    internet_message_id: str,
    subject: str,
    received_at,
    processing_run_id: int,
) -> Tuple[str, int]:
    """Handle FAB MT566 corporate action emails — parse PDFs and store to tab_mt566_parsed."""
    if email_already_processed(conn, internet_message_id):
        return ("ALREADY_PROCESSED", 0)

    message_id = msg["id"]
    attachments = get_message_attachments(token, mailbox, message_id)

    insert_settlement_email(
        conn=conn,
        internet_message_id=internet_message_id,
        message_id=message_id,
        sender="noreply@bankfab.com",
        subject=subject,
        received_at=received_at,
        status="RECEIVED",
        note="FAB MT566 received",
        mailbox=mailbox,
        attachment_count=len(attachments),
        parsed_trade_count=0,
        processing_run_id=processing_run_id,
    )

    parsed_count = 0
    for att in attachments:
        if att.get("@odata.type") != "#microsoft.graph.fileAttachment":
            continue
        filename = att.get("name") or "unnamed"
        if not filename.lower().endswith(".pdf"):
            continue
        content_b64 = att.get("contentBytes")
        if not content_b64:
            continue
        file_bytes = base64.b64decode(content_b64)
        text = extract_pdf_text(file_bytes)
        result = parse_mt566_pdf(text, filename)
        if not result:
            logging.warning("MT566: could not parse %s", filename)
            continue

        result["received_at"] = received_at
        gl_account_name = _lookup_gl_account(conn, result.get("account_number_key"))
        result["gl_account_name"] = gl_account_name
        result["status"] = "pending" if gl_account_name else "review_required"

        if not gl_account_name:
            logging.warning(
                "MT566: GL account not found for IBAN key=%s, file=%s",
                result.get("account_number_key"), filename,
            )

        _insert_mt566_parsed(conn, result)
        parsed_count += 1

    status = "PARSED" if parsed_count > 0 else "NO_TRADES_FOUND"
    insert_settlement_email(
        conn=conn,
        internet_message_id=internet_message_id,
        message_id=message_id,
        sender="noreply@bankfab.com",
        subject=subject,
        received_at=received_at,
        status=status,
        note=f"MT566 parsed: {parsed_count}",
        mailbox=mailbox,
        attachment_count=len(attachments),
        parsed_trade_count=parsed_count,
        processing_run_id=processing_run_id,
    )
    return (status, parsed_count)


# =============================================================================
# CMF (Cash Management Facilities) EMAIL PARSER
# =============================================================================

def _strip_html(html_text: str) -> str:
    """Strip HTML tags and decode entities, return plain text."""
    import re as _re
    from html import unescape
    text = _re.sub(r'<br\s*/?>', '\n', html_text, flags=_re.IGNORECASE)
    text = _re.sub(r'<[^>]+>', ' ', text)
    text = unescape(text)
    text = _re.sub(r'[^\S\n]+', ' ', text)
    return text.strip()


def _parse_number(s: str) -> Optional[float]:
    """Parse number from string like '392,227.00' or '4 846 481.85'."""
    if not s:
        return None
    s = s.replace(',', '').replace(' ', '').replace('\u00a0', '').replace('$', '').strip()
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date_cmf(s: str) -> Optional[str]:
    """Parse date from various CMF formats: MM/DD/YY, MM/DD/YYYY, YYYY-MM-DD."""
    if not s:
        return None
    s = s.strip()
    for fmt in ('%m/%d/%y', '%m/%d/%Y', '%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            d = datetime.strptime(s, fmt)
            return d.strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def parse_cmf_email(body_text: str) -> Optional[Dict[str, Any]]:
    """Parse CMF email body text and return structured data."""
    import re as _re

    text = body_text
    upper = text.upper()

    # Skip "No CMF movements" emails
    if 'NO CMF MOVEMENT' in upper:
        return None

    result = {
        'email_type': None,
        'counterparty': None,
        'isin': None,
        'famt_close': None,
        'famt_reopen': None,
        'amount_close': None,
        'amount_reopen': None,
        'interest': None,
        'net_amount': None,
        'net_nominal': None,
        'rate': None,
        'currency': 'USD',
        'trade_date': None,
        'settlement_date': None,
        'ssi': None,
    }

    # --- Counterparty ---
    m = _re.search(r'(?:Trade\s+)?VS\s*:\s*(\w[\w\s]*?)(?:\s*[-\u2013\u2014]|\s*$)', text, _re.IGNORECASE | _re.MULTILINE)
    if m:
        result['counterparty'] = m.group(1).strip()

    # --- SSI / E/C ---
    m = _re.search(r'(?:E/C|EC|FAB\s+SSI)\s*[:\s]*(\d{4,6})', text, _re.IGNORECASE)
    if m:
        result['ssi'] = 'E/C ' + m.group(1)

    # --- Detect email type ---
    has_closing = bool(_re.search(r'Closing\s+Deal', text, _re.IGNORECASE))
    has_reopen = bool(_re.search(r'Reopen\s+New\s+Deal', text, _re.IGNORECASE))
    has_fully = 'FULLY CLOSED' in upper
    has_new_trade = bool(_re.search(r'New\s+trade\s+opened', text, _re.IGNORECASE))
    has_new_repo = bool(_re.search(r'(?:opened|details\s+of)\s+(?:rev(?:erse)?\s+)?repo', text, _re.IGNORECASE))

    if has_closing and has_reopen:
        result['email_type'] = 'partial_close'
    elif has_closing or has_fully:
        result['email_type'] = 'fully_closed'
    elif has_new_trade or has_new_repo:
        result['email_type'] = 'new_trade'
    else:
        if _re.search(r'ISIN\s*:', text):
            result['email_type'] = 'new_trade'
        else:
            return None

    # --- Parse ISIN ---
    m = _re.search(r'ISIN\s*:\s*([A-Z]{2}[A-Z0-9]{9,10})', text)
    if m:
        result['isin'] = m.group(1)

    # --- Parse Rate ---
    m = _re.search(r'Rate\s*:\s*([\d.]+)', text, _re.IGNORECASE)
    if not m:
        m = _re.search(r'Repo\s+Rate\s*:\s*([\d.]+)', text, _re.IGNORECASE)
    if m:
        result['rate'] = float(m.group(1))

    # --- Parse Currency ---
    m = _re.search(r'Currency\s*:\s*([A-Z]{3})', text, _re.IGNORECASE)
    if m:
        result['currency'] = m.group(1).upper()

    if result['email_type'] == 'new_trade':
        # FAMT
        m = _re.search(r'FAMT\s*:\s*([\d,.\s]+)', text)
        if m:
            result['famt_close'] = _parse_number(m.group(1))
            result['net_nominal'] = result['famt_close']

        # Amount (Wired out)
        m = _re.search(r'Wired\s+out\s*:\s*\$?([\d,.\s]+)', text, _re.IGNORECASE)
        if not m:
            m = _re.search(r'Wired\s+(?:Amt|Amount)\s*[.:]\s*\$?([\d,.\s]+)', text, _re.IGNORECASE)
        if m:
            result['amount_close'] = _parse_number(m.group(1))
            result['net_amount'] = result['amount_close']

        # Cash field (e.g. "Cash: USD 854,352.00")
        if not result['amount_close']:
            m = _re.search(r'Cash\s*:\s*(?:USD\s*)?\$?([\d,.\s]+)', text, _re.IGNORECASE)
            if m:
                result['amount_close'] = _parse_number(m.group(1))
                result['net_amount'] = result['amount_close']

        # Trade Date
        m = _re.search(r'TD\s*:\s*([\d/]+)', text)
        if not m:
            m = _re.search(r'Trade\s+Date\s*:\s*([\d/.-]+)', text, _re.IGNORECASE)
        if m:
            result['trade_date'] = _parse_date_cmf(m.group(1))

        # Settlement Date
        m = _re.search(r'SD\s*:\s*([\d/]+)', text)
        if not m:
            m = _re.search(r'Settlement\s+Date\s*:\s*([\d/.-]+)', text, _re.IGNORECASE)
        if m:
            result['settlement_date'] = _parse_date_cmf(m.group(1))

        # Face Amount from structured field
        if not result['famt_close']:
            m = _re.search(r'Face\s+Amount\s*:\s*([\d,.\s]+)', text, _re.IGNORECASE)
            if m:
                result['famt_close'] = _parse_number(m.group(1))
                result['net_nominal'] = result['famt_close']

    elif result['email_type'] in ('partial_close', 'fully_closed'):
        # --- Netting Instruction block ---
        ni_block = _re.search(
            r'Netting\s+Instruction[:\s]*(.*?)(?:Kind\s+regards|$)',
            text, _re.IGNORECASE | _re.DOTALL
        )
        ni_text = ni_block.group(1) if ni_block else text

        # ISIN from netting block
        m = _re.search(r'ISIN\s*:\s*([A-Z]{2}[A-Z0-9]{9,10})', ni_text)
        if m:
            result['isin'] = m.group(1)

        # FAMT Close
        m = _re.search(r'FAMT\s+Close\s*:\s*([\d,.\s]+)', ni_text, _re.IGNORECASE)
        if m:
            result['famt_close'] = _parse_number(m.group(1))

        # Interest
        m = _re.search(r'Interest\s*:\s*\$?\s*([\d,.\s]+)', ni_text, _re.IGNORECASE)
        if m:
            result['interest'] = _parse_number(m.group(1))

        # Wired In (net amount for close)
        m = _re.search(r'Wired\s+In\s*:\s*\$?\s*([\d,.\s]+)', ni_text, _re.IGNORECASE)
        if m:
            result['net_amount'] = _parse_number(m.group(1))

        # Trade date from netting
        m = _re.search(r'Trade\s+date\s*:\s*([\d/.-]+)', ni_text, _re.IGNORECASE)
        if m:
            result['trade_date'] = _parse_date_cmf(m.group(1))

        # Settlement date (SD)
        m = _re.search(r'SD\s*:\s*([\d/.-]+)', ni_text)
        if not m:
            m = _re.search(r'Settlement\s+Date\s*:\s*([\d/.-]+)', ni_text, _re.IGNORECASE)
        if m:
            result['settlement_date'] = _parse_date_cmf(m.group(1))

        # --- Closing Deal table: get amount + interest ---
        close_block = _re.search(
            r'Closing\s+Deal[:\s]*(.*?)(?:Reopen|Netting|Kind\s+regards|$)',
            text, _re.IGNORECASE | _re.DOTALL
        )
        if close_block:
            cb = close_block.group(1)
            m = _re.search(r'Amount\s*[:\s]*\$?\s*([\d,.\s]+)', cb, _re.IGNORECASE)
            if m:
                result['amount_close'] = _parse_number(m.group(1))
            m = _re.search(r'Net\s+wired\s+amount\s*[:\s]*([\d,.\s]+)', cb, _re.IGNORECASE)
            if m:
                nwa = _parse_number(m.group(1))
                if nwa and not result['net_amount']:
                    result['net_amount'] = nwa
            if not result['interest']:
                m = _re.search(r'Interest\s*[:\s]*([\d,.\s]+)', cb, _re.IGNORECASE)
                if m:
                    result['interest'] = _parse_number(m.group(1))
            if not result['famt_close']:
                m = _re.search(r'FAMT\s*[:\s]*([\d,.\s]+)', cb, _re.IGNORECASE)
                if m:
                    result['famt_close'] = _parse_number(m.group(1))

        # --- Reopen New Deal table ---
        if result['email_type'] == 'partial_close':
            reopen_block = _re.search(
                r'Reopen\s+New\s+Deal[:\s]*(.*?)(?:Netting|Kind\s+regards|$)',
                text, _re.IGNORECASE | _re.DOTALL
            )
            if reopen_block:
                rb = reopen_block.group(1)
                m = _re.search(r'Amount\s*[:\s]*\$?\s*([\d,.\s]+)', rb, _re.IGNORECASE)
                if m:
                    result['amount_reopen'] = _parse_number(m.group(1))
                m = _re.search(r'FAMT\s+left\s*[:\s]*([\d,.\s]+)', rb, _re.IGNORECASE)
                if not m:
                    m = _re.search(r'FAMT\s*[:\s]*([\d,.\s]+)', rb, _re.IGNORECASE)
                if m:
                    result['famt_reopen'] = _parse_number(m.group(1))

        # --- Compute net nominal ---
        if result['famt_close'] and result.get('famt_reopen'):
            result['net_nominal'] = result['famt_close'] - result['famt_reopen']
        elif result['famt_close']:
            result['net_nominal'] = result['famt_close']

        # --- Compute net amount if not from netting block ---
        if not result['net_amount']:
            ac = result.get('amount_close') or 0
            interest = result.get('interest') or 0
            ar = result.get('amount_reopen') or 0
            if ac:
                result['net_amount'] = ac + interest - ar

    return result


def _process_cmf_message(
    conn,
    token: str,
    mailbox: str,
    msg: Dict[str, Any],
    sender: str,
    internet_message_id: str,
    subject: str,
    received_at,
    processing_run_id: int,
) -> Tuple[str, int]:
    """Handle CMF (Cash Management Facilities) emails — parse body and store to tab_cmf_parsed."""
    if email_already_processed(conn, internet_message_id):
        return ("ALREADY_PROCESSED", 0)

    message_id = msg["id"]

    # Get full message with body
    full_msg = get_message_full(token, mailbox, message_id)
    body_html = (full_msg.get("body") or {}).get("content") or ""
    body_text = _strip_html(body_html)

    insert_settlement_email(
        conn=conn,
        internet_message_id=internet_message_id,
        message_id=message_id,
        sender=sender,
        subject=subject,
        received_at=received_at,
        status="RECEIVED",
        note="CMF email received",
        mailbox=mailbox,
        attachment_count=0,
        parsed_trade_count=0,
        processing_run_id=processing_run_id,
    )

    parsed = parse_cmf_email(body_text)
    if not parsed:
        insert_settlement_email(
            conn=conn,
            internet_message_id=internet_message_id,
            message_id=message_id,
            sender=sender,
            subject=subject,
            received_at=received_at,
            status="SKIPPED",
            note="CMF: no data or 'No CMF movements'",
            mailbox=mailbox,
            attachment_count=0,
            parsed_trade_count=0,
            processing_run_id=processing_run_id,
        )
        return ("SKIPPED", 0)

    # Insert into tab_cmf_parsed
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO back_office_auto.tab_cmf_parsed
                (email_id, received_at, email_type, counterparty, isin,
                 famt_close, famt_reopen, amount_close, amount_reopen,
                 interest, net_amount, net_nominal, rate, currency,
                 trade_date, settlement_date, ssi, status)
            VALUES (%s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, 'pending')
            ON CONFLICT (email_id) DO NOTHING
        """, (
            internet_message_id,
            received_at,
            parsed['email_type'],
            parsed['counterparty'],
            parsed['isin'],
            parsed['famt_close'],
            parsed['famt_reopen'],
            parsed['amount_close'],
            parsed['amount_reopen'],
            parsed['interest'],
            parsed['net_amount'],
            parsed['net_nominal'],
            parsed['rate'],
            parsed['currency'],
            parsed['trade_date'],
            parsed['settlement_date'],
            parsed['ssi'],
        ))
    conn.commit()

    insert_settlement_email(
        conn=conn,
        internet_message_id=internet_message_id,
        message_id=message_id,
        sender=sender,
        subject=subject,
        received_at=received_at,
        status="PARSED",
        note=f"CMF {parsed['email_type']}: {parsed.get('isin') or 'no ISIN'}",
        mailbox=mailbox,
        attachment_count=0,
        parsed_trade_count=1,
        processing_run_id=processing_run_id,
    )
    return ("PARSED", 1)
